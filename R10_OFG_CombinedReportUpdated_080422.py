
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 16 07:21:14 2021

@author: jgcobb, christalrhigdon, kmmiles
"""

import sys
import pandas as pd
import numpy as np
import csv
import re
import os
import tempfile
import shutil
import traceback
from collections import OrderedDict
from contextlib import contextmanager
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, PatternFill, Border, Side, Font, Fill
import arcpy
from arcpy import env
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.cell_range import CellRange

from reporting import HuntingDetailReport


class Toolbox(object):
    def __init__(self):
        """Define the toolbox (the name of the toolbox is the name of the
        .pyt file)."""
        self.label = "R10 Outfitter Guide Tools"
        self.alias = ""

        # List of tool classes associated with this toolbox
        self.tools = [GuidedRecReport, BusinessXlsCreator, BusinessXlsUpdater, OutfittingReport, MendenhallReport, IcefieldReport,
                      IcefieldSummary, MendenhallSummary, HeliskiReport, HuntingDetailReport, FiveYearSummary, RVD_Report,
                      NEPAReview_Shoreline2, NEPAReview_KMRD, ConfirmActualUse, WildernessSummary, HeliskiSummary, VisitationSummary]

class GuidedRecDetail:

    def __init__(self, business_name, startdate, enddate, certification, save_path):

        self.row = 4
        wb_path = r'T:\FS\NFS\R10\Program\2700SpecialUsesMgmt\GIS\R10OGDatabase\ToolData\guided_rec_detail_template.xlsx'
        wb = load_workbook(wb_path)

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', business_name.title())
        savedates = startdate.strftime("%Y%m%d") + "_" + enddate.strftime("%Y%m%d")
        self.savefile = "GuidedRecDetail_" + savedates + '_' + savebusiness + ".xlsx"
        self.path = os.path.join(save_path, self.savefile)

        trip_style = NamedStyle('trip_style')
        thin = Side(border_style='thin', color='000000')
        trip_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        trip_style.fill = PatternFill("solid", fgColor="AAAAAA")
        wb.add_named_style(trip_style)

        int_style = NamedStyle('int_style', number_format='0')
        int_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        int_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(int_style)

        coord_style = NamedStyle('coord_style', number_format='###.##0')
        coord_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        coord_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(coord_style)

        float_style = NamedStyle('float_style', number_format='0.00')
        float_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        float_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(float_style)

        str_style = NamedStyle('str_style')
        str_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        str_style.alignment = Alignment(horizontal='left')
        wb.add_named_style(str_style)

        datetime_style = NamedStyle('datetime_style', number_format='MM/DD/YY')
        datetime_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        datetime_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(datetime_style)

        self.styles = ['datetime_style', 'int_style', 'str_style', 'coord_style',
                       'coord_style', 'str_style', 'int_style', 'int_style',
                       'int_style', 'float_style', 'datetime_style']

        ws = wb.active
        ws['C2'] = certification
        header = 'Business/Organization: {}\nOperating Season: {}'.format(business_name, enddate.strftime('%Y'))
        footer = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
        ws.firstHeader.left.text = header
        ws.firstFooter.left.text = footer
        ws.oddHeader.left.text = header
        ws.evenHeader.left.text = header
        ws.oddFooter.left.text = footer
        ws.evenFooter.left.text = footer
        wb.save(self.path)

    @contextmanager
    def __worksheet(self):

        wb = load_workbook(self.path)
        ws = wb.active
        try:
            yield ws
        finally:
            wb.save(self.path)

    def save_trips(self, trips):

        with self.__worksheet() as ws:
            for trip_guid in trips.keys():
                self._add_trip(ws, **trips[trip_guid])

    def _add_trip(self, ws, startdate, enddate, tripclients, category, submitter, activities):

        # Skip to next page if trip records span page.
        current_page_line = (self.row - 27) % 29 if self.row > 27 else self.row
        current_page_line = 29 if not current_page_line else current_page_line
        trip_end_line = current_page_line + len(activities)
        if (self.row <= 27) and (trip_end_line > 27):
            self.row += 28 - current_page_line
        elif trip_end_line > 29:
            self.row += 30 - current_page_line

        for col in range(1, 11):
            cell = ws.cell(row=self.row, column=col)
            cell.style = 'trip_style'

        ws.merge_cells('A{}:C{}'.format(self.row, self.row))
        ws.merge_cells('D{}:E{}'.format(self.row, self.row))
        ws.merge_cells('F{}:G{}'.format(self.row, self.row))
        ws.merge_cells('H{}:J{}'.format(self.row, self.row))
        ws['A{}'.format(self.row)] = "Trip Dates: {}-{}".format(startdate.strftime('%m/%d/%Y'),
                                                                enddate.strftime('%m/%d/%Y'))
        ws['D{}'.format(self.row)] = 'Clients on Trip: {}'.format(tripclients)
        ws['F{}'.format(self.row)] = 'Category: {}'.format(category)
        ws['H{}'.format(self.row)] = 'By: {}, {}'.format(submitter[0], submitter[1].strftime('%m/%d/%Y'))
        self.row += 1
        self._add_activities(ws, activities)

    def _add_activities(self, worksheet, activity_records):

        for record in activity_records:
            for col, value in enumerate(record, 1):
                cell = worksheet.cell(row=self.row, column=col)
                cell.value = value
                cell.style = self.styles[col - 1]
            self.row += 1


class GuidedRecReport(object):
    def __init__(self):
        self.label = "Run Guided Recreation Detail Report"
        self.canRunInBackground = False

        connection = 'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        tables = ['S_R10.R10_OFG_GUIDEDREC_ACTIVITY',
                  'S_R10.R10_OFG_GUIDEDREC_STOP',
                  'S_R10.R10_OFG_GUIDEDREC_DAY',
                  'S_R10.R10_OFG_GUIDEDREC_TRIP']
        self.tables = [connection + table for table in tables]

    def getParameterInfo(self):
        # Define parameter definitions

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"

        sc = arcpy.da.SearchCursor(self.tables[3], "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        param0.filter.list = [row[0] for row in sc]

        param1 = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param1.value = datetime(datetime.today().year, 1, 1).strftime('%x')

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param2.value = datetime.today().strftime('%x')

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]

        return params

    def isLicensed(self):  # optional
        return True

    def updateParameters(self, parameters):  # optional
        return

    def updateMessages(self, parameters):  # optional
        return

    def _get_rows(self, table, where):

        rows = []
        with arcpy.da.SearchCursor(table, '*', where) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                rows.append(row)
        return pd.DataFrame(rows)

    def _read_activities(self, trip_guids):

        where_trip = "TRIP_GUID IN ({})".format(','.join(["'{}'".format(guid) for guid in trip_guids]))

        activities = self._get_rows(self.tables[0], where_trip)
        stops = self._get_rows(self.tables[1], where_trip)
        days = self._get_rows(self.tables[2], where_trip)

        use_report = (
            activities[['ACTIVITY', 'CLIENTNUMBER', 'GROUPNUMBER', 'GUIDENUMBER', 'HOURSSPENTONFS', 'STOP_GUID']]
            .merge(stops[['STOP_GUID', 'USELOCATION', 'LOCATION_ID', 'DAY_GUID']], on='STOP_GUID')
            .merge(days[['DAY_GUID', 'TRIPDATE', 'TOTALCLIENTSONDAY', 'TRIP_GUID']], on='DAY_GUID'))

        return sorted(use_report.to_dict(orient='records'), key=lambda x: x['TRIPDATE'])

    def execute(self, parameters, messages):

        businessname = parameters[0].value
        startdate = parameters[1].value
        enddate = parameters[2].value
        savepath = parameters[3].value.value

        trips = OrderedDict()
        certification = None
        query = """"BUSINESSNAME" = '{}' AND "ENDDATE" >= date '{}' AND "ENDDATE" <= date '{}'"""
        query = query.format(businessname, startdate.strftime('%Y-%m-%d'), enddate.strftime('%Y-%m-%d'))
        with arcpy.da.SearchCursor(self.tables[-1], '*', query) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                certification = certification or row['CERTIFICATION']
                trips[row['TRIP_GUID']] = {'startdate': row['STARTDATE'],
                                           'enddate': row['ENDDATE'],
                                           'tripclients': row['MAXCLIENTS'],
                                           'category': row['USECATEGORY'],
                                           'submitter': (row['REPORTERNAME'], row['CREATIONDATE']),
                                           'activities': []}

        trips = OrderedDict(sorted(trips.items(), key=lambda x: (x[1]['startdate'], x[1]['enddate'])))

        if trips:
            arcpy.AddMessage("Processed {} trip records...".format(len(trips)))
        else:
            arcpy.AddMessage("No trips to report in date range.")
            return

        use_reports = 0
        for row in self._read_activities(trips.keys()):
            use_reports += 1
            if row['LOCATION_ID'].startswith('+'):
                latitude, longitude = [float(coord) for coord in row['LOCATION_ID'].split('_')]
            else:
                latitude, longitude = (None, None)

            record = [row['TRIPDATE'],
                      row['TOTALCLIENTSONDAY'],
                      row['USELOCATION'],
                      latitude,
                      longitude,
                      row['ACTIVITY'],
                      row['CLIENTNUMBER'],
                      row['GROUPNUMBER'],
                      row['GUIDENUMBER'],
                      row['HOURSSPENTONFS']]
            if row['TRIP_GUID'] in trips.keys():
                trips[row['TRIP_GUID']]['activities'].append(record)

        for trip in trips.keys():
            try:
                trips[trip]['activities'] = sorted(trips[trip]['activities'], key=lambda x: x[0])
            except Exception as e:
                arcpy.AddError("ERROR: {}\nTRIPID {}: {}\nACTIVITY DATA: {}".format(e, trips[trip], trips[trip]['activities']))
                raise e

        arcpy.AddMessage("Processed {} use report records.".format(use_reports))

        report = GuidedRecDetail(businessname, startdate, enddate, certification, savepath)

        arcpy.AddMessage('Saving {}...'.format(report.savefile))
        report.save_trips(trips)


class BusinessXlsCreator(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Business Name Excel Creator (Step 1)"
        self.description = "Used to create the Excel file used for business name updates in" + \
                           "the tool Business Name Updater (Step 2). First create the Excel file, " + \
                           "and then edit it to populate the 'New' column. Then proceed to Step 2 to update in SDE."
        self.canRunInBackground = False

    def getParameterInfo(self):
        """Define parameter definitions"""
        env.overwriteOutput = True
        param0 = arcpy.Parameter(
            displayName="SDE Connection",
            name="sdeConn",
            datatype="DEWorkspace",
            parameterType="Required",
            direction="Input")
        param0.value = r'T:\FS\Reference\GeoTool\agency\DatabaseConnection\r10_default_as_myself.sde'
        # param1.filter.type = "Workspace"
        # param1.filter.list = ["Remote Database"]

        param1 = arcpy.Parameter(
            displayName="Excel File",
            name="xlsPath",
            datatype="DEFile",
            parameterType="Required",
            direction="Output")
        param1.filter.list = ['xlsx', 'xls']
        param1.value = r'T:\FS\NFS\R10\Program\2700SpecialUsesMgmt\GIS\R10OGDatabase\ToolData\BusinessNameUpdates.xlsx'

        param2 = arcpy.Parameter(
            displayName="Dataset Name",
            name="datasetName",
            datatype="String",
            parameterType="Required",
            direction="Input")
        param2.filter.list = ['GUIDEDREC', 'HUNTING', 'HELISKI', 'OUTFITTING', 'MENDENHALL', 'ICEFIELD', 'ALL']

        params = [param0, param1, param2]
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""

        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""

        return

    def execute(self, parameters, messages):
        """The source code of the tool."""
        env.overwriteOutput = True
        # xl = parameters[1].value.value
        xl = parameters[1].valueAsText
        ds = parameters[2].valueAsText
        arcpy.AddMessage("     Output Excel file: " + xl)
        # sde = parameters[0].value.value
        sde = parameters[0].valueAsText
        env.workspace = sde
        objFilter = 'S_R10.R10_OFG_{}_*'.format(ds)
        dsList = [os.path.basename(obj) for obj in arcpy.ListFeatureClasses(objFilter) if
                  not obj.endswith('_MV') and not obj.endswith('_VIEW') and 'CHRISTAL' not in obj]
        dsList += [os.path.basename(obj) for obj in arcpy.ListTables(objFilter) if
                   not obj.endswith('_MV') and not obj.endswith('__ATTACH') and not obj.endswith(
                       '_VIEW') and 'CHRISTAL' not in obj]

        if ds == 'ALL':
            ofgList = ['GUIDEDREC', 'HELISKI', 'HUNTING', 'ICEFIELD', 'MENDENHALL', 'OUTFITTING']
        else:
            ofgList = [ds]
        arcpy.AddMessage("     SDE connection: " + sde)
        arcpy.AddMessage("     Dataset: " + ds)
        df_all = pd.DataFrame(columns=['Current', 'New', 'Dataset'])
        for o in ofgList:
            objFilter = 'S_R10.R10_OFG_{}_*'.format(o)
            dsList = [os.path.basename(obj) for obj in arcpy.ListFeatureClasses(objFilter) if
                      not obj.endswith('_MV') and not obj.endswith('_VIEW') and 'CHRISTAL' not in obj]
            dsList += [os.path.basename(obj) for obj in arcpy.ListTables(objFilter) if
                       not obj.endswith('_MV') and not obj.endswith('__ATTACH') and not obj.endswith(
                           '_VIEW') and 'CHRISTAL' not in obj]
            i = 0
            while i in range(len(dsList)):
                d = dsList[i]
                inData = os.path.join(sde, d)
                arcpy.AddMessage("Reading BUSINESSNAME data from {}".format(d))
                npArr = np.unique(arcpy.da.FeatureClassToNumPyArray(inData, ["BUSINESSNAME"], skip_nulls=True))
                npArr.dtype.names = ('Current',)
                df = pd.DataFrame(npArr, columns=['Current', 'New'])
                df.insert(2, "Dataset", d.split(".")[1])
                df_all = df_all.append(df, ignore_index=True)
                i += 1
        arcpy.AddMessage("Converting to Excel...")
        if ds == 'ALL':
            df_all_out = df_all.drop_duplicates(subset=['Current', 'Dataset'])
            df_all_out.to_excel(xl, sheet_name=ds, index=False, columns=['Current', 'Dataset'])
        else:
            df_ds_out = df_all.drop_duplicates(subset=['Current'])
            df_ds_out.to_excel(xl, sheet_name=ds, index=False, columns=['Current', 'New'])

        arcpy.AddMessage("Completed business name export to Excel for {}".format(ds))
        arcpy.AddMessage("Output Excel: {}".format(xl))


class BusinessXlsUpdater(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Business Name Updater (Step 2)"
        self.description = "Used to update the business name in the Outfitter Guide data tables. " + \
                           "You must first run Step 1 and choose the business area, then edit the xlsx " + \
			   "spreadsheet to populate the 'New' column before running this tool."
        self.canRunInBackground = False

    def getParameterInfo(self):
        """Define parameter definitions"""
        param0 = arcpy.Parameter(
            displayName="Excel File",
            name="xlsPath",
            datatype="DEFile",
            parameterType="Required",
            direction="Input")
        param0.value = r'T:\FS\NFS\R10\Program\2700SpecialUsesMgmt\GIS\R10OGDatabase\ToolData\BusinessNameUpdates.xlsx'

        param1 = arcpy.Parameter(
            displayName="SDE Connection",
            name="sdeConn",
            datatype="DEWorkspace",
            parameterType="Required",
            direction="Input")
        param1.value = r'T:\FS\Reference\GeoTool\agency\DatabaseConnection\r10_default_as_myself.sde'
        # param1.filter.type = "Workspace"
        # param1.filter.list = ["Remote Database"]

        param2 = arcpy.Parameter(
            displayName="Edit Version",
            name="editVer",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        params = [param0, param1, param2]
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        if parameters[0].altered:
            parameters[0].filter.list = ['xls', 'xlsx']
        if parameters[1].value:
            parameters[2].filter.list = [v.name for v in arcpy.da.ListVersions(parameters[1].value)
                                         if os.getenv('username').upper() in v.name]

        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""

        return

    def execute(self, parameters, messages):
        """The source code of the tool."""
        # xl = parameters[0].value.value
        xl = parameters[0].valueAsText
        arcpy.AddMessage("     Excel file: " + xl)
        # sde = parameters[1].value.value
        sde = parameters[1].valueAsText
        arcpy.AddMessage("     SDE connection: " + sde)
        env.workspace = sde
        vers = parameters[2].value
        arcpy.AddMessage("     Edit Version: " + vers)
        # create pandas data frame of excel file and drop all blanks in the New column
        arcpy.AddMessage("Reading Excel data....")
        ofg = pd.ExcelFile(xl).sheet_names[0]
        df = pd.read_excel(xl, header=0).replace("\'", "\'\'", regex=True).dropna()
        # df = df.replace("\'", "\'\'", regex=True)
        # list of fc and tables to update
        objFilter = 'S_R10.R10_OFG_{}_*'.format(ofg)
        updtObj = [os.path.basename(obj) for obj in arcpy.ListFeatureClasses(objFilter) if
                   not obj.endswith('_MV') and not obj.endswith(
                       '_VIEW') and 'CHRISTAL' not in obj]
        updtObj += [os.path.basename(obj) for obj in arcpy.ListTables(objFilter) if
                    not obj.endswith('_MV') and not obj.endswith('__ATTACH') and not obj.endswith(
                        '_VIEW') and 'CHRISTAL' not in obj]

        # turn excel data frame into a dictionary with Current:New
        data_dict = df.set_index('Current')['New'].to_dict()
        print(data_dict)

        # create temp sde connection file to user's version
        sdeTempPath = tempfile.mkdtemp()
        arcpy.AddMessage("Creating SDE connection to edit version....")
        arcpy.CreateDatabaseConnection_management(
            out_folder_path=sdeTempPath,
            out_name="MyEditVersion.sde", database_platform="ORACLE", instance="SDE_R10",
            account_authentication="OPERATING_SYSTEM_AUTH", username="", password="#", save_user_pass="SAVE_USERNAME",
            database="", schema="S_R10", version_type="TRANSACTIONAL",
            version=vers, date="")

        # Connect to SDE
        mySDE = os.path.join(sdeTempPath, 'MyEditVersion.sde')
        # arcpy.AddMessage("Temp SDE conn: {}".format(mySDE))
        env.workspace = mySDE
        # start edit session
        arcpy.AddMessage("Starting edit session....")
        try:
            edit = arcpy.da.Editor(mySDE)
            # Iterate through Excel data and update table(s)
            field = "BUSINESSNAME"
            arcpy.AddMessage("Beginning Business Name updates")
            valueList = list(data_dict.keys())
            nameCnt = len(valueList)
            arcpy.AddMessage("{} business names:".format(nameCnt))
            sql = """{} IN ('{}')""".format(arcpy.AddFieldDelimiters(os.path.join(sde, updtObj[0]), field),
                                            '\', \''.join(map(str, valueList)))
            for tbl in updtObj:
                if 'ACTIVITY' in tbl:
                    arcpy.MakeFeatureLayer_management(os.path.join(mySDE, tbl), "tblLyr", sql)
                else:
                    arcpy.MakeTableView_management(os.path.join(mySDE, tbl), "tblLyr", sql)
                arcpy.AddMessage("Updating {}".format(tbl))
                edit.startEditing(False, True)
                edit.startOperation()
                cnt = 0
                with arcpy.da.UpdateCursor("tblLyr", field) as uCur:
                    for row in uCur:
                        for bName in valueList:
                            if row[0] == bName.replace("\'\'", "\'"):
                                row[0] = data_dict[bName]
                                uCur.updateRow(row)
                                cnt += 1
                    else:
                        print("No rows in cursor")
                edit.stopOperation()
                edit.stopEditing(True)
                if arcpy.Exists("tblLyr"):
                    arcpy.Delete_management("tblLyr")
                if uCur:
                    del uCur
                if row:
                    del row
                arcpy.AddMessage("     Updated {} with {} records affected".format(tbl, cnt))
            arcpy.AddMessage("Finished updating BUSINESSNAME in edit version {}".format(vers))
            arcpy.AddMessage(
                "\n***** Verify edits, then contact SDE Manager to reconcile/post your version to DEFAULT *****\n")
        except:
            err = sys.exc_info()[:2]
            errtype = err[0]
            errval = str(err[1])
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)
            errmsg = str(errtype) + ': ' + errval
            arcpy.AddError(errmsg)
            arcpy.AddError("PYTHON ERRORS: \nTraceback Info:\n{0}".format(tbinfo))
            arcpy.AddError('Script failed on Line {}'.format(tb.tb_lineno))
            raise errtype(errval)
        finally:
            shutil.rmtree(sdeTempPath)
            
            
class OutfittingDetail:

    def __init__(self, business_name, startdate, enddate, reporter, certification, forestName, save_path):

        self.row = 4
        wbPath = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\outfitting_detail_template.xlsx'
        wb = load_workbook(wbPath)

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', business_name.title())
        savedates = startdate.strftime("%Y%m%d") + "_" + enddate.strftime("%Y%m%d")
        self.savefile = "OutfittingDetail_" + savedates + '_' + savebusiness + ".xlsx"
        self.path = os.path.join(save_path, self.savefile)

        trip_style = NamedStyle('trip_style')
        thin = Side(border_style='thin', color='000000')
        trip_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
#        trip_style.fill = PatternFill("solid", fgColor="AAAAAA")
        wb.add_named_style(trip_style)

        int_style = NamedStyle('int_style', number_format='0')
        int_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        int_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(int_style)

        coord_style = NamedStyle('coord_style', number_format='###.##0')
        coord_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        coord_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(coord_style)

        float_style = NamedStyle('float_style', number_format='0.000')
        float_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        float_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(float_style)

        str_style = NamedStyle('str_style')
        str_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        str_style.alignment = Alignment(horizontal='left', wrapText = True)
        wb.add_named_style(str_style)

        datetime_style = NamedStyle('datetime_style', number_format='MM/DD/YY')
        datetime_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        datetime_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(datetime_style)

        currancy_style = NamedStyle('curr_style', number_format=('#,##0.00'))
        currancy_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        currancy_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(currancy_style)
        
        
        self.styles = ['datetime_style', 'int_style', 'str_style',
                       'str_style', 'datetime_style', 'str_style']

        ws = wb.active
        ws['C2'] = certification
        header1 = 'Business/Organization: {}\nOperating Season: {}'.format(business_name,  enddate.strftime('%Y'))
        
        header2 = 'Outfitting Actual Use Report \n{} National Forest'.format(forestName) 
        footer = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
        ws.firstHeader.left.text = header1
        ws.firstFooter.left.text = footer
        ws.firstHeader.right.text = header2
        ws.oddHeader.left.text = header1
        ws.evenHeader.left.text = header1
        ws.oddFooter.left.text = footer
        ws.evenFooter.left.text = footer
        ws.oddHeader.right.text = header2
        ws.evenHeader.right.text = header2
        wb.save(self.path)
        
    @contextmanager
    def __worksheet(self):
    
        wb = load_workbook(self.path)
        ws = wb.active
        try:
            yield ws
        finally:
            wb.save(self.path)
    
    def save_trips(self, trips):
    
        with self.__worksheet() as ws:
            for trip in trips.keys():
                self._add_trip(ws, **trips[trip])
                
    def _add_trip(self, ws, tripDate, activities):

        # Skip to next page if trip records span page.
        current_page_line = (self.row - 27) % 29 if self.row > 27 else self.row
        current_page_line = 29 if not current_page_line else current_page_line
        trip_end_line = current_page_line + len(activities)
        if (self.row <= 27) and (trip_end_line > 27):
            self.row += 28 - current_page_line
        elif trip_end_line > 29:
            self.row += 30 - current_page_line

        for col in range(1, 6):
            cell = ws.cell(row=self.row, column=col)
            cell.style = 'trip_style'


#        self.row += 1
        self._add_activities(ws, activities)

    def _add_activities(self, worksheet, activity_records):

        for record in activity_records:
            for col, value in enumerate(record, 1):
                cell = worksheet.cell(row=self.row, column=col)
                cell.value = value
                cell.style = self.styles[col-1]
            self.row += 1
            
class OutfittingReport(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run Outfitting Detail Report"
        self.description = ""
        self.canRunInBackground = False
        
        
    def getParameterInfo(self):
        # Define parameter definitions

        connection = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        tables = ['S_R10.R10_OFG_OUTFITTING_ACTIVITY']
        self.tables = [connection + table for table in tables]

        sc = arcpy.da.SearchCursor(self.tables[0], "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        self.business_names = [row[0] for row in sc]
        self.parameters = self.getParameterInfo()

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = self.business_names

        param1  = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param1.value = datetime(datetime.today().year, 1, 1).strftime('%x')

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param2.value = datetime.today().strftime('%x')

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]

        return params


    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True


    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        
        businessname = parameters[0].value
        startdate = parameters[1].value
        enddate = parameters[2].value
        savepath = parameters[3].value.value
    
        arcpy.MakeQueryTable_management(self.tables, 'QueryTable', 'NO_KEY_FIELD')

            
    
        trips = OrderedDict()
        reporter = None
        certification = None
        forestName = None
        query1 = "S_R10.R10_OFG_OUTFITTING_ACTIVITY.BUSINESSNAME = '{}' AND S_R10.R10_OFG_OUTFITTING_ACTIVITY.TRIPDATE >= date '{}' AND S_R10.R10_OFG_OUTFITTING_ACTIVITY.TRIPDATE <= date '{}'".format(businessname, startdate.strftime('%Y-%m-%d'), enddate.strftime('%Y-%m-%d'))

    
        with arcpy.da.SearchCursor(self.tables[-1], '*', query1) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                reporter = reporter or row['REPORTERNAME']
                certification = certification or row['CERTIFICATION']
                forestName = forestName or row['FORESTNAME']
                trips[row['TRIP_GUID']] = {'tripDate' : row['TRIPDATE'],
                                                  'activities': []}
                     
    
        trips = OrderedDict(sorted(trips.items(), key=lambda x:(x[1]['tripDate'])))
    
        if trips: 
            arcpy.AddMessage("Processing {} trip records...".format(len(trips)))
        else:
            arcpy.AddWarning("No trips to report in date range.")
            return
    
        query2 = "S_R10.R10_OFG_OUTFITTING_ACTIVITY.BUSINESSNAME = '{}' AND S_R10.R10_OFG_OUTFITTING_ACTIVITY.TRIPDATE >= date '{}' AND S_R10.R10_OFG_OUTFITTING_ACTIVITY.TRIPDATE <= date '{}'".format(businessname, startdate.strftime('%Y-%m-%d'), enddate.strftime('%Y-%m-%d'))
        use_reports = 0
        with arcpy.da.SearchCursor("QueryTable", '*', query2) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                use_reports += 1
                row = dict(zip(fields, row))
                record = [row['TRIPDATE'],
                          row['TOTALCLIENTSONDAY'],
                          row['USELOCATION'],
                          row['ACTIVITY'],
                          row['CREATIONDATE'],
                          row['REPORTERNAME']]
                if row['TRIP_GUID'] in trips.keys():
                    trips[row['TRIP_GUID']]['activities'].append(record)
    
        for trip in trips.keys():
            try:
                trips[trip]['activities'] = sorted(trips[trip]['activities'], key=lambda x:x[0])
            except Exception as e:
                arcpy.AddError("ERROR: {}\nOUTFITTING_GUID {}: {}\nACTIVITY DATA: {}")
                raise e
    
        arcpy.AddMessage("Processing {} use report records.".format(use_reports))
    
        report = OutfittingDetail(businessname, startdate, enddate,
                                     reporter, certification, forestName, savepath)
    
        arcpy.AddMessage('Saving {}...'.format(report.savefile))
        report.save_trips(trips)
        

class MendenhallDetail:

    def __init__(self, business_name, startdate, enddate, reporter, certification, save_path):

        self.row = 4
        wbPath = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\mendenhall_detail_template.xlsx'
        wb = load_workbook(wbPath)

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', business_name.title())
        savedates = str(int(startdate)) + "_" + str(int(enddate))
        self.savefile = "MendenhallDetail_" + savedates + '_' + savebusiness + ".xlsx"
        self.path = os.path.join(save_path, self.savefile)

        trip_style = NamedStyle('trip_style')
        thin = Side(border_style='thin', color='000000')
        trip_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        trip_style.fill = PatternFill("solid", fgColor="AAAAAA")
        wb.add_named_style(trip_style)

        int_style = NamedStyle('int_style', number_format='0')
        int_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        int_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(int_style)

        coord_style = NamedStyle('coord_style', number_format='###.##0')
        coord_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        coord_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(coord_style)

        float_style = NamedStyle('float_style', number_format='0.000')
        float_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        float_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(float_style)

        str_style = NamedStyle('str_style')
        str_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        str_style.alignment = Alignment(horizontal='left', wrapText = True)
        wb.add_named_style(str_style)

        datetime_style = NamedStyle('datetime_style', number_format='MM/DD/YY')
        datetime_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        datetime_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(datetime_style)

        self.styles = ['str_style', 'int_style',
                       'int_style', 'int_style', 'str_style', 'datetime_style', 'str_style']

        ws = wb.active
        ws['C2'] = certification
        header = 'Business/Organization: {}\nOperating Season: {} - {}'.format(business_name, int(startdate),
                                                                                                   int(enddate))
        footer = 'Report Generated: {}\n*None is not zero. It means that no data was collected.'.format(datetime.today().strftime('%m/%d/%Y'))
        ws.firstHeader.left.text = header
        ws.firstFooter.left.text = footer
        ws.oddHeader.left.text = header
        ws.evenHeader.left.text = header
        ws.oddFooter.left.text = footer
        ws.evenFooter.left.text = footer
        wb.save(self.path)

    @contextmanager
    def __worksheet(self):

        wb = load_workbook(self.path)
        ws = wb.active
        try:
            yield ws
        finally:
            wb.save(self.path)

    def save_trips(self, trips):

        with self.__worksheet() as ws:
            for trip_guid in trips.keys():
                self._add_trip(ws, **trips[trip_guid])

    def _add_trip(self, ws, reportMonth, tripMonth, clientMonth, sum16up, sum15down, activities):

        # Skip to next page if trip records span page.
        current_page_line = (self.row - 27) % 29 if self.row > 27 else self.row
        current_page_line = 29 if not current_page_line else current_page_line
        trip_end_line = current_page_line + len(activities)
        if (self.row <= 27) and (trip_end_line > 27):
            self.row += 28 - current_page_line
        elif trip_end_line > 29:
            self.row += 30 - current_page_line

        for col in range(1, 7):
            cell = ws.cell(row=self.row, column=col)
            cell.style = 'trip_style'

        ws.merge_cells('B{}:C{}'.format(self.row, self.row))
        ws.merge_cells('F{}:G{}'.format(self.row, self.row))

        
        ws['A{}'.format(self.row)] = "Month: {}".format(reportMonth)
        ws['B{}'.format(self.row)] = 'Trips Per Month: {}'.format(tripMonth)
        ws['D{}'.format(self.row)] = 'Clients: {}'.format(clientMonth)
        ws['E{}'.format(self.row)] = 'Clients > 16: {}'.format(sum16up)
        ws['F{}'.format(self.row)] = 'Clients < 15: {}'.format(sum15down)
        self.row += 1
        self._add_activities(ws, activities)

    def _add_activities(self, worksheet, activity_records):

        for record in activity_records:
            for col, value in enumerate(record, 1):
                cell = worksheet.cell(row=self.row, column=col)
                cell.value = value
                cell.style = self.styles[col-1]
            self.row += 1
            

class MendenhallReport(object):
    def __init__(self):
        self.label = "Run Mendenhall Detail Report"
        self.canRunInBackground = False




    def getParameterInfo(self):
        # Define parameter definitions
        
        connection = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        tables = ['S_R10.R10_OFG_MENDENHALL_ACTIVITY',
                  'S_R10.R10_OFG_MENDENHALL_TRIPMONTH']
        self.tables = [connection + table for table in tables]

        sc = arcpy.da.SearchCursor(self.tables[1], "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        self.business_names = [row[0] for row in sc]
        self.parameters = self.getParameterInfo()

        self.whereClause = "S_R10.R10_OFG_MENDENHALL_TRIPMONTH.TRIP_GUID ="\
                            "S_R10.R10_OFG_MENDENHALL_ACTIVITY.TRIP_GUID"

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = self.business_names

        param1  = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param1.value = startyear

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param2.value = endyear

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]

        return params

    def isLicensed(self):  # optional
        return True

    def updateParameters(self, parameters):  # optional
        return

    def updateMessages(self, parameters):  # optional
        return

    def execute(self, parameters, messages):

        businessname = parameters[0].value
        startdate = parameters[1].value
        enddate = parameters[2].value
        savepath = parameters[3].value.value

        arcpy.MakeQueryTable_management(self.tables, 'QueryTable', 'NO_KEY_FIELD', '', '', self.whereClause)

        trips = OrderedDict()
        reporter = None
        certification = None
        query1 = "S_R10.R10_OFG_MENDENHALL_TRIPMONTH.BUSINESSNAME = '{}' AND S_R10.R10_OFG_MENDENHALL_TRIPMONTH.REPORTYEAR >= '{}' AND S_R10.R10_OFG_MENDENHALL_TRIPMONTH.REPORTYEAR <= '{}'".format(businessname, startdate, enddate)

        with arcpy.da.SearchCursor(self.tables[-1], '*', query1) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                reporter = reporter or row['REPORTERNAME']
                certification = certification or row['CERTIFICATION']
                trips[row['TRIP_GUID']] = {'reportMonth' : row['REPORTMONTH'],
                                            'tripMonth' : row['TRIPSPERMONTH'],
                                            'clientMonth' : row['CLIENTMONTH'],
                                            'sum16up' : row['SUMCLIENT16UP'],
                                            'sum15down' : row['SUMCLIENT15BELOW'],
                                            'activities': []}

        monthDict = {'January': 1, 'February':2, 'March':3, 'April':4, 'May':5, 'June':6, 'July':7, 'August':8, 'September':9, 'October':10, 'November':11, 'December':12}
        
        
        ordered_records = OrderedDict()
        for key in sorted(trips, key=lambda x: monthDict[trips[x]['reportMonth']]):
            ordered_records[key] = trips[key]
        
        trips = ordered_records


        if trips:
            arcpy.AddMessage("Processing {} trip records...".format(len(trips)))
        else:
            arcpy.AddWarning("No trips to report in date range.")
            return

        query2 = "S_R10.R10_OFG_MENDENHALL_TRIPMONTH.BUSINESSNAME = '{}' AND S_R10.R10_OFG_MENDENHALL_TRIPMONTH.REPORTYEAR >= '{}' AND S_R10.R10_OFG_MENDENHALL_TRIPMONTH.REPORTYEAR <= '{}'".format(businessname, startdate, enddate)
        use_reports = 0
                    
                    
        
        with arcpy.da.SearchCursor("QueryTable", '*', query2) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                use_reports += 1
                actCol = {}
                for i, j in zip(row, fields):
                    actCol.setdefault(i, []).append(j)
                actCol = ','.join(actCol[1]) if 1 in actCol else ''
                
                row = dict(zip(fields, row)) 
                record = [row['USELOCATION'],
                          row['CLIENTSLOCATION'],
                          row['CLIENTS16OLDER'],
                          row['CLIENTS15YOUNGER'],
                          actCol,
                          row['CREATIONDATE'],
                          row['REPORTERNAME']]
                if row['TRIP_GUID'] in trips.keys():
                    trips[row['TRIP_GUID']]['activities'].append(record)

        for trip in trips.keys():
            try:
                trips[trip]['activities'] = sorted(trips[trip]['activities'], key=lambda x:x[0])
            except Exception as e:
                arcpy.AddError("ERROR: {}\nTRIPID {}: {}\nACTIVITY DATA: {}".format(e, trips[trip], trips[trip]['activities']))
                raise e

        arcpy.AddMessage("Processing {} use report records.".format(use_reports))

        report = MendenhallDetail(businessname, startdate, enddate,
                                 reporter, certification, savepath)

        arcpy.AddMessage('Saving {}...'.format(report.savefile))
        report.save_trips(trips)
        
class IcefieldDetail:

    def __init__(self, business_name, startdate, enddate, reporter, certification, forestName, save_path):

        self.row = 4
        wbPath = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\icefields_detail_template.xlsx'
        wb = load_workbook(wbPath)

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', business_name.title())
        savedates = str(int(startdate)) + "_" + str(int(enddate))
        self.savefile = "Icefields_" + savedates + '_' + savebusiness + ".xlsx"
        self.path = os.path.join(save_path, self.savefile)

        trip_style = NamedStyle('trip_style')
        thin = Side(border_style='thin', color='000000')
        trip_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        trip_style.fill = PatternFill("solid", fgColor="AAAAAA")
        wb.add_named_style(trip_style)

        int_style = NamedStyle('int_style', number_format='0')
        int_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        int_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(int_style)

        coord_style = NamedStyle('coord_style', number_format='###.##0')
        coord_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        coord_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(coord_style)

        float_style = NamedStyle('float_style', number_format='0.000')
        float_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        float_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(float_style)

        str_style = NamedStyle('str_style')
        str_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        str_style.alignment = Alignment(horizontal='left', wrapText = True)
        wb.add_named_style(str_style)

        datetime_style = NamedStyle('datetime_style', number_format='MM/DD/YY')
        datetime_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        datetime_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(datetime_style)

        self.styles = ['str_style', 'int_style', 'int_style', 'int_style', 'int_style', 'int_style', 'int_style', 'datetime_style', 'str_style']

        ws = wb.active
        ws['C2'] = certification
        header1 = 'Business/Organization: {}\nOperating Season: {}-{}'.format(business_name, int(startdate),
                                                                                                   int(enddate))
        header2 = 'Icefields Actual Use Report \n{} National Forest'.format(forestName)
        footer = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
        ws.firstHeader.left.text = header1
        ws.firstFooter.left.text = footer
        ws.oddHeader.left.text = header1
        ws.evenHeader.left.text = header1
        ws.oddFooter.left.text = footer
        ws.evenFooter.left.text = footer
        ws.firstHeader.right.text = header2
        ws.oddHeader.right.text = header2
        ws.evenHeader.right.text = header2
        wb.save(self.path)

    @contextmanager
    def __worksheet(self):

        wb = load_workbook(self.path)
        ws = wb.active
        try:
            yield ws
        finally:
            wb.save(self.path)

    def save_trips(self, trips):

        with self.__worksheet() as ws:
            for trip_guid in trips.keys():
                self._add_trip(ws, **trips[trip_guid])

    def _add_trip(self, ws, reportMonth, landMonth, clientMonth, activities):

        # Skip to next page if trip records span page.
        current_page_line = (self.row - 27) % 29 if self.row > 27 else self.row
        current_page_line = 29 if not current_page_line else current_page_line
        trip_end_line = current_page_line + len(activities)
        if (self.row <= 27) and (trip_end_line > 27):
            self.row += 28 - current_page_line
        elif trip_end_line > 29:
            self.row += 30 - current_page_line

        for col in range(1, 9):
            cell = ws.cell(row=self.row, column=col)
            cell.style = 'trip_style'

        ws.merge_cells('A{}:C{}'.format(self.row, self.row))
        ws.merge_cells('D{}:E{}'.format(self.row, self.row))
        ws.merge_cells('F{}:I{}'.format(self.row, self.row))
        ws['A{}'.format(self.row)] = "Month: {}".format(reportMonth)
        ws['D{}'.format(self.row)] = 'Landings: {}'.format(landMonth)
        ws['F{}'.format(self.row)] = 'Clients: {}'.format(clientMonth)
        self.row += 1
        self._add_activities(ws, activities)

    def _add_activities(self, worksheet, activity_records):

        for record in activity_records:
            for col, value in enumerate(record, 1):
                cell = worksheet.cell(row=self.row, column=col)
                cell.value = value
                cell.style = self.styles[col-1]
            self.row += 1
            
class IcefieldReport(object):
    def __init__(self):
        self.label = "Run Icefields Detail Report"
        self.canRunInBackground = False




    def getParameterInfo(self):
        # Define parameter definitions
        
        connection = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        tables = ['S_R10.R10_OFG_ICEFIELD_ACTIVITY',
                  'S_R10.R10_OFG_ICEFIELD_TRIPMONTH']
        self.tables = [connection + table for table in tables]

        sc = arcpy.da.SearchCursor(self.tables[0], "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        self.business_names = [row[0] for row in sc]
        self.parameters = self.getParameterInfo()

        self.whereClause = "S_R10.R10_OFG_ICEFIELD_TRIPMONTH.TRIP_GUID ="\
                            "S_R10.R10_OFG_ICEFIELD_ACTIVITY.TRIP_GUID"

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = self.business_names

        param1  = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param1.value = startyear 

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param2.value = endyear

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]

        return params

    def isLicensed(self):  # optional
        return True

    def updateParameters(self, parameters):  # optional
        return

    def updateMessages(self, parameters):  # optional
        return

    def execute(self, parameters, messages):

        businessname = parameters[0].value
        startdate = parameters[1].value
        enddate = parameters[2].value
        savepath = parameters[3].value.value

        arcpy.MakeQueryTable_management(self.tables, 'QueryTable', 'NO_KEY_FIELD', '', '', self.whereClause)

        trips = OrderedDict()
        reporter = None
        certification = None
        forestName = 'Tongass'
        query1 = "S_R10.R10_OFG_ICEFIELD_TRIPMONTH.BUSINESSNAME = '{}' AND S_R10.R10_OFG_ICEFIELD_TRIPMONTH.REPORTYEAR >= '{}' AND S_R10.R10_OFG_ICEFIELD_TRIPMONTH.REPORTYEAR <= '{}'".format(businessname, startdate, enddate)

        with arcpy.da.SearchCursor(self.tables[-1], '*', query1) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                reporter = reporter or row['REPORTERNAME']
                certification = certification or row['CERTIFICATION']
                trips[row['TRIP_GUID']] = {'reportMonth' : row['REPORTMONTH'],
                                            'landMonth' : row['LDNGMONTH'],
                                            'clientMonth' : row['CLIENTMONTH'],
                                            'activities': []}
                
        monthDict = {'January': 1, 'February':2, 'March':3, 'April':4, 'May':5, 'June':6, 'July':7, 'August':8, 'September':9, 'October':10, 'November':11, 'December':12}
        
        
        ordered_records = OrderedDict()
        for key in sorted(trips, key=lambda x: monthDict[trips[x]['reportMonth']]):
            ordered_records[key] = trips[key]
        
        trips = ordered_records                


        if trips:
            arcpy.AddMessage("Processing {} trip records...".format(len(trips)))
        else:
            arcpy.AddWarning("No trips to report in date range.")
            return

        query2 = "S_R10.R10_OFG_ICEFIELD_TRIPMONTH.BUSINESSNAME = '{}' AND S_R10.R10_OFG_ICEFIELD_TRIPMONTH.REPORTYEAR >= '{}' AND S_R10.R10_OFG_ICEFIELD_TRIPMONTH.REPORTYEAR <= '{}'".format(businessname, startdate, enddate)
        use_reports = 0
        with arcpy.da.SearchCursor("QueryTable", '*', query2) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                use_reports += 1
                row = dict(zip(fields, row))
                record = [row['USELOCATION'],
                          row['LDNGOPS'],
                          row['LDNGGRATUITY'],
                          row['LDNGPAIDCLIENTS'],
                          row['CLIENTSGLACTREK'],
                          row['CLIENTSDOGSLED'],
                          row['CLIENTSHIKE'],
                          row['CREATIONDATE'], 
                          row['REPORTERNAME']]
                if row['TRIP_GUID'] in trips.keys():
                    trips[row['TRIP_GUID']]['activities'].append(record)

        for trip in trips.keys():
            try:
                trips[trip]['activities'] = sorted(trips[trip]['activities'], key=lambda x:x[0])
            except Exception as e:
                arcpy.AddError("ERROR: {}\nTRIPID {}: {}\nACTIVITY DATA: {}".format(e, trips[trip], trips[trip]['activities']))
                raise e

        arcpy.AddMessage("Processing {} use report records.".format(use_reports))

        report = IcefieldDetail(businessname, startdate, enddate,
                                 reporter, certification, forestName, savepath)

        arcpy.AddMessage('Saving {}...'.format(report.savefile))
        report.save_trips(trips)
        
        
class IcefieldSummary(object):
    """ This tool totals the clients and landings for all Icefiled activities for the previous 10 years."""
    
    def __init__(self):
        self.label = "Run Icefields Summary Report"
        self.canRunInBackground = False


    def getParameterInfo(self):

        param0 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param0.filter.list = ["File System"]

        params = [param0]

        return params
    
    def execute(self, parameters, messages):
        savepath = parameters[0].value.value        
        savedates = str(datetime.today().strftime('%m%d%Y'))
        savefile = "IcefieldsSummary_" + savedates +".xlsx"
        fullpath = os.path.join(savepath, savefile)
        startYear = str(datetime.today().year -10)
        endYear = str(datetime.today().year)

        connection = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        actTable = '{}S_R10.R10_OFG_ICEFIELD_ACTIVITY'.format(connection)
        tripTable = '{}S_R10.R10_OFG_ICEFIELD_TRIPMONTH'.format(connection)

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames 
        
        query = "REPORTYEAR >= '{}' AND REPORTYEAR <= '{}'".format(startYear, endYear)
        tripFields = getFieldNames(tripTable)
        dataTrip = [row for row in arcpy.da.SearchCursor(tripTable, tripFields, where_clause=query)]
        tripDF = pd.DataFrame(dataTrip, columns=tripFields)
        
        actFields = getFieldNames(actTable)
        dataAct = [row for row in arcpy.da.SearchCursor(actTable, actFields)]
        actDF = pd.DataFrame(dataAct, columns=actFields)
        
        allDF = actDF.merge(tripDF, how = 'inner', on = 'TRIP_GUID')

        valColumns = [u'CLIENTSGLACTREK', u'CLIENTSDOGSLED', u'CLIENTSHIKE', u'CLIENTSHELITOURSLEGACY']
        valLandings = [u'LDNGGRATUITY', u'LDNGPAIDCLIENTS']
        allDF['ClientActTotal'] = allDF[valColumns].sum(axis=1)
        allDF['LandingActTotal'] = allDF[valLandings].sum(axis=1)
        
        allDF.rename(columns={'BUSINESSNAME_x':'BUSINESSNAME'}, inplace=True)

        pvt = pd.pivot_table(allDF, values=('ClientActTotal'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        pvtLand = pd.pivot_table(allDF, values=('LandingActTotal'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        pvtClientAll = pd.pivot_table(allDF, values=('ClientActTotal'), index=[u'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        pvtLandAll = pd.pivot_table(allDF, values=('LandingActTotal'), index=[u'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
         
        """Write Pivot table and apply styles and formating"""
        writer_args = {
            'path': fullpath,
            'mode': 'w',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')
        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def excelUpdate(ws, headerString):
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column)
            rows=ws.max_row
            title_row = '1'
            value_cells = 'B2:{}{}'.format(col, rows)
            targetRange = CellRange(range_string=value_cells)
            value_col = 2
            index_column = 'A'
            ws.column_dimensions[index_column].width = 42.57
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')  
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 12.29
                value_col += 1
            mergedcells=[]    
            for cell_group in ws.merged_cells.ranges:
                    if cell_group.issubset(targetRange):
                        mergedcells.append(cell_group)                
            for cell_group in mergedcells:
                min_col, min_row, max_col, max_row = cell_group.bounds
                ws.unmerge_cells(str(cell_group))
                for irow in range(min_row, max_row+1):
                	for jcol in range(min_col, max_col+1):
                		ws.cell(row =irow, column=jcol, value=None)       
            for row in ws[value_cells]:
                for cell in row:
                    cell.style = value_style            
            for cell in ws[index_column]:
                cell.style = index_style                
            for cell in ws[title_row]:
                cell.style = title_style
            header = headerString
            ws.oddHeader.left.text = header
            ws.oddHeader.left.size = 8
            ws.oddHeader.left.font = "Calibri, bold"
            ws.oddHeader.left.color = "000000"
        
        with pd.ExcelWriter(**writer_args) as xlsx:
            """All the variables needed to apply styles"""
            pvtClientAll.to_excel(xlsx, 'Totals')
            pvt.to_excel(xlsx, 'Clients')
            pvtLand.to_excel(xlsx, 'Landings')
            ws = xlsx.sheets['Clients']
            wsLand = xlsx.sheets['Landings']
            wsTotal = xlsx.sheets['Totals']
            date = datetime.today().strftime('%m/%d/%Y')
            totalHead = 'Juneau/Skagway Icefields Summary Report\nTotal Clients and Landings (Revenue and Gratuity) for last 10 years\nReport Generated: {}'.format(date)
            landHead = 'Juneau/Skagway Icefields Summary Report\nLandings (Revenue and Gratuity)\nReport Generated: {}'.format(date)
            clientHead = 'Juneau/Skagway Icefields Summary Report\nClients (Revenue and Gratuity)\nReport Generated: {}'.format(date)
            
            excelUpdate(ws, clientHead)
            excelUpdate(wsLand, landHead)
            excelUpdate(wsTotal, totalHead)                    
                            
            """Creates Formatting for Totals Sheet"""
            clientValue = 'B3:O13'
            landValue = 'B17:O27'
            clientIndex = '2'
            landIndex = '16'
            colIntTot= (wsTotal.max_column)
            value_col = 2            
            rowsTot=wsTotal.max_row

            wsTotal.insert_rows(1)
            wsTotal['A1']= 'Total Clients'
        
            
            cellEnd = wsTotal.cell(row=(rowsTot+3), column=1)
            cellEnd.value = 'Total Landings'
                
            pvtLandAll.to_excel(xlsx, 'Totals', startrow=(rowsTot+3), startcol=0)
            
            clientsTitle = '1'
            landTitle = '15'
            
            while value_col <= colIntTot:
                i = get_column_letter(value_col)
                wsTotal.column_dimensions[i].width = 12.29
                value_col += 1    
            
            for row in wsTotal[clientValue]:
                for cell in row:
                    cell.style = value_style
                    
            for row in wsTotal[landValue]:
                for cell in row:
                    cell.style = value_style
                
            for cell in wsTotal[clientIndex]:
                cell.style = title_style
                
            for cell in wsTotal[landIndex]:
                cell.style = title_style
                
            for cell in wsTotal[clientsTitle]:
                cell.style = title_style    
            
            for cell in wsTotal[landTitle]:
                cell.style = title_style 
            
class MendenhallSummary(object):
    """This tool summarizes the number of clients by location, by business for all Mendenhall activities for the year selected. It then breaks that out by month."""
    
    def __init__(self):
        self.label = "Run Mendenhall Summary Report"
        self.description = ""
        self.canRunInBackground = False
        
    def getParameterInfo(self):

        param0 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param0.filter.list = ["File System"]
        
        param1 = arcpy.Parameter(
            displayName="Report Year",
            name="reportYr",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param1.value = endyear

        params = [param0, param1]

        return params
    
    def execute(self, parameters, messages):

        savepath = parameters[0].value.value
        reportYr = parameters[1].value
        savedates = str(datetime.today().strftime('%m%d%Y'))
        savefile = "MendenhallSummary_" + savedates +".xlsx"
        fullpath = os.path.join(savepath, savefile)
  
        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames 

        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'
        mendTrip = '{}S_R10.R10_OFG_MENDENHALL_TRIPMONTH'.format(connection)
        mendAct = '{}S_R10.R10_OFG_MENDENHALL_ACTIVITY'.format(connection)
        
        
        query = "REPORTYEAR = {}".format(int(reportYr))
        tripCol = getFieldNames(mendTrip)
        mendTripSC = [row for row in arcpy.da.SearchCursor(mendTrip, tripCol, where_clause=query)]
        mendTripDF = pd.DataFrame(mendTripSC, columns = tripCol)
        
        actCol = getFieldNames(mendAct)
        mendActSC = [row for row in arcpy.da.SearchCursor(mendAct, actCol)]  
        mendActDF = pd.DataFrame(mendActSC, columns = actCol)
        
        allDF = mendActDF.merge(mendTripDF, how = 'inner', on = "TRIP_GUID")
        allDF.rename(columns={'BUSINESSNAME_x':'BUSINESSNAME'}, inplace=True)
        
        """This checks to see if there are trips by month. If there are a pivot table is created, if not a message is created that is then added to that month tab that there were zero trips"""

        allMonths = allDF.empty
        if allMonths == True:
            allEmpty = 'There were no trips in {}'.format(reportYr)
        else:
            pvt = pd.pivot_table(allDF, values=('CLIENTSLOCATION'), index=['BUSINESSNAME', 'REPORTYEAR'], columns=['USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        
        apr = allDF[allDF.REPORTMONTH =='April'].empty
        if apr == True:
            aprEmpty = 'There were no trips in April, {}'.format(reportYr)
        else:
            pvtApr = pd.pivot_table(allDF[allDF.REPORTMONTH == 'April'], values=('CLIENTSLOCATION'), index=['BUSINESSNAME', 'REPORTYEAR'], columns=['USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        
        may = allDF[allDF.REPORTMONTH == 'May'].empty
        if may == True:
            mayEmpty = 'There were no trips in May, {}'.format(reportYr)
        else:
            pvtMay =  pd.pivot_table(allDF[allDF.REPORTMONTH == 'May'], values=('CLIENTSLOCATION'), index=['BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')

        jun = allDF[allDF.REPORTMONTH == 'June'].empty
        if jun == True:
            junEmpty = 'There were no trips in June, {}'.format(reportYr)
        else: 
            pvtJun = pd.pivot_table(allDF[allDF.REPORTMONTH == 'June'], values=('CLIENTSLOCATION'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        
        jul = allDF[allDF.REPORTMONTH == 'July'].empty
        if jul == True:
            julEmpty = 'There were no trips in July, {}'.format(reportYr)
        else:
            pvtJul = pd.pivot_table(allDF[allDF.REPORTMONTH == 'July'], values=('CLIENTSLOCATION'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')

        aug = allDF[allDF.REPORTMONTH == 'August'].empty
        if aug == True:
            augEmpty = 'There were no trips in August, {}'.format(reportYr)
        else:
            pvtAug = pd.pivot_table(allDF[allDF.REPORTMONTH == 'August'], values=('CLIENTSLOCATION'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')

        sept = allDF[allDF.REPORTMONTH == 'September'].empty
        if sept == True:
            septEmpty = 'There were no trips in September, {}'.format(reportYr)
        else:
            pvtSept = pd.pivot_table(allDF[allDF.REPORTMONTH == 'September'], values=('CLIENTSLOCATION'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')
        
        octo = allDF[allDF.REPORTMONTH == 'October'].empty
        if octo == True:
            octoEmpty = 'There were no trips in October, {}'.format(reportYr)
        else:
            pvtOct = pd.pivot_table(allDF[allDF.REPORTMONTH == 'October'], values=('CLIENTSLOCATION'), index=[u'BUSINESSNAME', 'REPORTYEAR'], columns=[u'USELOCATION'], aggfunc=np.sum, margins = True, dropna = True, margins_name = 'Total')        
        
        """Write Pivot table and apply styles and formating"""
        writer_args = {
            'path': fullpath,
            'mode': 'w',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')
        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')


        """ Function applies all excel formating"""

        def excelUpdate(ws):
            title_row = 1    
            value_col = 2
            index_column = 'A'
            date = datetime.today().strftime('%m/%d/%Y')
            rowsTot=ws.max_row
            colInt= (ws.max_column)
            colTot=get_column_letter(ws.max_column)
            value_cellsTot = 'B2:{}{}'.format(colTot, rowsTot)
            targetRangeTot = CellRange(range_string=value_cellsTot)
            ws.column_dimensions[index_column].width = 42.57
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            mergedcells=[]            
            for cell_group in ws.merged_cells.ranges:
                    if cell_group.issubset(targetRangeTot):
                        mergedcells.append(cell_group)                    
            for cell_group in mergedcells:
                min_col, min_row, max_col, max_row = cell_group.bounds
                ws.unmerge_cells(str(cell_group))
                for irow in range(min_row, max_row+1):
                	for jcol in range(min_col, max_col+1):
                		ws.cell(row =irow, column=jcol, value=None)
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 12.29
                value_col += 1
            for row in ws[value_cellsTot]:
                for cell in row:
                    cell.style = value_style
            for cell in ws[index_column]:
                cell.style = index_style                    
            for cell in ws[title_row]:
                cell.style = title_style
            header1 = 'Mendenhall Annual Use Summary Report\nReport Year: {}\nReport Generated: {}'.format(reportYr, date)
            ws.oddHeader.left.text = header1
            ws.oddHeader.left.size = 8
            ws.oddHeader.left.font = "Calibri, bold"
            ws.oddHeader.left.color = "000000"
            ws.cell(column=2, row=rowsTot, value=None)
    


        with pd.ExcelWriter(**writer_args) as xlsx:
            """Writes the pivot table or message to the excel workbook. If the pivot is written, all formatting is applied."""
            
            if allMonths == True:   
                totDF = pd.DataFrame({'Message': [allEmpty]})
                totDF.to_excel(xlsx, 'Totals')
            else: 
                pvt.to_excel(xlsx, 'Totals') 
                wsTotals = xlsx.sheets['Totals']
                excelUpdate(wsTotals)   
                
            if apr == True:
                aprDF = pd.DataFrame({'Message': [aprEmpty]})
                aprDF.to_excel(xlsx, 'April')               
            else:
                pvtApr.to_excel(xlsx, 'April')
                wsApr = xlsx.sheets['April'] 
                excelUpdate(wsApr) 
                
            if may == True:
                mayDF = pd.DataFrame({'Message': [mayEmpty]})
                mayDF.to_excel(xlsx, 'May')
            else: 
                pvtMay.to_excel(xlsx, 'May')
                wsMay = xlsx.sheets['May'] 
                excelUpdate(wsMay)
            
            if jun == True:
                junDF = pd.DataFrame({'Message': [junEmpty]})
                junDF.to_excel(xlsx, 'June')
            else: 
                pvtJun.to_excel(xlsx, 'June')
                wsJun = xlsx.sheets['June'] 
                excelUpdate(wsJun)
                
            if jul == True:
                julDF = pd.DataFrame({'Message': [julEmpty]})
                julDF.to_excel(xlsx, 'July')
            else:
                pvtJul.to_excel(xlsx, 'July')
                wsJul = xlsx.sheets['July']   
                excelUpdate(wsJul)

            if aug == True:
                augDF = pd.DataFrame({'Message': [augEmpty]})
                augDF.to_excel(xlsx, 'August')
            else:
                pvtAug.to_excel(xlsx, 'August')
                wsAug = xlsx.sheets['August']
                excelUpdate(wsAug)
                
            if sept == True:
                septDF = pd.DataFrame({'Message': [septEmpty]})
                septDF.to_excel(xlsx, 'September')
            else:
                pvtSept.to_excel(xlsx, 'September')
                wsSept = xlsx.sheets['September']  
                excelUpdate(wsSept)

            if octo == True:
                octDF = pd.DataFrame({'Message': [octoEmpty]})
                octDF.to_excel(xlsx, 'October')
            else:
                pvtOct.to_excel(xlsx, 'October')
                wsOct = xlsx.sheets['October']  
                excelUpdate(wsOct)                  
        return  

class HeliskiDetail:

    def __init__(self, business_name, startdate, enddate, reporter, certification, forestName, save_path):

        self.row = 4
        wbPath = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\heliski_detail_template.xlsx'
        wb = load_workbook(wbPath)

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', business_name.title())
        savedates = str((startdate.strftime("%Y%m%d"))) + "_" + str((enddate.strftime("%Y%m%d")))
        self.savefile = "Heliski_" + savedates + '_' + savebusiness + ".xlsx"
        self.path = os.path.join(save_path, self.savefile)

        trip_style = NamedStyle('trip_style')
        thin = Side(border_style='thin', color='000000')
        trip_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        trip_style.fill = PatternFill("solid", fgColor="AAAAAA")
        wb.add_named_style(trip_style)

        int_style = NamedStyle('int_style', number_format='0')
        int_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        int_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(int_style)

        coord_style = NamedStyle('coord_style', number_format='###.##0')
        coord_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        coord_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(coord_style)

        float_style = NamedStyle('float_style', number_format='0.000')
        float_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        float_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(float_style)

        str_style = NamedStyle('str_style')
        str_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        str_style.alignment = Alignment(horizontal='left', wrapText = True)
        wb.add_named_style(str_style)

        datetime_style = NamedStyle('datetime_style', number_format='MM/DD/YY')
        datetime_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        datetime_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(datetime_style)
        
        time_style = NamedStyle('time_style', number_format= 'h:mm' )
        time_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        time_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(time_style)        

        self.styles = ['datetime_style', 'int_style', 'str_style', 'coord_style', 'coord_style', 'str_style', 'int_style', 'int_style', 'time_style', 'time_style', 'datetime_style', 'str_style']

        ws = wb.active
        ws['C2'] = certification
        header1 = 'Business/Organization: {}\nOperating Season: {}-{}'.format(business_name, startdate.strftime('%Y-%m-%d'),
                                                                                                   enddate.strftime('%Y-%m-%d'))
        header2 = 'Heliski Actual Use Report \n{} National Forest'.format(forestName)
        footer = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
        ws.firstHeader.left.text = header1
        ws.firstFooter.left.text = footer
        ws.oddHeader.left.text = header1
        ws.evenHeader.left.text = header1
        ws.oddFooter.left.text = footer
        ws.evenFooter.left.text = footer
        ws.firstHeader.right.text = header2
        ws.oddHeader.right.text = header2
        ws.evenHeader.right.text = header2
        wb.save(self.path)

    @contextmanager
    def __worksheet(self):

        wb = load_workbook(self.path)
        ws = wb.active
        try:
            yield ws
        finally:
            wb.save(self.path)

    def save_trips(self, trips):

        with self.__worksheet() as ws:
            for trip_guid in trips.keys():
                self._add_trip(ws, **trips[trip_guid])

    def _add_trip(self, ws, tripDate, activities):

        # Skip to next page if trip records span page.
        current_page_line = (self.row - 27) % 29 if self.row > 27 else self.row
        current_page_line = 29 if not current_page_line else current_page_line
        trip_end_line = current_page_line + len(activities)
        if (self.row <= 27) and (trip_end_line > 27):
            self.row += 28 - current_page_line
        elif trip_end_line > 29:
            self.row += 30 - current_page_line

        for col in range(1, 13):
            cell = ws.cell(row=self.row, column=col)
            cell.style = 'trip_style'

        self.row += 1
        self._add_activities(ws, activities)

    def _add_activities(self, worksheet, activity_records):

        for record in activity_records:
            for col, value in enumerate(record, 1):
                cell = worksheet.cell(row=self.row, column=col)
                cell.value = value
                cell.style = self.styles[col-1]
            self.row += 1


class HeliskiReport(object):
    def __init__(self):
        self.label = "Run Heliski Detail Report"
        self.canRunInBackground = False




    def getParameterInfo(self):
        # Define parameter definitions
        
        connection = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/'
        tables = ['S_R10.R10_OFG_HELISKI_ACTIVITY',
                  'S_R10.R10_OFG_HELISKI_TRIP']
        self.tables = [connection + table for table in tables]

        sc = arcpy.da.SearchCursor(self.tables[0], "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        self.business_names = [row[0] for row in sc]
        self.parameters = self.getParameterInfo()

        self.whereClause = "S_R10.R10_OFG_HELISKI_TRIP.TRIP_GUID ="\
                            "S_R10.R10_OFG_HELISKI_ACTIVITY.TRIP_GUID"        

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = self.business_names

        param1  = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param1.value = datetime(datetime.today().year, 1, 1).strftime('%x')

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param2.value = datetime.today().strftime('%x')

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]
        
        return params

    def isLicensed(self):  # optional
        return True

    def updateParameters(self, parameters):  # optional
        return

    def updateMessages(self, parameters):  # optional
        return

    def execute(self, parameters, messages):

        businessname = parameters[0].value
        startdate = parameters[1].value
        enddate = parameters[2].value
        savepath = parameters[3].value.value

        arcpy.MakeQueryTable_management(self.tables, 'QueryTable', 'NO_KEY_FIELD', '', '', self.whereClause)

        trips = OrderedDict()
        reporter = None
        certification = None
        forestName = None
        query1 = "S_R10.R10_OFG_HELISKI_TRIP.BUSINESSNAME = '{}' AND S_R10.R10_OFG_HELISKI_TRIP.TRIPDATE >= date '{}' AND S_R10.R10_OFG_HELISKI_TRIP.TRIPDATE <= date '{}'".format(businessname, startdate.strftime('%Y-%m-%d'), enddate.strftime('%Y-%m-%d'))

        with arcpy.da.SearchCursor(self.tables[-1], '*', query1) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                row = dict(zip(fields, row))
                reporter = reporter or row['REPORTERNAME']
                certification = certification or row['CERTIFICATION']
                forestName = forestName or row['FORESTNAME']
                trips[row['TRIP_GUID']] = {'tripDate' : row['TRIPDATE'],
                                                  'activities': []}
                     
    
        trips = OrderedDict(sorted(trips.items(), key=lambda x:(x[1]['tripDate'])))

        if trips:
            arcpy.AddMessage("Processing {} trip records...".format(len(trips)))
        else:
            arcpy.AddWarning("No trips to report in date range.")
            return

        query2 = "S_R10.R10_OFG_HELISKI_TRIP.BUSINESSNAME = '{}' AND S_R10.R10_OFG_HELISKI_TRIP.TRIPDATE >= date '{}' AND S_R10.R10_OFG_HELISKI_TRIP.TRIPDATE <= date '{}'".format(businessname, startdate.strftime('%Y-%m-%d'), enddate.strftime('%Y-%m-%d'))
        use_reports = 0
        with arcpy.da.SearchCursor("QueryTable", '*', query2) as sc:
            fields = [field.split('.')[-1] for field in sc.fields]
            for row in sc:
                use_reports += 1
                row = dict(zip(fields, row))
                enter = ()
                if row['ENTERTIME'] is not None: 
                    enter = datetime.strftime(row['ENTERTIME']+timedelta(days=365), '%H:%M')
                else:
                    enter = row['ENTERTIME']
                depart = ()
                if row['DEPARTTIME'] is not None:
                    depart = datetime.strftime(row['DEPARTTIME']+timedelta(days=365), '%H:%M')
                else:
                    depart = row['DEPARTTIME']
                record = [row['TRIPDATE'],
                          row['TOTALCLIENTSONDAY'],
                          row['USELOCATION'],
                          row['LATITUDE'],
                          row['LONGITUDE'],
                          row['ACTIVITY'],
                          row['CLIENTS_LOCATION'],
                          row['HOURSSPENTONFS'],
                          enter,
                          depart, 
                          row['CREATIONDATE'],
                          row['REPORTERNAME']]
                if row['TRIP_GUID'] in trips.keys():
                    trips[row['TRIP_GUID']]['activities'].append(record)

        for trip in trips.keys():
            try:
                trips[trip]['activities'] = sorted(trips[trip]['activities'], key=lambda x:x[0])
            except Exception as e:
                arcpy.AddError("ERROR: {}\nTRIPID: {}\nACTIVITY DATA: {}".format(e, trips[trip], trips[trip]['activities']))
                raise e

        arcpy.AddMessage("Processing {} use report records.".format(use_reports))

        report = HeliskiDetail(businessname, startdate, enddate,
                                 reporter, certification, forestName, savepath)

        arcpy.AddMessage('Saving {}...'.format(report.savefile))
        report.save_trips(trips)
        
        
class FiveYearSummary(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Five Year Summary"
        self.description = ""
        self.canRunInBackground = False
        



    def getParameterInfo(self):
        """Define parameter definitions"""
        
        self.heli = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_HELISKI_TRIP'
        self.ice = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_ICEFIELD_TRIPMONTH'
        self.guide = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_GUIDEDREC_DAY'
        self.hunt = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_HUNTING_DAY'
        self.mend = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_MENDENHALL_TRIPMONTH'
        self.outfit = r'T:/FS/Reference/GeoTool/agency/DatabaseConnection/r10_default_as_myself.sde/S_R10.R10_OFG_OUTFITTING_ACTIVITY'
             
       
        sc = arcpy.da.SearchCursor(self.heli, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus = [row[0] for row in sc]
        
        sc1 = arcpy.da.SearchCursor(self.ice, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus1 = [row[0] for row in sc1]
        
        sc2 = arcpy.da.SearchCursor(self.guide, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus2 = [row[0] for row in sc2]
        
        sc3 = arcpy.da.SearchCursor(self.hunt, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus3 = [row[0] for row in sc3]
        
        sc4 = arcpy.da.SearchCursor(self.mend, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus4 = [row[0] for row in sc4]
        
        sc5 = arcpy.da.SearchCursor(self.outfit, "BUSINESSNAME", sql_clause=("DISTINCT", "ORDER BY BUSINESSNAME"))
        bus5 = [row[0] for row in sc5]
        
            
        busList = bus1+bus2+bus3+bus4+bus+bus5   
        
        self.business_name = list(set(busList))
         
        self.business_name.sort()
               
        self.parameters = self.getParameterInfo()        

        param0 = arcpy.Parameter(
            displayName="For trips conducted by",
            name="business",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = self.business_name

        param1  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param1.value = startyear -5
        
        param2  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param2.value = endyear -1


        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]
        
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        
        """ Function to get list of field names in a table"""

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames




        """The source code of the tool."""
        

        
        businessname = parameters[0].value
        startYear = parameters[1].value
        endYear = parameters[2].value
        savepath = parameters[3].value.value
        
        start = str(int(startYear))
        end = str(int(endYear))

        savebusiness = re.sub(r'[^A-Za-z0-9]+', '', businessname.title())
        self.savefile = "FiveYearReview_" + savebusiness + "_" + start + "_" + end + ".xlsx"
        self.path = os.path.join(savepath, self.savefile)

        username = os.environ.get("USERNAME")
        
        
        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

        iceTrip = '{}S_R10.R10_OFG_ICEFIELD_TRIPMONTH'.format(connection) 
        iceAct = '{}S_R10.R10_OFG_ICEFIELD_ACTIVITY'.format(connection)
        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntTrip = '{}S_R10.R10_OFG_HUNTING_TRIP'.format(connection)
        huntHunt = '{}S_R10.R10_OFG_HUNTING_HUNTER'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        heliTrip ='{}S_R10.R10_OFG_HELISKI_TRIP'.format(connection)
        heliActivity = '{}S_R10.R10_OFG_HELISKI_ACTIVITY'.format(connection)
        mendTrip = '{}S_R10.R10_OFG_MENDENHALL_TRIPMONTH'.format(connection)
        mendActivity = '{}S_R10.R10_OFG_MENDENHALL_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)
        

            
        """ ReFormatting start and end date for Guided Recreation and Hunting so that it matches the timestamp format"""

        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
            
        """Creating Icefield dataframes and filtering data by the report timeframe and business name"""

        iceTripCol = getFieldNames(iceTrip)
        iceActCol = getFieldNames(iceAct)
        
        iceWhere = "REPORTYEAR>= {} AND REPORTYEAR<= {} AND BUSINESSNAME = '{}'".format(startYear, endYear, businessname)
        iceTripData =[row for row in arcpy.da.SearchCursor(iceTrip, iceTripCol, where_clause=iceWhere)]
        
        iceActWhere = "BUSINESSNAME = '{}'".format(businessname)
        iceActData = [row for row in arcpy.da.SearchCursor(iceAct, iceActCol, where_clause=iceActWhere)]

        iceTripDF =  pd.DataFrame(iceTripData, columns = iceTripCol)
        iceActDF = pd.DataFrame(iceActData, columns = iceActCol)
        
        if (iceTripDF.empty == False):
            if (iceActDF.empty == False):
                iceDF = iceActDF.merge(iceTripDF, how = 'inner', on = 'TRIP_GUID')
        
        """Creating Guided Recreation DataFrames and filtering data by the report timeframe and business name"""
        
        guideDayCol = getFieldNames(guideDay)
        guideStopCol = getFieldNames(guideStop)
        guideActCol = getFieldNames(guideActivity)
        
        guideWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND BUSINESSNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), businessname)
        guideData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideWhere)]
        guideDayDF = pd.DataFrame(guideData, columns = guideDayCol)
        
        guideStopWhere = "BUSINESSNAME = '{}'".format(businessname)
        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol, where_clause=guideStopWhere)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol)

        guideActData =[row for row in arcpy.da.SearchCursor(guideActivity, guideActCol, where_clause=guideStopWhere)]  
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)
        
        
        if (guideDayDF.empty == False):
            if (guideStopDF.empty ==False):
                if (guideActDF.empty == False):
                    guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
                    guideDF = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')
        
        """Creating Hunting Data Frames to only get the non-hunting activities and filtering data by the report timeframe and business name"""        

        huntDayCol = getFieldNames(huntDay)
        huntStopCol = getFieldNames(huntStop)
        huntActCol = getFieldNames(huntActivity)
    

        huntWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND BUSINESSNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), businessname)
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns = huntDayCol)
        
        huntStopWhere = "BUSINESSNAME = '{}'".format(businessname)
        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol, where_clause=huntStopWhere)]
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)
        
        huntActWhere = "BUSINESSNAME = '{}'".format(businessname) 
        huntActData =  [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol, where_clause=huntActWhere)]  
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
        
        if (huntDayDF.empty == False):
            if (huntStopDF.empty == False):
                if (huntActDF.empty == False):
                    huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
                    huntDF = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')        

        nonHuntWhere = "BUSINESSNAME = '{}'  AND SERVICE_DAYS_NONHUNTER > 0".format(businessname)
        nonHuntData = [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol, where_clause=nonHuntWhere)]
        nonHuntAct = pd.DataFrame(nonHuntData, columns = huntActCol)
        
        if (huntDayDF.empty == False):
            if (huntStopDF.empty == False):
                if (nonHuntAct.empty == False):
                    nonHuntDF1 = nonHuntAct.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
                    nonHuntDF = nonHuntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')       

        """Creating Hunting dataframes to get the count of hunts in the date range for the specific business name"""

        huntCountCol = getFieldNames(huntTrip)
        huntHuntCol = getFieldNames(huntHunt)
        
        huntCountWhere = "ENDDATE >= timestamp '{}' AND ENDDATE <= timestamp '{}' AND BUSINESSNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), businessname)          
        huntCountData = [row for row in arcpy.da.SearchCursor(huntTrip, huntCountCol, where_clause=huntCountWhere)]            
        huntTripDF = pd.DataFrame(huntCountData, columns = huntCountCol)
        
        huntHuntWhere = "BUSINESSNAME = '{}'".format(businessname) 
        huntHuntData = [row for row in arcpy.da.SearchCursor(huntHunt, huntHuntCol, where_clause=huntHuntWhere)]  
        huntHuntDF = pd.DataFrame(huntHuntData, columns = huntHuntCol)  
        
        if (huntTripDF.empty == False):
            if (huntHuntDF.empty == False):
                huntCountDF = huntHuntDF.merge(huntTripDF, how = 'inner', on = 'TRIP_GUID')
        
        
        """ Creating Heliski dataframes in the date range and for the specific business name"""

        heliTripCol = getFieldNames(heliTrip)
        heliActCol = getFieldNames(heliActivity)

        heliWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND BUSINESSNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), businessname)
        heliData = [row for row in arcpy.da.SearchCursor(heliTrip, heliTripCol, where_clause=heliWhere)]
        heliTripDF = pd.DataFrame(heliData, columns=heliTripCol)   
        
        heliActWhere = "BUSINESSNAME = '{}'".format(businessname)
        heliActData = [row for row in arcpy.da.SearchCursor(heliActivity, heliActCol, where_clause=heliActWhere)]
        heliActDF = pd.DataFrame(heliActData, columns = heliActCol)
        
        if (heliTripDF.empty == False):
            if (heliActDF.empty == False):
                heliDF = heliActDF.merge(heliTripDF, how = 'inner', on='TRIP_GUID')
        
        
        """Creating Mendenhall datafrmaes in the date range and for the specific business name"""

        mendTripCol = getFieldNames(mendTrip)
        mendActCol = getFieldNames(mendActivity)
        
        
        mendWhere = "REPORTYEAR >= {} AND REPORTYEAR <= {} AND BUSINESSNAME = '{}'".format(startYear, endYear, businessname)
        mendData = [row for row in arcpy.da.SearchCursor(mendTrip, mendTripCol, where_clause=mendWhere)]
        mendTripDF = pd.DataFrame(mendData, columns=mendTripCol)
        
        mendActWhere = "BUSINESSNAME = '{}'".format(businessname)
        mendActData = [row for row in arcpy.da.SearchCursor(mendActivity, mendActCol, where_clause=mendActWhere)]
        mendActDF = pd.DataFrame(mendActData, columns=mendActCol) 
        
        if (mendTripDF.empty == False):
            if (mendActDF.empty == False):
                mendDF = mendActDF.merge(mendTripDF, how='inner', on = 'TRIP_GUID')
        
        """Creating Outfitting data frame in the date range and for the specific business name"""

        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND BUSINESSNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), businessname)
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)
        
               

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames    
        
        """ Activity Dictionary to account for differences in the activity categories and what is listed in each of the datasets. """
        
        activities = {'Boating, Stand Up Paddle Boarding, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Pack Rafting, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 
           'Boating, Raft, Canoe, Kayak or Other Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Canoeing (Mendenhall form)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Rafting (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Canoeing':'Boating (Non-Motorized, Freshwater)', 'Boating (Non-Motorized, Freshwater)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking':'Boating (Non-Motorized, Freshwater)', 'Rafting':'Boating (Non-Motorized, Freshwater)', 'Camping':'Camping','Dog Sledding':'Dog Sled Tours', 'Dog Sled Tours':'Dog Sled Tours', 'Flightseeing Landing Tours':'Flightseeing Landing Tours',
           'Freshwater Fishing':'Freshwater Fishing', 'Glacier Trekking':'Helicopter Landing Tours', 'Helicopter Landing Tours':'Helicopter Landing Tours', 'Heli-skiing Tours':'Heli-skiing Tours', 
           'Heliski':'Heli-skiing Tours', 'Hunting, Brown Bear':'Hunting, Brown Bear', 'Hunting, Deer':'Hunting, Deer', 'Remote Setting Nature Tour, on Foot': 'Remote Setting Nature Tour', 
           'Hunting, Elk':'Hunting, Elk', 'Hunting, Moose':'Hunting, Moose', 'Hunting, Mountain Goat':'Hunting, Mountain Goat', 'Remote Setting Nature Tour':'Remote Setting Nature Tour',
           'Hunting, Wolf':'Hunting, Waterfowl/Small game/Wolf - Service Day', 'Outfitting':'Outfitting (Delivery and/or pick-up of vehicles, equipment, etc. to/from National Forest System lands; Total per day - no limit on equipment numbers or number of trips)', 
           'Over-Snow Vehicle Tours':'Over-Snow Vehicle Tours', 'Bikepacking':'Remote Setting Nature Tour', 'Horseback Riding':'Remote Setting Nature Tour', 'Nature Tour, on Foot':'Remote Setting Nature Tour',
           'Nature Tour, on Ski':'Remote Setting Nature Tour','Nature Tour, Bicycle':'Remote Setting Nature Tour', 'Biking (Mendenhall form)':'Remote Setting Nature Tour', 'Biking':'Remote Setting Nature Tour', 
           'Hiking (Mendenhall form)':'Remote Setting Nature Tour', 'Hiking':'Remote Setting Nature Tour', 'Nature Tour, ATV/OHV':'Road Based Nature Tours', 'Nature Tour, Vehicle':'Road Based Nature Tours', 
           'Visitor Center (Begich Boggs, MGVC, SEADC)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Visitor Center Transport (Mendenhall form)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 
           'Visitor_Center_Transport':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Hunting, Black Bear ':'Hunting, Black Bear', 'Hunting, Dall Sheep ':'Hunting, Dall Sheep', 'Hunting, Waterfowl/Small game ':'Hunting, Waterfowl/Small game/Wolf - Service Day',
           'Assigned Site' : 'Assigned Site', 'Minimum Fee': 'Minimum Fee', 'Hunting, Black Bear':'Hunting, Black Bear', 'Hunting, Waterfowl/Small game/Wolf - Service Day Rate':'Hunting, Waterfowl/Small game/Wolf - Service Day'}
  

        """Creating Location dataframe. This will be joined to all the other dataframes so that we can get ancillary data, like ranger district"""
    
        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['FULLNAME'].fillna(locDF['ZONE_NAME'], inplace=True)
        locDF['USE_AREA'] = locDF['FULLNAME']
        locDF['USE_AREA'].fillna('Use Area Unknown', inplace=True)
        locDF['USELOCATION'].str.upper()
        locDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
        
        """ This takes the start year and end year parameters and reformats it so that it matches the date format in Guided Recreation and Hunting."""
        
        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
        
        
            
        """ Updating the Icefields dataframe if it exists. If it exists, we have to find out what the activities were. There are documenmented in seperate columns with a 0 or 1. 
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Icefield data, it is the total number of clients by day. 
        When these are grouped, there can be some duplication. When sorting by activity, if the same group of people do two different activities on the same day, they are counted twice, once for each
        activity. If a group do two of the same activities in two differet locations, they will only be counted once in the activity tale, but will be counted twice in the location group. 
        """


        if 'iceDF' not in locals():
            arcpy.AddMessage('This business had no icefield activities for this year.')
        else:
            iceDF.loc[iceDF['LDNGOPS'] >0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['LDNGGRATUITY']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['LDNGPAIDCLIENTS']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['CLIENTSGLACTREK']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['CLIENTSDOGSLED']>0, 'ACTIVITY'] = 'Dog Sled Tours'
            iceDF.loc[iceDF['CLIENTSHIKE']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF['Activity2'] = iceDF['ACTIVITY'].map(activities)

            iceDF['USELOCATION'].str.upper()
            iceLoc = iceDF.merge(locDF, how='left', left_on = ['LOCATION_ID'] , right_on = ['LOCATION_ID'])
            iceLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
            iceLoc['Year'] = iceLoc['REPORTYEAR']

            iceLoc['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            iceLoc['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            iceLoc['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)

            iceLoc['LandTourTotal'] = iceLoc['CLIENTSGLACTREK'] + iceLoc['CLIENTSHIKE']
            iceLoc['LandTourTotal'].fillna(0, axis=0, inplace=True)
            iceLoc['DogSledTotal'] = iceLoc['CLIENTSDOGSLED']
            iceLoc['DogSledTotal'].fillna(0, axis=0, inplace=True)
            iceLoc['ice_CLIENTNUMBER'] = iceLoc['LandTourTotal']  + iceLoc['DogSledTotal']

            iceActGroup = iceLoc.groupby(['TRIP_GUID', 'Activity2'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'ice_CLIENTNUMBER': 'sum'})
            iceActGroup.loc[iceActGroup['ice_CLIENTNUMBER'] > iceActGroup['CLIENTMONTH'], 'ice_SERVICE_DAYS'] = iceActGroup['CLIENTMONTH']
            iceActGroup['ice_SERVICE_DAYS'].fillna(iceActGroup['ice_CLIENTNUMBER'], axis = 0, inplace = True)

            iceUseGroup = iceLoc.groupby(['TRIP_GUID', 'USE_AREA'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'ice_CLIENTNUMBER': 'sum'})
            iceUseGroup.loc[iceUseGroup['ice_CLIENTNUMBER'] > iceUseGroup['CLIENTMONTH'], 'ice_SERVICE_DAYS'] = iceUseGroup['CLIENTMONTH']
            iceUseGroup['ice_SERVICE_DAYS'].fillna(iceUseGroup['ice_CLIENTNUMBER'], axis = 0, inplace = True)            
 
            iceLocGroup = iceLoc.groupby(['TRIP_GUID', 'USELOCATION'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'ice_CLIENTNUMBER': 'sum'})
            iceLocGroup.loc[iceLocGroup['ice_CLIENTNUMBER'] > iceLocGroup['CLIENTMONTH'], 'ice_SERVICE_DAYS'] = iceLocGroup['CLIENTMONTH']
            iceLocGroup['ice_SERVICE_DAYS'].fillna(iceLocGroup['ice_CLIENTNUMBER'], axis = 0, inplace = True)            
            
            
            iceActivity = iceActGroup[['TRIP_GUID', 'Activity2', 'ice_SERVICE_DAYS', 'DISTRICTNAME', 'Year']].copy()
            iceUse = iceUseGroup[['TRIP_GUID', 'USE_AREA', 'ice_SERVICE_DAYS', 'DISTRICTNAME', 'Year']].copy()
            iceLocName = iceLocGroup[['TRIP_GUID', 'USELOCATION', 'ice_SERVICE_DAYS', 'DISTRICTNAME', 'Year']].copy()  
            arcpy.AddMessage("Icefields DataFrame Created")

        """ Updating the Guided Recreation dataframe if it exists. If it exists, we capitalize all the Use Location data to accomidate the different formats that this data comes in as. 
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Guided Recreation data, it is the total number of clients by day. 
        When these are grouped, there can be some duplication. When sorting by activity, if the same group of people do two different activities on the same day, they are counted twice, once for each
        activity. If a group do two of the same activities in two differet locations, they will only be counted once in the activity table, but will be counted twice in the location group. To minimize
        the double counting, we compare the total number of people by activity, location, or Use Area, and if the sum of total clients is greater than the total clients on the day, then we use the 
        total clients on the day field for the service days. 
        """

        if 'guideDF' not in locals():
            arcpy.AddMessage('This business had no guided recreation activities for this year.')   
        else:
            if (guideDF['USELOCATION'].loc() == guideDF['USELOCATION'].loc()):
                guideDF['USELOCATION'] = guideDF.loc(guideDF['USELOCATION'].map(lambda x: x.title()))
            guideDF['Activity2'] = guideDF['ACTIVITY'].map(activities)
            guideDF['Year'] = guideDF['TRIPDATE'].dt.year

            
            guideDF['USELOCATION'].str.upper()
            guideLoc = guideDF.merge(locDF, how = 'left', on = ['LOCATION_ID'])

            guideLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
            guideLoc['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            guideLoc['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            guideLoc['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)


            guideActGroup = guideLoc.groupby(['DAY_GUID', 'Activity2'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
            guideActGroup.loc[guideActGroup['CLIENTNUMBER'] > guideActGroup['TOTALCLIENTSONDAY'], 'Guide_SERVICE_DAYS'] = guideActGroup['TOTALCLIENTSONDAY']
            guideActGroup['Guide_SERVICE_DAYS'].fillna(guideActGroup['CLIENTNUMBER'], axis = 0, inplace = True)

            guideUseGroup = guideLoc.groupby(['DAY_GUID', 'USE_AREA'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
            guideUseGroup.loc[guideUseGroup['CLIENTNUMBER'] > guideUseGroup['TOTALCLIENTSONDAY'], 'Guide_SERVICE_DAYS'] = guideUseGroup['TOTALCLIENTSONDAY']
            guideUseGroup['Guide_SERVICE_DAYS'].fillna(guideUseGroup['CLIENTNUMBER'], axis = 0, inplace = True)            

            guideLocGroup = guideLoc.groupby(['DAY_GUID', 'USELOCATION'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
            guideLocGroup.loc[guideLocGroup['CLIENTNUMBER'] > guideLocGroup['TOTALCLIENTSONDAY'], 'Guide_SERVICE_DAYS'] = guideLocGroup['TOTALCLIENTSONDAY']
            guideLocGroup['Guide_SERVICE_DAYS'].fillna(guideLocGroup['CLIENTNUMBER'], axis = 0, inplace = True)            
            guideAct = guideActGroup[['Activity2', 'DISTRICTNAME', 'Year', 'Guide_SERVICE_DAYS']].copy()
            guideUse = guideUseGroup[['USE_AREA', 'DISTRICTNAME', 'Year', 'Guide_SERVICE_DAYS']].copy()
            guideLocName = guideLocGroup[['USELOCATION', 'DISTRICTNAME', 'Year', 'Guide_SERVICE_DAYS']].copy()

            
            arcpy.AddMessage("GuidedRec DataFrame Created")            
                

        """ In the hunting data, there are hunts, and non-hunting activities. This next set of code is only handeling the non-hunting activities.Updating the Hunting dataframe if it exists. 
        If it exists, we capitalize all the Use Location data to accomidate the different formats that this data comes in as.  
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Hunting data, it is the total number of clients by day. 
        When these are grouped, there can be some duplication. When sorting by activity, if the same group of people do two different activities on the same day, they are counted twice, once for each
        activity. If a group do two of the same activities in two differet locations, they will only be counted once in the activity table, but will be counted twice in the location group. To minimize
        the double counting, we compare the total number of people by activity, location, or Use Area, and if the sum of total clients is greater than the total clients on the day, then we use the 
        total clients on the day field for the service days. For the Activity tab, we only want activites that are not hunting. Those hunting activites are captured in the hunt tab. But we still 
        need to capture the hunters for Use Area and Location, that is why nonHuntActOnly is only used for the activity group by. It has filtered out all hunting activities. 
        """

        
        if 'huntDF' not in locals():
            if 'nonHuntDF' not in locals():
                arcpy.AddMessage('This business had no hunting associated activities for this year.')   
        else:
            huntDF['Activity2'] = huntDF['ACTIVITY'].map(activities)
            huntDF['Year'] = huntDF['TRIPDATE'].dt.year           
            huntDF['USELOCATION'].str.upper()

            huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID']) 
            huntLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
            hunts = huntLoc[['Year', 'Activity2', 'HUNTERS', 'USE_AREA', 'USELOCATION', 'DISTRICTNAME', 'DAY_GUID','TOTALCLIENTSONDAY']].copy()
            
            if 'nonHuntDF' in locals():
                nonHuntDF['Activity2'] = 'Remote Setting Nature Tour'            
                nonHuntDF['Year'] = nonHuntDF['TRIPDATE'].dt.year            
                nonHuntDF['USELOCATION'].str.upper()
                nonHuntLoc = nonHuntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
                nonHuntLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
                nonHuntOnly = nonHuntLoc[['Year', 'Activity2', 'SERVICE_DAYS_NONHUNTER', 'USE_AREA', 'USELOCATION', 'DISTRICTNAME', 'DAY_GUID','TOTALCLIENTSONDAY']].copy()            
                huntAll = hunts.append(nonHuntOnly)
            else: huntAll = hunts
            
            huntAll['Total Clients'] = huntAll['HUNTERS']
            if 'nonHuntDF' in locals():
                huntAll['Total Clients'].fillna(huntAll['SERVICE_DAYS_NONHUNTER'], inplace = True)           
            
            
            huntAll['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            huntAll['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            huntAll['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
            
            nonHuntActOnly = huntAll[huntAll['Activity2'].str.contains("Hunting") == False]
            
            huntActGroup = nonHuntActOnly.groupby(['DAY_GUID', 'Activity2'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'Total Clients': 'sum'})
            huntActGroup.loc[huntActGroup['Total Clients'] > huntActGroup['TOTALCLIENTSONDAY'], 'Hunting_SERVICE_DAYS'] = huntActGroup['TOTALCLIENTSONDAY']
            huntActGroup['Hunting_SERVICE_DAYS'].fillna(huntActGroup['Total Clients'], axis = 0, inplace = True)

            huntUseGroup = huntAll.groupby(['DAY_GUID', 'USE_AREA'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'Total Clients': 'sum'})
            huntUseGroup.loc[huntUseGroup['Total Clients'] > huntUseGroup['TOTALCLIENTSONDAY'], 'Hunting_SERVICE_DAYS'] = huntUseGroup['TOTALCLIENTSONDAY']
            huntUseGroup['Hunting_SERVICE_DAYS'].fillna(huntUseGroup['Total Clients'], axis = 0, inplace = True)            

            huntLocGroup = huntAll.groupby(['DAY_GUID', 'USELOCATION'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'Total Clients': 'sum'})
            huntLocGroup.loc[huntLocGroup['Total Clients'] > huntLocGroup['TOTALCLIENTSONDAY'], 'Hunting_SERVICE_DAYS'] = huntLocGroup['TOTALCLIENTSONDAY']
            huntLocGroup['Hunting_SERVICE_DAYS'].fillna(huntLocGroup['Total Clients'], axis = 0, inplace = True)            
  
            huntAct1 = huntActGroup[['Activity2', 'DISTRICTNAME', 'Year', 'Hunting_SERVICE_DAYS']].copy()
            huntUse = huntUseGroup[['USE_AREA', 'DISTRICTNAME', 'Year', 'Hunting_SERVICE_DAYS']].copy()
            huntLocName = huntLocGroup[['USELOCATION', 'DISTRICTNAME', 'Year', 'Hunting_SERVICE_DAYS']].copy()
            
            arcpy.AddMessage("Hunting Associated Activities DataFrame Created")
                 
                
                
        """ In the hunting data, there are hunts, and non-hunting activities. This next set of code is only handeling hunting activities. We are not working with service days with this data. 
            We are just counting the number of hunts that take place. The exception to that is small game and wolves. For those species, we do use Service days, as the total number of clients
            by day that do this activity. These totals are grouped into Hunting Service Days and refelcted in the activity tab for use locations. The hunt counts
            will all be calculated on the Hunt tab. We do not join to the location table, because of the heiarchy of the hunting tables, the hunter table is a child of the trip table, but it does not have a realationship to the 
            stop table, which is where all the location data is saved. We have to toatl all the hunts by species first. Then we handle the species as the "activity" and totaly all the hunt counts 
            in the hunts column. This pivot below is what is exported to excel. 
        """
        
        
        if 'huntCountDF' not in locals():
            arcpy.AddMessage('This business had no hunts for this year.')
        else:

            huntCountDF['Year'] = huntCountDF['ENDDATE'].dt.year
            huntCountDF['DISTRICTNAME'] = 'All Districts'
                     

            blkBear = huntCountDF[['TRIP_GUID', 'BLACKBEAR', 'Year', 'DISTRICTNAME']]
            blkBear.loc[blkBear['BLACKBEAR']>0, 'Activity2'] = 'Hunting, Black Bear'

            brwnBear = huntCountDF[['TRIP_GUID', 'BROWNBEAR', 'Year', 'DISTRICTNAME']]
            brwnBear.loc[brwnBear['BROWNBEAR']>0, 'Activity2'] = 'Hunting, Brown Bear'
            
            dallSheep = huntCountDF[['TRIP_GUID', 'DALLSHEEP', 'Year', 'DISTRICTNAME']]
            dallSheep.loc[dallSheep['DALLSHEEP']>0, 'Activity2'] = 'Hunting, Dall Sheep' 

            deer = huntCountDF[['TRIP_GUID', 'DEER', 'Year', 'DISTRICTNAME']]  
            deer.loc[deer['DEER']>0, 'Activity2'] = 'Hunting, Deer'

            elk = huntCountDF[['TRIP_GUID', 'ELK', 'Year', 'DISTRICTNAME']]
            elk.loc[elk['ELK']>0, 'Activity2'] = 'Hunting, Elk'
            
            moose = huntCountDF[['TRIP_GUID', 'MOOSE', 'Year', 'DISTRICTNAME']]
            moose.loc[moose['MOOSE']>0, 'Activity2'] = 'Hunting, Moose'
            
            goat = huntCountDF[['TRIP_GUID', 'MOUNTAINGOAT', 'Year', 'DISTRICTNAME']]
            goat.loc[goat['MOUNTAINGOAT']>0, 'Activity2'] = 'Hunting, Mountain Goat'
            
            smGame = huntCountDF[['TRIP_GUID', 'DAYS_WTRFOWLSMALLGAME', 'Year', 'DISTRICTNAME']]
            smGame.loc[smGame['DAYS_WTRFOWLSMALLGAME']>0, 'Activity2'] = 'Hunting, Waterfowl/Small game/Wolf - Service Day'
            smGame.rename(columns= {'DAYS_WTRFOWLSMALLGAME':'Hunting_SERVICE_DAYS'}, inplace = True)
            
            wolf = huntCountDF[['TRIP_GUID', 'DAYS_WOLF', 'Year', 'DISTRICTNAME']]
            wolf.loc[wolf['DAYS_WOLF']>0, 'Activity2'] = 'Hunting, Waterfowl/Small game/Wolf - Service Day'
            wolf.rename(columns= {'DAYS_WOLF':'Hunting_SERVICE_DAYS'}, inplace = True)
            
            hunting = blkBear.append([brwnBear, deer, dallSheep, moose, elk, goat])
            
            if 'huntAct1' in locals():
                huntAct = huntAct1.append([smGame, wolf])
            else:
                huntAct = smGame.append(wolf)

            
            columns = ['BLACKBEAR', 'BROWNBEAR','DALLSHEEP', 'DEER', 'ELK', 'MOOSE', 'MOUNTAINGOAT']
            
            hunting['HUNTS'] = hunting[columns].sum(axis = 1)
            
            hunting1 = hunting[['TRIP_GUID', 'Year', 'DISTRICTNAME', 'HUNTS', 'Activity2']]

            hunting1.rename(columns={'Activity2':'Activity', 'DISTRICTNAME':'Ranger District'}, inplace=True)
           
            
            huntActPivot = pd.pivot_table(hunting1, index = ['Activity', 'Ranger District'], columns = ['Year'], values = ['HUNTS'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')

            arcpy.AddMessage("Hunting DataFrame Created")

        """ Updating the Heliski dataframe if it exists. 
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Heliski data, it is the total number of clients by day. 
        When these are grouped, there can be some duplication. When sorting by activity, if the same group of people do two different activities on the same day, they are counted twice, once for each
        activity. If a group do two of the same activities in two differet locations, they will only be counted once in the activity tale, but will be counted twice in the location group.
        To minimize the double counting, we compare the total number of people by activity, location, or Use Area, and if the sum of total clients is greater than the total clients on the day, then we use the 
        total clients on the day field for the service days. 
        """
        
        
        if 'heliDF' not in locals():
            arcpy.AddMessage('This business had no heliski activities for this year.') 
        else:        
            heliDF['Activity2'] = heliDF['ACTIVITY'].map(activities)
            heliDF['Year'] = heliDF['TRIPDATE'].dt.year
            heliDF['USELOCATION'].str.upper()
            heliLoc = heliDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
            heliLoc.rename(columns = {'USELOCATION_x':'USELOCATION'}, inplace=True)
            heliLoc['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            heliLoc['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            heliLoc['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
            heliActGroup = heliLoc.groupby(['TRIP_GUID', 'Activity2'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTS_LOCATION': 'sum'})
            heliActGroup.loc[heliActGroup['CLIENTS_LOCATION'] > heliActGroup['TOTALCLIENTSONDAY'], 'Heli_SERVICE_DAYS'] = heliActGroup['TOTALCLIENTSONDAY']
            heliActGroup['Heli_SERVICE_DAYS'].fillna(heliActGroup['CLIENTS_LOCATION'], axis = 0, inplace = True)

            heliUseGroup = heliLoc.groupby(['TRIP_GUID', 'USE_AREA'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTS_LOCATION': 'sum'})
            heliUseGroup.loc[heliUseGroup['CLIENTS_LOCATION'] > heliUseGroup['TOTALCLIENTSONDAY'], 'Heli_SERVICE_DAYS'] = heliUseGroup['TOTALCLIENTSONDAY']
            heliUseGroup['Heli_SERVICE_DAYS'].fillna(heliUseGroup['CLIENTS_LOCATION'], axis = 0, inplace = True)            

            heliLocGroup = heliLoc.groupby(['TRIP_GUID', 'USELOCATION'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTS_LOCATION': 'sum'})
            heliLocGroup.loc[heliLocGroup['CLIENTS_LOCATION'] > heliLocGroup['TOTALCLIENTSONDAY'], 'Heli_SERVICE_DAYS'] = heliLocGroup['TOTALCLIENTSONDAY']
            heliLocGroup['Heli_SERVICE_DAYS'].fillna(heliLocGroup['CLIENTS_LOCATION'], axis = 0, inplace = True)            


            heliAct = heliActGroup[['TRIP_GUID', 'Activity2', 'DISTRICTNAME', 'Year', 'Heli_SERVICE_DAYS']].copy()
            heliUse = heliUseGroup[['TRIP_GUID', 'USE_AREA', 'DISTRICTNAME', 'Year', 'Heli_SERVICE_DAYS']].copy()
            heliLocName = heliLocGroup[['TRIP_GUID', 'USELOCATION', 'DISTRICTNAME', 'Year', 'Heli_SERVICE_DAYS']].copy()
        
            arcpy.AddMessage("Heliski DataFrame Created")          
        
        """ Updating the Mendenhall dataframe if it exists. If it exists, we have to find out what the activities were. There are documenmented in seperate columns with a 0 or 1.
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Mendenhall data, it is the total number of clients by day. 
        When these are grouped, there can be some duplication. When sorting by activity, if the same group of people do two different activities on the same day, they are counted twice, once for each
        activity. If a group do two of the same activities in two differet locations, they will only be counted once in the activity tale, but will be counted twice in the location group.
        To minimize the double counting, we compare the total number of people by activity, location, or Use Area, and if the sum of total clients is greater than the total clients on the day, then we use the 
        total clients on the day field for the service days. 
        """


        if 'mendDF' not in locals():
            arcpy.AddMessage('This business had no mendenhall activities for this year.')      
        else:            
            mendDF.loc[mendDF['BIKING']>0, 'ACTIVITY'] = 'Remote Setting Nature Tour'
            mendDF.loc[mendDF['HIKING']>0, 'ACTIVITY'] = 'Remote Setting Nature Tour'
            mendDF.loc[mendDF['CANOEING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['KAYAKING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['RAFTING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['VCTRANSPORT']>0, 'ACTIVITY'] = 'Visitor Center (Begich Boggs, MGVC, SEADC)'
            
            mendDF.rename(columns = {'USELOCATION':'USELOCATION'}, inplace=True)
            mendDF['Activity2'] = mendDF['ACTIVITY'].map(activities)
            mendDF['USELOCATION'].str.upper()
            mendLoc = mendDF.merge(locDF, how='left', left_on= ['LOCATION_ID'], right_on= ['LOCATION_ID'])
            mendLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)


            mendLoc['Year'] = mendLoc['REPORTYEAR']
            mendLoc['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            mendLoc['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            mendLoc['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)   
            mendLoc['Activity2'].fillna('Activity Unknown', inplace = True)
            
            mendActGroup = mendLoc.groupby(['TRIP_GUID', 'Activity2'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'CLIENTSLOCATION': 'sum'})
            mendActGroup.loc[mendActGroup['CLIENTSLOCATION'] > mendActGroup['CLIENTMONTH'], 'mend_SERVICE_DAYS'] = mendActGroup['CLIENTMONTH']
            mendActGroup['mend_SERVICE_DAYS'].fillna(mendActGroup['CLIENTSLOCATION'], axis = 0, inplace = True)

            mendUseGroup = mendLoc.groupby(['TRIP_GUID', 'USE_AREA'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'CLIENTSLOCATION': 'sum'})
            mendUseGroup.loc[mendUseGroup['CLIENTSLOCATION'] > mendUseGroup['CLIENTMONTH'], 'mend_SERVICE_DAYS'] = mendUseGroup['CLIENTMONTH']
            mendUseGroup['mend_SERVICE_DAYS'].fillna(mendUseGroup['CLIENTSLOCATION'], axis = 0, inplace = True)            
 
            mendLocGroup = mendLoc.groupby(['TRIP_GUID', 'USELOCATION'], as_index = False).agg({'Year': 'first', 'DISTRICTNAME' : 'first', 'CLIENTMONTH': 'max', 'CLIENTSLOCATION': 'sum'})
            mendLocGroup.loc[mendLocGroup['CLIENTSLOCATION'] > mendLocGroup['CLIENTMONTH'], 'mend_SERVICE_DAYS'] = mendLocGroup['CLIENTMONTH']
            mendLocGroup['mend_SERVICE_DAYS'].fillna(mendLocGroup['CLIENTSLOCATION'], axis = 0, inplace = True)            

            
            mendAct = mendActGroup[['TRIP_GUID', 'Activity2', 'DISTRICTNAME', 'Year', 'mend_SERVICE_DAYS']].copy()
            mendUse = mendUseGroup[['TRIP_GUID', 'USE_AREA', 'DISTRICTNAME', 'Year', 'mend_SERVICE_DAYS']].copy()
            mendLocName = mendLocGroup[['TRIP_GUID', 'USELOCATION', 'DISTRICTNAME', 'Year', 'mend_SERVICE_DAYS']].copy()
            
            arcpy.AddMessage("Mendenhall DataFrame Created")
        


        """ Updating the Outfitting dataframe if it exists. 
        Then the activity dictionary is applied and the dataframe is joined to the location dataframe. To figure out service days for Outfitting data, each day is one service day. It does not 
        matter how many people were there. 
        """


        if (outfitDF.empty == True):
            arcpy.AddMessage('This business had no outfitting activities for this year.')  
        else:
            outfitDF['Activity2'] = outfitDF['ACTIVITY'].map(activities)
            outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
            outfitDF['USELOCATION'].str.upper()
            outfitLoc = outfitDF.merge(locDF, how = 'left', on = ['LOCATION_ID'])
            outfitLoc['Outfitting_SERVICE_DAYS'] = 1
            outfitLoc['USELOCATION'].fillna('Use Location Other or Unknown', inplace = True)
            outfitLoc['USE_AREA'].fillna('Use Area Unknown', inplace=True)
            outfitLoc['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)            
            outfitAct = outfitLoc[['Activity2', 'DISTRICTNAME', 'Year', 'Outfitting_SERVICE_DAYS']].copy()
            outfitUse = outfitLoc[['USE_AREA', 'DISTRICTNAME', 'Year', 'Outfitting_SERVICE_DAYS']].copy()
            outfitLocName = outfitLoc[['USELOCATION', 'DISTRICTNAME', 'Year', 'Outfitting_SERVICE_DAYS']].copy()
            
            arcpy.AddMessage("Outfitting DataFrame Created")          
            
        """ If no dataframes are created, this means there is no data for that business in that date range and the tool exits."""

        if 'iceDF' not in locals():  
             if 'guideDF' not in locals():
                 if 'heliDF' not in locals():
                     if 'mendDF' not in locals():
                         if (outfitDF.empty == True):
                             if 'huntDF' not in locals():
                                 arcpy.AddMessage('This business has no activities between the start and end date.')
                                 exit()
                             
        
        def listYear(r1, r2):
            return [item for item in range(r1, r2+1)]
        
        yearStart = int(startYear)
        yearEnd = int(endYear)
        years = listYear(yearStart, yearEnd)


        """This step in the process is combining all of the individual dataframes into one large dataframe. Because we don't know if there is data in each dataframe, we have to check each one first.
        """

        if 'guideAct' in locals():
            if 'heliAct' in locals():
                heliGuideAct = guideAct.append(heliAct)
                arcpy.AddMessage("Heliski and Guided Rec Merged DataFrame Created")                
            else:   
                heliGuideAct = guideAct

                arcpy.AddMessage("Heliski Merged DataFrame Created")                
        elif 'heliAct' in locals():
            heliGuideAct = heliAct
           
            arcpy.AddMessage("Guided Rec Merged DataFrame Created")
        
        
        if 'iceActivity' in locals():
            if 'mendAct' in locals():
                iceMendAct = iceActivity.append(mendAct)
                arcpy.AddMessage("Icefields and Mendenhall Merged DataFrame Created")
            else:   
                iceMendAct = iceActivity

                arcpy.AddMessage("Icefields Merged DataFrame Created")              
        elif 'mendAct' in locals():
            iceMendAct = mendAct
            
            arcpy.AddMessage("Mendenhall Merged DataFrame Created")            
        
        if 'huntAct' in locals():
            if 'outfitAct' in locals():
                huntOutAct = huntAct.append(outfitAct)
                arcpy.AddMessage("Hunting and Outfitting Merge DataFrame Created")
            else:   
                huntOutAct = huntAct

                arcpy.AddMessage("Hunting Merge DataFrame Created")
        elif 'outfitAct' in locals():
            huntOutAct = outfitAct 
          
            arcpy.AddMessage("Outfitting Merge DataFrame Created")   
             

        if 'heliGuideAct' in locals():
            if 'iceMendAct' in locals():
                hGIMDF_Act = heliGuideAct.append(iceMendAct)
            else:   
                hGIMDF_Act = heliGuideAct
 
        elif 'iceMendAct' in locals():
            hGIMDF_Act = iceMendAct

            
        if 'huntOutAct' in locals():
            if 'hGIMDF_Act' in locals():
                actAllDF = huntOutAct.append(hGIMDF_Act)
            else:   
                actAllDF = huntOutAct

        elif 'hGIMDF_Act' in locals():
            actAllDF = hGIMDF_Act 
  
        
        dfYears = actAllDF['Year'].tolist()
        otherYears = list(set(years).difference(dfYears)) 
        
        
        addYears = pd.DataFrame(otherYears, columns = ['Year'])
        actAllDF = actAllDF.append(addYears)        


        actAllDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace=True)   
        actAllDF['Activity2'].dropna(axis = 0, how='any', inplace=True)

        columns = ['Guide_SERVICE_DAYS', 'Outfitting_SERVICE_DAYS', 'ice_SERVICE_DAYS', 'mend_SERVICE_DAYS', 'Heli_SERVICE_DAYS', 'Hunting_SERVICE_DAYS']
        for x in columns: 
            if x not in actAllDF.columns:
                actAllDF[x] = np.nan
                
        actAllDF['SERVICE DAYS'] = actAllDF[columns].sum(axis=1)
        actAllDF = actAllDF.replace(0,np.nan)
                
        actAllDF.rename(columns ={'Activity2':'Activity', 'DISTRICTNAME':'Ranger District'}, inplace = True)

        
        arcpy.AddMessage("All Activities DataFrame Created") 
        

        if 'guideUse' in locals():
            if 'heliUse' in locals():
                heliGuideUse = guideUse.append(heliUse)
                arcpy.AddMessage("Heliski and Guided Rec Use Merged DataFrame Created")
            else:   
                heliGuideUse = guideUse

                arcpy.AddMessage("Guided Rec Use Merged DataFrame Created")
        elif 'heliUse' in locals():
            heliGuideUse = heliUse 

            arcpy.AddMessage("Heliski Use Merged DataFrame Created")  
            
        if 'iceUse' in locals():
            if 'mendUse' in locals():
                iceMendUse = iceUse.append(mendUse) 
                arcpy.AddMessage("Icefields and Mendenhall Use Merge DataFrame Created")
            else:   
                iceMendUse = iceUse

                arcpy.AddMessage("Icefields Use Merge DataFrame Created")                
        elif 'mendAct' in locals():
            iceMendUse = mendUse
            
            arcpy.AddMessage("Mendenhall Use Merge DataFrame Created")  
            
        if 'huntUse' in locals():
            if 'outfitUse' in locals():
                huntOutUse = huntUse.append(outfitUse)
                arcpy.AddMessage("Hunting and Outfitting Merge DataFrame Created")
            else:   
                huntOutUse = huntUse
 
                arcpy.AddMessage("Hunting Use Merge DataFrame Created")
        elif 'outfitUse' in locals():
            huntOutUse = outfitUse  
            
            arcpy.AddMessage("Outfitting Use Merge DataFrame Created")           

        if 'heliGuideUse' in locals():
            if 'iceMendUse' in locals():
                hGIMDF_Use = heliGuideUse.append(iceMendUse)
            else:   
                hGIMDF_Use = heliGuideUse
                
        elif 'iceMendUse' in locals():
            hGIMDF_Use = iceMendUse
  
            
        if 'huntOutUse' in locals():
            if 'hGIMDF_Use' in locals():
                useAllDF = huntOutUse.append(hGIMDF_Use)
            else:   
                useAllDF = huntOutUse

        elif 'hGIMDF_Use' in locals():
            useAllDF = hGIMDF_Use

            
        useAllDF = useAllDF.append(addYears)
            

        useAllDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace=True)   

        for x in columns: 
            if x not in useAllDF.columns:
                useAllDF[x] = np.nan
                      
        
        useAllDF['SERVICE DAYS - ALL ACTIVITIES'] = useAllDF[columns].sum(axis=1)
        useAllDF = useAllDF.replace(0,np.nan)         
        
        useAllDF.rename(columns ={'USE_AREA':'Use Area', 'DISTRICTNAME':'Ranger District'}, inplace = True)
        
        arcpy.AddMessage("All Use Areas DataFrame Created")          

        if 'guideLocName' in locals():
            if 'heliLocName' in locals():
                heliGuideLocName = guideLocName.append(heliLocName)       
                arcpy.AddMessage("Heliski And Guided Rec Location Merge DataFrame Created")
            else:   
                heliGuideLocName = guideLocName

                arcpy.AddMessage("Guided Rec Location Merge DataFrame Created") 
        elif 'heliLocName' in locals():
            heliGuideLocName = heliLocName

            arcpy.AddMessage("Heliski Location Merge DataFrame Created")       

        if 'iceLocName' in locals():
            if 'mendLocName' in locals():
                iceMendLocName = iceLocName.append(mendLocName)
                arcpy.AddMessage("Icefields and Mendenhall Location Merge DataFrame Created")
            else:   
                iceMendLocName = iceLocName

                arcpy.AddMessage("Icefields Location Merge DataFrame Created")
        elif 'mendLocName' in locals():
            iceMendLocName = mendLocName 

            arcpy.AddMessage("Mendenhall Location Merge DataFrame Created")

            
        if 'huntLocName' in locals():
            if 'outfitLocName' in locals():
                huntOutLocName = huntLocName.append(outfitLocName, ignore_index=True)
                arcpy.AddMessage("Hunting and Outfitting Merge DataFrame Created")
            else:   
                huntOutLocName = huntLocName
               
                arcpy.AddMessage("Hunting Location Merge DataFrame Created")
        elif 'outfitLocName' in locals():    
            huntOutLocName = outfitLocName   
           
            arcpy.AddMessage("Outfitting Location Merge DataFrame Created")            
            
        if 'heliGuideLocName' in locals():
            if 'iceMendLocName' in locals():
                hGIMDF_LocName = heliGuideLocName.append(iceMendLocName)
            else:   
                hGIMDF_LocName = heliGuideLocName

        elif 'iceMendLocName' in locals():
            hGIMDF_LocName = iceMendLocName

            
        if 'huntOutLocName' in locals():
            if 'hGIMDF_LocName' in locals():
                locNameAllDF = huntOutLocName.append(hGIMDF_LocName, ignore_index=True)
            else:   
                locNameAllDF = huntOutLocName
               
        elif 'hGIMDF_LocName' in locals():
            locNameAllDF = hGIMDF_LocName 


        locNameAllDF = locNameAllDF.append(addYears)            

       
        locNameAllDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace=True)
        
        
        for x in columns: 
            if x not in locNameAllDF.columns:
                locNameAllDF[x] = np.nan

        locNameAllDF['SERVICE DAYS - ALL ACTIVITIES'] = locNameAllDF[columns].sum(axis=1)
        locNameAllDF = locNameAllDF.replace(0,np.nan)
                
        locNameAllDF.rename(columns ={'USELOCATION':'Use Location', 'DISTRICTNAME':'Ranger District'}, inplace = True)
        

        arcpy.AddMessage("All Locaitons Merge DataFrame Created")

        """Create three pivot tables that are added to the activity, use area, and location tabs""" 

        actAllPivot = pd.pivot_table(actAllDF, index = ['Activity', 'Ranger District'], columns = ['Year'], values = ['SERVICE DAYS'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')

        useAllPivot = pd.pivot_table(useAllDF, index = ['Use Area'], columns = ['Year'], values = ['SERVICE DAYS - ALL ACTIVITIES'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')

        locNameAllPivot = pd.pivot_table(locNameAllDF, index = ['Ranger District', 'Use Location'], columns = ['Year'], values = ['SERVICE DAYS - ALL ACTIVITIES'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')

        

        """ Writing pivot tables to excel and additional formatting. The activity tab and hunt tab have to be added with conditional statements because not all businesses will have hunts, and some
        businesses that have hunts do not have other activities.
        """

        writer_args = {
            'path': self.path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='left', wrapText = True, vertical='center')
        index_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        title_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def update(data, startNumb, columnNumb, ws):
            for row, text in enumerate(data, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)
                
        def headers(ws):
            header1 = ' Outfitter/Guide Priority Use Permit\nFive-Year Review'
            ws.oddHeader.center.text = header1
            ws.oddHeader.center.size = 11
            ws.oddHeader.center.font = "Calibri, bold"
            ws.oddHeader.center.color = "000000"
            
            header2 = 'Business/Organization:\n {}'.format(businessname)
            ws.oddHeader.left.text = header2
            ws.oddHeader.left.size = 8
            ws.oddHeader.left.font = "Calibri"
            ws.oddHeader.left.color = "000000"
            
            footer = 'Prepared By: {}\nDate: {}'.format(username, datetime.today().strftime('%m/%d/%Y'))
            ws.oddFooter.left.text = footer
            
        def excelRow(ws, rows, styles):
            for row in ws[rows]:
                for cell in row:
                    cell.style = styles
        
        def excelStyle(ws, cells, styles):
            for cell in ws[cells]:
                cell.style = styles

        def excelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 3
            index_column = 'A'
            index_column1 = 'B'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'C4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            ws.column_dimensions[index_column1].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, index_column1, index_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)
        
    
        with pd.ExcelWriter(**writer_args) as xlsx:
            
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 3
            index_column = 'A'
            index_column1 = 'B'

            if (actAllPivot.empty == False):
                actAllPivot.to_excel(xlsx, 'Activity Summary')
                ws = xlsx.sheets['Activity Summary']   
                excelUpdate(ws)  

            
            if 'huntActPivot' in locals():
                huntActPivot.to_excel(xlsx, 'Hunt Summary')
                wsHunt = xlsx.sheets['Hunt Summary']
                excelUpdate(wsHunt)

            
            useAllPivot.to_excel(xlsx, 'Use Area Summary')            
            wsUseArea = xlsx.sheets['Use Area Summary']
            colUse=get_column_letter(wsUseArea.max_column)
            UseCol = wsUseArea.max_column
            rowsUse = wsUseArea.max_row
            value_use ='B4:{}{}'.format(colUse, rowsUse)
            wsUseArea.column_dimensions[index_column].width = 35            
            lastCellUse = 'B{}'.format(wsUseArea.max_row-1) 
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(wsUseArea, paper_size = 1, orientation='landscape')
            wsUseArea.sheet_view.view = 'pageLayout'
            excelRow(wsUseArea, value_use, value_style)
            excelStyle(wsUseArea, index_column, index_style)
            excelStyle(wsUseArea, title_row, title_style)
            excelStyle(wsUseArea, title_row1, title_style)
            excelStyle(wsUseArea, title_row2, title_style)
            headers(wsUseArea) 
            while value_col <= UseCol:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                wsUseArea.column_dimensions[i].width = 8
                value_col += 1

            
            locNameAllPivot.to_excel(xlsx, 'Location Summary')
            wsLocation = xlsx.sheets['Location Summary']            
            excelUpdate(wsLocation)

        return


        
class RVD_Report(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run NEPA Review-RVD Report"
        self.description = ""
        self.canRunInBackground = False
        
    def getParameterInfo(self):
        """Define parameter definitions"""
    
        param0 = arcpy.Parameter(
            displayName="For activities on ",
            name="RangerDistrict",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = ['Petersburg Ranger District', 'Wrangell Ranger District']

        param1  = arcpy.Parameter(
            displayName="that end between",
            name="startdate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param1.value = datetime(datetime.today().year, 1, 1).strftime('%x')

        param2 = arcpy.Parameter(
            displayName="and",
            name="enddate",
            datatype="GPDate",
            parameterType="Required",
            direction="Input")

        param2.value = datetime.today().strftime('%x')

        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]

        return params
              
    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        """The source code of the tool."""
        

        
        rangerDistrict = parameters[0].value
        startDate = parameters[1].value  
        endDate = parameters[2].value            
        savepath = parameters[3].value.value 


        self.savefile = "RVD_Report_{}_{}_{}.xlsx".format(rangerDistrict, startDate.strftime('%Y%m%d'), endDate.strftime('%Y%m%d'))
        self.path = os.path.join(savepath, self.savefile)   


        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        heliTrip ='{}S_R10.R10_OFG_HELISKI_TRIP'.format(connection)
        heliActivity = '{}S_R10.R10_OFG_HELISKI_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)


          
        def getFieldNames(shp):
            fieldnames = [f.name for f in arcpy.ListFields(shp)]
            return fieldnames

        
        """Creating Guided Recreation DataFrames and filtering data by the report timeframe and business name"""
        
        guideDayCol = getFieldNames(guideDay)
        guideStopCol = getFieldNames(guideStop)
        guideActCol = getFieldNames(guideActivity)
        
        guideWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(startDate.strftime('%Y-%m-%d %H:%M:%S'), endDate.strftime('%Y-%m-%d %H:%M:%S'))
        guideData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideWhere)]
        guideDayDF = pd.DataFrame(guideData, columns = guideDayCol)

        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol)

        guideActData =[row for row in arcpy.da.SearchCursor(guideActivity, guideActCol)]  
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)
        
        
        if (guideDayDF.empty == False):
            if (guideStopDF.empty ==False):
                if (guideActDF.empty == False):
                    guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
                    guideDF = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')

        
        """Creating Hunting Data Frames to only get the non-hunting activities and filtering data by the report timeframe and business name"""        

        huntDayCol = getFieldNames(huntDay)
        huntStopCol = getFieldNames(huntStop)
        huntActCol = getFieldNames(huntActivity)
    

        huntWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(startDate.strftime('%Y-%m-%d %H:%M:%S'), endDate.strftime('%Y-%m-%d %H:%M:%S'))
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns = huntDayCol)
        
        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol)]
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)


        huntActData =  [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol)]  
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
      
        
        if (huntDayDF.empty == False):
            if (huntStopDF.empty == False):
                if (huntActDF.empty == False):
                    huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
                    huntDF = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')        
    

        
        
        """ Creating Heliski dataframes in the date range and for the specific business name"""

        heliTripCol = getFieldNames(heliTrip)
        heliActCol = getFieldNames(heliActivity)

        heliWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(startDate.strftime('%Y-%m-%d %H:%M:%S'), endDate.strftime('%Y-%m-%d %H:%M:%S'))
        heliData = [row for row in arcpy.da.SearchCursor(heliTrip, heliTripCol, where_clause=heliWhere)]
        heliTripDF = pd.DataFrame(heliData, columns=heliTripCol)   

        heliActData = [row for row in arcpy.da.SearchCursor(heliActivity, heliActCol)]
        heliActDF = pd.DataFrame(heliActData, columns = heliActCol)
        
        if (heliTripDF.empty == False):
            if (heliActDF.empty == False):
                heliDF = heliActDF.merge(heliTripDF, how = 'inner', on='TRIP_GUID')
     
        
        """Creating Outfitting data frame in the date range and for the specific business name"""

        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(startDate.strftime('%Y-%m-%d %H:%M:%S'), endDate.strftime('%Y-%m-%d %H:%M:%S'))
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)

        
        activities = {'Boating, Stand Up Paddle Boarding, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Pack Rafting, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 
           'Boating, Raft, Canoe, Kayak or Other Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Canoeing (Mendenhall form)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Rafting (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Canoeing':'Boating (Non-Motorized, Freshwater)', 'Boating (Non-Motorized, Freshwater)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking':'Boating (Non-Motorized, Freshwater)', 'Rafting':'Boating (Non-Motorized, Freshwater)', 'Camping':'Camping','Dog Sledding':'Dog Sled Tours', 'Dog Sled Tours':'Dog Sled Tours', 'Flightseeing Landing Tours':'Flightseeing Landing Tours',
           'Freshwater Fishing':'Freshwater Fishing', 'Glacier Trekking':'Helicopter Landing Tours', 'Helicopter Landing Tours':'Helicopter Landing Tours', 'Heli-skiing Tours':'Heli-skiing Tours', 
           'Heliski':'Heli-skiing Tours', 'Hunting, Brown Bear':'Hunting, Brown Bear', 'Hunting, Deer':'Hunting, Deer', 'Remote Setting Nature Tour, on Foot': 'Remote Setting Nature Tour', 
           'Hunting, Elk':'Hunting, Elk', 'Hunting, Moose':'Hunting, Moose', 'Hunting, Mountain Goat':'Hunting, Mountain Goat', 'Remote Setting Nature Tour':'Remote Setting Nature Tour',
           'Hunting, Wolf':'Hunting, Waterfowl/Small game/Wolf', 'Outfitting':'Outfitting (Delivery and/or pick-up of vehicles, equipment, etc. to/from National Forest System lands; Total per day - no limit on equipment numbers or number of trips)', 
           'Over-Snow Vehicle Tours':'Over-Snow Vehicle Tours', 'Bikepacking':'Remote Setting Nature Tour', 'Horseback Riding':'Remote Setting Nature Tour', 'Nature Tour, on Foot':'Remote Setting Nature Tour',
           'Nature Tour, on Ski':'Remote Setting Nature Tour','Nature Tour, Bicycle':'Remote Setting Nature Tour', 'Biking (Mendenhall form)':'Remote Setting Nature Tour', 'Biking':'Remote Setting Nature Tour', 
           'Hiking (Mendenhall form)':'Remote Setting Nature Tour', 'Hiking':'Remote Setting Nature Tour', 'Nature Tour, ATV/OHV':'Road Based Nature Tours', 'Nature Tour, Vehicle':'Road Based Nature Tours', 
           'Visitor Center (Begich Boggs, MGVC, SEADC)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Visitor Center Transport (Mendenhall form)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 
           'Visitor_Center_Transport':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Hunting, Black Bear ':'Hunting, Black Bear', 'Hunting, Dall Sheep ':'Hunting, Dall Sheep', 'Hunting, Waterfowl/Small game ':'Hunting, Waterfowl/Small game/Wolf',
           'Assigned Site' : 'Assigned Site', 'Minimum Fee': 'Minimum Fee', 'Hunting, Black Bear':'Hunting, Black Bear', 'Hunting, Waterfowl/Small game/Wolf - Service Day Rate':'Hunting, Waterfowl/Small game/Wolf'}
        
        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['FULLNAME'].fillna(locDF['ZONE_NAME'], inplace=True)
        locDF['USE_AREA'] = locDF['FULLNAME']
        locDF['USE_AREA'].fillna('Use Area Unknown', inplace=True)
        locDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
        locDF['USELOCATION'].str.upper()

       
        if (guideDF['USELOCATION'].loc() == guideDF['USELOCATION'].loc()):
            guideDF['USELOCATION'] = guideDF.loc(guideDF['USELOCATION'].map(lambda x: x.title()))
        if 'guideDF' not in locals():
            arcpy.AddMessage('This business had no guided recreation activities for this time frame.') 
            guideFinal = pd.DataFrame()
        else:
            guideDF['Activity2'] = guideDF['ACTIVITY'].map(activities)      

            guideLoc = guideDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])

            guideFilter = guideLoc.loc[guideLoc['DISTRICTNAME'] == rangerDistrict]

            guideFilter.rename(columns= {'HOURSSPENTONFS': 'HOURS', 'AREANUMBER': 'AREA', 'BUSINESSNAME_x':'BUSINESS NAME', 'CLIENTNUMBER': 'CLIENTS', 'USELOCATION_x':'USELOCATION'}, inplace = True)
            guideFilter['DATE'] = guideFilter['TRIPDATE'].dt.strftime('%m/%d/%Y')
            guideFilter['RVD'] = (guideFilter['CLIENTS'] * guideFilter['HOURS'])/12
            guideFinal = guideFilter[['DATE', 'AREA', 'USELOCATION', 'BUSINESS NAME', 'Activity2', 'CLIENTS', 'HOURS', 'RVD']].copy()
   
        if 'huntDF' not in locals():
            arcpy.AddMessage('This business had no hunting activities for this time frame.')  
            huntFinal = pd.DataFrame()
        else:
            huntDF['Activity2'] = huntDF['ACTIVITY'].map(activities)
            huntDF.rename(columns= {'CLIENTNUMBER': 'CLIENTS'}, inplace = True)  
            huntDF['USELOCATION'].str.upper()
            huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
            huntFilter = huntLoc.loc[huntLoc['DISTRICTNAME'] == rangerDistrict]
            huntFilter.rename(columns= {'HOURSSPENTONFS': 'HOURS', 'AREANUMBER': 'AREA', 'BUSINESSNAME_x':'BUSINESS NAME', 'USELOCATION_x':'USELOCATION'}, inplace = True)            
            huntFilter['DATE'] = huntFilter['TRIPDATE'].dt.strftime('%m/%d/%Y')
            huntFilter['RVD'] = (huntFilter['CLIENTS'] * huntFilter['HOURS'])/12 
            campDF = huntFilter.loc[huntFilter['Activity2'] == 'Camping']
            if not campDF.empty:
                tripDF = campDF['TRIP_GUID']
                for t in tripDF: 
                    if t == huntFilter['TRIP_GUID']:
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Black Bear', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Elk', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Brown Bear', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Deer', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Mountain Goat', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Moose', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Waterfowl/Small game/Wolf', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
                        huntFilter.loc[huntFilter['Activity2']=='Hunting, Dall Sheep', 'RVD x 3 for Hunting'] = huntFilter['RVD']*3
            else: 
                huntFilter['RVD x 3 for Hunting']=""
            huntFinal = huntFilter[['DATE', 'AREA', 'USELOCATION', 'BUSINESS NAME', 'Activity2', 'CLIENTS', 'HOURS', 'RVD', 'RVD x 3 for Hunting']].copy()       
            
        if 'heliDF' not in locals():
            arcpy.AddMessage('This business had no heliski activities for this time frame.') 
            heliFinal = pd.DataFrame()
        else:        
            heliDF['Activity2'] = heliDF['ACTIVITY'].map(activities)
            heliDF['USELOCATION'].str.upper()
            heliLoc = heliDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
            heliFilter = heliLoc.loc[heliLoc['DISTRICTNAME'] == rangerDistrict]            
            heliFilter.rename(columns= {'CLIENTS_LOCATION':'CLIENTS', 'HOURSSPENTONFS': 'HOURS', 'AREANUMBER': 'AREA', 'BUSINESSNAME_x':'BUSINESS NAME', 'USELOCATION_x':'USELOCATION'}, inplace = True) 
            heliFilter['DATE'] = heliFilter['TRIPDATE'].dt.strftime('%m/%d/%Y')
            heliFilter['RVD'] = (heliFilter['CLIENTS'] * heliFilter['HOURS'])/12     
            heliFinal = heliFilter[['DATE', 'AREA', 'USELOCATION', 'BUSINESS NAME', 'Activity2', 'CLIENTS', 'HOURS', 'RVD']].copy()

        if 'outfitDF' not in locals():
            arcpy.AddMessage('This business had no outfitting activities for this time frame.')  
            outfitFinal = pd.DataFrame()
        else:
            outfitDF['Activity2'] = outfitDF['ACTIVITY'].map(activities)
            outfitDF['USELOCATION'].str.upper()
            outfitLoc = outfitDF.merge(locDF, how = 'left', on = ['LOCATION_ID'])
            outfitFilter = outfitLoc.loc[outfitLoc['DISTRICTNAME'] == rangerDistrict]           
            outfitFilter.rename(columns = {'TOTALCLIENTSONDAY':'CLIENTS', 'BUSINESSNAME':'BUSINESS NAME', 'AREANUMBER': 'AREA', 'USELOCATION_x':'USELOCATION'}, inplace=True)
            outfitFilter['HOURS'] =0               
            outfitFilter['DATE'] = outfitFilter['TRIPDATE'].dt.strftime('%m/%d/%Y')
            outfitFilter['RVD'] = (outfitFilter['CLIENTS'] * outfitFilter['HOURS'])/12 
            outfitFinal = outfitFilter[['DATE', 'AREA', 'USELOCATION', 'BUSINESS NAME', 'Activity2', 'CLIENTS', 'HOURS', 'RVD']].copy()            
          
        if 'guideFinal' not in locals():
             if 'heliFinal' not in locals():
                 if 'outfitFinal' not in locals():
                     if 'huntFinal' not in locals():
                         arcpy.AddMessage('There are no activities between the start and end date.')
                         exit()
        
        
        
        rvdDF = guideFinal.append([huntFinal, heliFinal, outfitFinal])
        rvdDF.rename(columns = {'Activity2':'ACTIVITY'}, inplace=True)
        pd.to_datetime(rvdDF['DATE'], format="%x")
        
        
        columnsTitles = ['DATE', 'AREA', 'USELOCATION', 'BUSINESS NAME', 'ACTIVITY', 'CLIENTS', 'HOURS', 'RVD', 'RVD x 3 for Hunting']
        rvdDF.sort_values(by=['BUSINESS NAME', 'DATE'], inplace = True, ascending=True)
        rvdDF = rvdDF.reindex(columns=columnsTitles)
        
         
        writer_args = {
            'path': self.path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')

        with pd.ExcelWriter(**writer_args) as xlsx:
            rvdDF.to_excel(xlsx, 'All Areas')
            ws = xlsx.sheets['All Areas']

            cols = ['A','B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            rows=ws.max_row

            for c in cols: 
                ws.column_dimensions[c].width = 35
 
            for row in range(1, rows):
                ws["I{}".format(row)].number_format = '#,##0.00'
                ws["J{}".format(row)].number_format = '#,##0.00'
            ws.delete_cols(1)            


class NEPAReview_Shoreline2(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run NEPA Review-Shoreline II Report"
        self.description = ""
        self.canRunInBackground = False

    def getParameterInfo(self):
        """Define parameter definitions"""


        param0  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param0.value = startyear -5
        
        param1  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param1.value = endyear -1


        param2 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param2.filter.list = ["File System"]

        params = [param0, param1, param2]
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        """This tool is sued to populate the Shoreline II Allocation status template. It calculates the number of clients by use are by season."""
        
        startYear = parameters[0].value
        endYear = parameters[1].value
        savepath = parameters[2].value.value  
        start = str(int(startYear))
        end = str(int(endYear))


        template = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\ShorelineIIAllocationStatus_template.xlsx'
        wb= openpyxl.Workbook()
        wb= openpyxl.load_workbook(template)
        ws = wb['AllocStatus']
        ws.sheet_state = 'visible'
        savefile = "NEPAReview_Shoreline2" + "_" + start + "_" + end + ".xlsx"
        path = os.path.join(savepath, savefile)
        wb.save(path)
        

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames

        """List of Full Name use areas that are in the Shoreline II area."""
        
        shoreline2 = ['01-01 SKAGWAY AREA','01-03 EAST CHILKATS','04-03 SITKA AREA','04-08 NE ADMIRALTY','04-12 TENAKEE INLET','04-13 PERIL STRAIT','04-14 SLOCUM ARM','01-04A BERNERS BAY',
                     '01-04B N. JUNEAU COAST','01-04C TAKU INLET','01-04D SLOCUM INLET','01-05A TAKU HARBOR','01-05B PORT SNETTISHAM','01-05C WINDHAM BAY','01-05D TRACY ARM','01-05E FORDS TERROR',
                     '01-05F ENDICOTT ARM','04-01A GUT BAY, BARANOF','04-01B PORT ARMSTRONG','04-01C NELSON BAY','04-02A REDOUBT LAKE','04-02B WHALE BAY','04-02C NECKER ISLANDS','04-02D SW BARANOF',
                     '04-04A RODMAN BAY','04-04B KELP BAY','04-04C BARANOF WARM SPRINGS','04-05A SW ADMIRALTY','04-05B MITCHELL BAY','04-06A PYBUS BAY','04-06B ELIZA HARBOR','04-07A GAMBIER BAY',
                     '04-07B CANOE ROUTE','04-09A SEYMOUR CANAL','04-09B PACK CREEK','04-10A GREENS CREEK','04-10B NW ADMIRALTY','04-11A PORT FREDERICK','04-11B FRESHWATER BAY','04-15A LISIANSKI',
                     '04-15B WEST YAKOBI ISLAND','04-15C STAG BAY','04-15D PORTLOCK HARBOR','04-16A POINT ADOLPHUS','04-16B NORTH CHICHAGOF','04-16C IDAHO INLET','04-16D PLI WILDERNESS',
                     '04-16D PORT ALTHORP', '04-01A GUT BAY, BARANOF ']




        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'
        
        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)


        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['USELOCATION'].str.upper()



        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
        
        start = int(startYear)
        end = int(endYear)

        """ Calculating the total sum number of clients by use area and by season. But because the same group of people can visit two stop locations in the same use are, if we only summed client number
        we would count that group twice. But if a group of people visited two use areas in one trip, we need to count them twice. So we sum the total nubmer of clients by use area and season, then make 
        sure that sum is not greater than the total number of clinets on the day by use area and season. If the client number sum is greater than the total clents on the day, we use total clients on 
        the day. """

        guideDayCol = getFieldNames(guideDay)
        guideDayWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        guideDayData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideDayWhere)]
        guideDayDF = pd.DataFrame(guideDayData, columns=guideDayCol)
       
        guideStopCol = getFieldNames(guideStop)
        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol) 
        
        guideActCol = getFieldNames(guideActivity)
        guideActData = [row for row in arcpy.da.SearchCursor(guideActivity, guideActCol)]
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)

        guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
        guideDF = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')                 

        
        guideDF['Year'] = guideDF['TRIPDATE'].dt.year            
        guideDF['USELOCATION'].str.upper()
        guideLoc = guideDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID']) 
        
        guideLoc['keep'] = guideLoc['FULLNAME'].apply(lambda x: 'True' if x in shoreline2 else 'False')
        
        guideLoc1 = guideLoc[guideLoc["keep"].str.contains("True")]
        
        
        
        guideActGroup = guideLoc1.groupby(['DAY_GUID', 'FULLNAME'], as_index = False).agg({'Year': 'first','SEASON':'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
        guideActGroup.loc[guideActGroup['CLIENTNUMBER'] > guideActGroup['TOTALCLIENTSONDAY'], 'USE'] = guideActGroup['TOTALCLIENTSONDAY']
        guideActGroup['USE'].fillna(guideActGroup['CLIENTNUMBER'], axis = 0, inplace = True)
        
        guideRec = guideActGroup[['SEASON', 'FULLNAME', 'Year', 'USE']]


        
        huntDayCol = getFieldNames(huntDay)
        huntDayWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntDayWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns=huntDayCol)
       
        huntStopCol = getFieldNames(huntStop)
        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol)]  
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)
        
        huntActCol = getFieldNames(huntActivity)
        huntActData = [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol)] 
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
        
        huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
        huntDF = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')
        
        huntDF['Year'] = huntDF['TRIPDATE'].dt.year        
        
        huntDF['USELOCATION'].str.upper()
        huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
        
        huntLoc['keep'] = huntLoc['FULLNAME'].apply(lambda x: 'True' if x in shoreline2 else 'False')
        
        huntLoc1 = huntLoc[huntLoc["keep"].str.contains("True")]
        
        huntActGroup = huntLoc1.groupby(['DAY_GUID', 'FULLNAME'], as_index = False).agg({'Year': 'first', 'SEASON' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
        huntActGroup.loc[huntActGroup['CLIENTNUMBER'] > huntActGroup['TOTALCLIENTSONDAY'], 'USE'] = huntActGroup['TOTALCLIENTSONDAY']
        huntActGroup['USE'].fillna(huntActGroup['TOTALCLIENTSONDAY'], axis = 0, inplace = True)        
        
        hunting = huntActGroup[['SEASON', 'FULLNAME', 'Year', 'USE']]
        

        
        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)
        outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
        outfitDF['USELOCATION'].str.upper()
        outfitLoc = outfitDF.merge(locDF, how = 'inner', on = ['LOCATION_ID', 'USELOCATION'])
        outfitLoc.rename(columns = {'TOTALCLIENTSONDAY':'USE'}, inplace=True)
        
        outfitLoc['keep']= outfitLoc['FULLNAME'].apply(lambda x: 'True' if x in shoreline2 else 'False')

        outfitLoc1 = outfitLoc[outfitLoc["keep"].str.contains("True")]        
        
        outfitting = outfitLoc1[['SEASON', 'FULLNAME', 'Year', 'USE']] 
        

        
        allDF = guideRec.append([hunting, outfitting])
        

        allPivot = pd.pivot_table(allDF, index = ['SEASON', 'FULLNAME'], columns = ['Year'], values = ['USE'], aggfunc=np.sum, margins = False, dropna = True)
       
        allPivot.fillna(0, inplace=True)
        allPivot.reset_index(inplace=True)
        allPivot['ActualUseAvg'] = (allPivot.sum(axis=1))/5


        
        allPivot = allPivot.round()
        
        """ Because we are writing specific values to specific cells in the template sheet, we use the update function to do that. """

        writer_args = {
            'path': path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        def update(season, fullname, startNumb, columnNumb):
            x = allPivot.loc[(allPivot.SEASON == season) & (allPivot.FULLNAME==fullname), 'ActualUseAvg']
            for row, text in enumerate(x, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)


        with pd.ExcelWriter(**writer_args) as xlsx:
            wb = openpyxl.load_workbook(path)
            ws = wb['AllocStatus']
            xlsx.book = wb
            
            update('Early Spring', '01-01 SKAGWAY AREA', 3, 4)
            update('Early Spring', '01-02 HAINES AREA', 4, 4)
            update('Early Spring', '01-03 EAST CHILKATS', 5, 4)
            update('Early Spring', '01-04A BERNERS BAY', 6, 4)
            update('Early Spring', '01-04B N. JUNEAU COAST', 7, 4)
            update('Early Spring', '01-04C TAKU INLET', 8, 4)
            update('Early Spring', '01-04D SLOCUM INLET', 9, 4)
            update('Early Spring', '01-05A TAKU HARBOR', 10, 4)
            update('Early Spring', '01-05B PORT SNETTISHAM', 11, 4)
            update('Early Spring', '01-05C WINDHAM BAY', 12, 4)
            update('Early Spring', '01-05D TRACY ARM', 13, 4)
            update('Early Spring', '01-05E FORDS TERROR', 14, 4)
            update('Early Spring', '01-05F ENDICOTT ARM', 15, 4)
            update('Early Spring', '04-01A GUT BAY, BARANOF', 16, 4)
            update('Early Spring', '04-01B PORT ARMSTRONG', 17, 4)
            update('Early Spring', '04-01C NELSON BAY', 18, 4)
            update('Early Spring', '04-02A REDOUBT LAKE', 19, 4)
            update('Early Spring', '04-02B WHALE BAY', 20, 4)
            update('Early Spring', '04-02C NECKER ISLANDS', 21, 4)
            update('Early Spring', '04-02D SW BARANOF', 22, 4)
            update('Early Spring', '04-03 SITKA AREA', 23, 4)
            update('Early Spring', '04-04A RODMAN BAY', 24, 4)
            update('Early Spring', '04-04B KELP BAY', 25, 4)
            update('Early Spring', '04-04C BARANOF WARM SPRINGS', 26, 4)
            update('Early Spring', '04-05A SW ADMIRALTY', 27, 4)
            update('Early Spring', '04-06A PYBUS BAY', 28, 4)
            update('Early Spring', '04-06B ELIZA HARBOR', 29, 4)
            update('Early Spring', '04-07A GAMBIER BAY', 30, 4)
            update('Early Spring', '04-07B CROSS-ADMIRALTY CANOE RTE', 31, 4)
            update('Early Spring', '04-08 NE ADMIRALTY', 32, 4)
            update('Early Spring', '04-09A SEYMOUR CANAL', 33, 4)
            update('Early Spring', '04-09B PACK CREEK ZOOLOGICAL AREA', 34, 4)
            update('Early Spring', '04-10A GREENS CREEK', 35, 4)
            update('Early Spring', '04-10B NW ADMIRALTY', 36, 4)
            update('Early Spring', '04-11A PORT FREDERICK', 37, 4)
            update('Early Spring', '04-11B FRESHWATER BAY', 38, 4)
            update('Early Spring', '04-12 TENAKEE INLET', 39, 4)
            update('Early Spring', '04-13 PERIL STRAIT', 40, 4)
            update('Early Spring', '04-14 SLOCUM ARM', 41, 4)
            update('Early Spring', '04-15A LISIANSKI', 42, 4)
            update('Early Spring', '04-15B WEST YAKOBI ISLAND', 43, 4)
            update('Early Spring', '04-15C STAG BAY', 44, 4)
            update('Early Spring', '04-15D PORTLOCK HARBOR', 45, 4)
            update('Early Spring', '04-16A POINT ADOLPHUS', 46, 4)
            update('Early Spring', '04-16B NORTH CHICHAGOF', 47, 4)
            update('Early Spring', '04-16C IDAHO INLET', 48, 4)
            update('Early Spring', '04-16D PLI WILDERNESS', 49, 4)
            update('Early Spring', '04-16E PORT ALTHORP', 50, 4)
            
            
            update('Late Spring', '01-01 SKAGWAY AREA', 3, 9)
            update('Late Spring', '01-02 HAINES AREA', 4, 9)
            update('Late Spring', '01-03 EAST CHILKATS', 5, 9)
            update('Late Spring', '01-04A BERNERS BAY', 6, 9)
            update('Late Spring', '01-04B N. JUNEAU COAST', 7, 9)
            update('Late Spring', '01-04C TAKU INLET', 8, 9)
            update('Late Spring', '01-04D SLOCUM INLET', 9, 9)
            update('Late Spring', '01-05A TAKU HARBOR', 10, 9)
            update('Late Spring', '01-05B PORT SNETTISHAM', 11, 9)
            update('Late Spring', '01-05C WINDHAM BAY', 12, 9)
            update('Late Spring', '01-05D TRACY ARM', 13, 9)
            update('Late Spring', '01-05E FORDS TERROR', 14, 9)
            update('Late Spring', '01-05F ENDICOTT ARM', 15, 9)
            update('Late Spring', '04-01A GUT BAY, BARANOF', 16, 9)
            update('Late Spring', '04-01B PORT ARMSTRONG', 17, 9)
            update('Late Spring', '04-01C NELSON BAY', 18, 9)
            update('Late Spring', '04-02A REDOUBT LAKE', 19, 9)
            update('Late Spring', '04-02B WHALE BAY', 20, 9)
            update('Late Spring', '04-02C NECKER ISLANDS', 21, 9)
            update('Late Spring', '04-02D SW BARANOF', 22, 9)
            update('Late Spring', '04-03 SITKA AREA', 23, 9)
            update('Late Spring', '04-04A RODMAN BAY', 24, 9)
            update('Late Spring', '04-04B KELP BAY', 25, 9)
            update('Late Spring', '04-04C BARANOF WARM SPRINGS', 26, 9)
            update('Late Spring', '04-05A SW ADMIRALTY', 27, 9)
            update('Late Spring', '04-06A PYBUS BAY', 28, 9)
            update('Late Spring', '04-06B ELIZA HARBOR', 29, 9)
            update('Late Spring', '04-07A GAMBIER BAY', 30, 9)
            update('Late Spring', '04-07B CROSS-ADMIRALTY CANOE RTE', 31, 9)
            update('Late Spring', '04-08 NE ADMIRALTY', 32, 9)
            update('Late Spring', '04-09A SEYMOUR CANAL', 33, 9)
            update('Late Spring', '04-09B PACK CREEK ZOOLOGICAL AREA', 34, 9)
            update('Late Spring', '04-10A GREENS CREEK', 35, 9)
            update('Late Spring', '04-10B NW ADMIRALTY', 36, 9)
            update('Late Spring', '04-11A PORT FREDERICK', 37, 9)
            update('Late Spring', '04-11B FRESHWATER BAY', 38, 9)
            update('Late Spring', '04-12 TENAKEE INLET', 39, 9)
            update('Late Spring', '04-13 PERIL STRAIT', 40, 9)
            update('Late Spring', '04-14 SLOCUM ARM', 41, 9)
            update('Late Spring', '04-15A LISIANSKI', 42, 9)
            update('Late Spring', '04-15B WEST YAKOBI ISLAND', 43, 9)
            update('Late Spring', '04-15C STAG BAY', 44, 9)
            update('Late Spring', '04-15D PORTLOCK HARBOR', 45, 9)
            update('Late Spring', '04-16A POINT ADOLPHUS', 46, 9)
            update('Late Spring', '04-16B NORTH CHICHAGOF', 47, 9)
            update('Late Spring', '04-16C IDAHO INLET', 48, 9)
            update('Late Spring', '04-16D PLI WILDERNESS', 49, 9)
            update('Late Spring', '04-16E PORT ALTHORP', 50, 9)
            
            update('Summer', '01-01 SKAGWAY AREA', 3, 14)
            update('Summer', '01-02 HAINES AREA', 4, 14)
            update('Summer', '01-03 EAST CHILKATS', 5, 14)
            update('Summer', '01-04A BERNERS BAY', 6, 14)
            update('Summer', '01-04B N. JUNEAU COAST', 7, 14)
            update('Summer', '01-04C TAKU INLET', 8, 14)
            update('Summer', '01-04D SLOCUM INLET', 9, 14)
            update('Summer', '01-05A TAKU HARBOR', 10, 14)
            update('Summer', '01-05B PORT SNETTISHAM', 11, 14)
            update('Summer', '01-05C WINDHAM BAY', 12, 14)
            update('Summer', '01-05D TRACY ARM', 13, 14)
            update('Summer', '01-05E FORDS TERROR', 14, 14)
            update('Summer', '01-05F ENDICOTT ARM', 15, 14)
            update('Summer', '04-01A GUT BAY, BARANOF', 16, 14)
            update('Summer', '04-01B PORT ARMSTRONG', 17, 14)
            update('Summer', '04-01C NELSON BAY', 18, 14)
            update('Summer', '04-02A REDOUBT LAKE', 19, 14)
            update('Summer', '04-02B WHALE BAY', 20, 14)
            update('Summer', '04-02C NECKER ISLANDS', 21, 14)
            update('Summer', '04-02D SW BARANOF', 22, 14)
            update('Summer', '04-03 SITKA AREA', 23, 14)
            update('Summer', '04-04A RODMAN BAY', 24, 14)
            update('Summer', '04-04B KELP BAY', 25, 14)
            update('Summer', '04-04C BARANOF WARM SPRINGS', 26, 14)
            update('Summer', '04-05A SW ADMIRALTY', 27, 14)
            update('Summer', '04-06A PYBUS BAY', 28, 14)
            update('Summer', '04-06B ELIZA HARBOR', 29, 14)
            update('Summer', '04-07A GAMBIER BAY', 30, 14)
            update('Summer', '04-07B CROSS-ADMIRALTY CANOE RTE', 31, 14)
            update('Summer', '04-08 NE ADMIRALTY', 32, 14)
            update('Summer', '04-09A SEYMOUR CANAL', 33, 14)
            update('Summer', '04-09B PACK CREEK ZOOLOGICAL AREA', 34, 14)
            update('Summer', '04-10A GREENS CREEK', 35, 14)
            update('Summer', '04-10B NW ADMIRALTY', 36, 14)
            update('Summer', '04-11A PORT FREDERICK', 37, 14)
            update('Summer', '04-11B FRESHWATER BAY', 38, 14)
            update('Summer', '04-12 TENAKEE INLET', 39, 14)
            update('Summer', '04-13 PERIL STRAIT', 40, 14)
            update('Summer', '04-14 SLOCUM ARM', 41, 14)
            update('Summer', '04-15A LISIANSKI', 42, 14)
            update('Summer', '04-15B WEST YAKOBI ISLAND', 43, 14)
            update('Summer', '04-15C STAG BAY', 44, 14)
            update('Summer', '04-15D PORTLOCK HARBOR', 45, 14)
            update('Summer', '04-16A POINT ADOLPHUS', 46, 14)
            update('Summer', '04-16B NORTH CHICHAGOF', 47, 14)
            update('Summer', '04-16C IDAHO INLET', 48, 14)
            update('Summer', '04-16D PLI WILDERNESS', 49, 14)
            update('Summer', '04-16E PORT ALTHORP', 50, 14)            

            update('Fall', '01-01 SKAGWAY AREA', 3, 19)
            update('Fall', '01-02 HAINES AREA', 4, 19)
            update('Fall', '01-03 EAST CHILKATS', 5, 19)
            update('Fall', '01-04A BERNERS BAY', 6, 19)
            update('Fall', '01-04B N. JUNEAU COAST', 7, 19)
            update('Fall', '01-04C TAKU INLET', 8, 19)
            update('Fall', '01-04D SLOCUM INLET', 9, 19)
            update('Fall', '01-05A TAKU HARBOR', 10, 19)
            update('Fall', '01-05B PORT SNETTISHAM', 11, 19)
            update('Fall', '01-05C WINDHAM BAY', 12, 19)
            update('Fall', '01-05D TRACY ARM', 13, 19)
            update('Fall', '01-05E FORDS TERROR', 14, 19)
            update('Fall', '01-05F ENDICOTT ARM', 15, 19)
            update('Fall', '04-01A GUT BAY, BARANOF', 16, 19)
            update('Fall', '04-01B PORT ARMSTRONG', 17, 19)
            update('Fall', '04-01C NELSON BAY', 18, 19)
            update('Fall', '04-02A REDOUBT LAKE', 19, 19)
            update('Fall', '04-02B WHALE BAY', 20, 19)
            update('Fall', '04-02C NECKER ISLANDS', 21, 19)
            update('Fall', '04-02D SW BARANOF', 22, 19)
            update('Fall', '04-03 SITKA AREA', 23, 19)
            update('Fall', '04-04A RODMAN BAY', 24, 19)
            update('Fall', '04-04B KELP BAY', 25, 19)
            update('Fall', '04-04C BARANOF WARM SPRINGS', 26, 19)
            update('Fall', '04-05A SW ADMIRALTY', 27, 19)
            update('Fall', '04-06A PYBUS BAY', 28, 19)
            update('Fall', '04-06B ELIZA HARBOR', 29, 19)
            update('Fall', '04-07A GAMBIER BAY', 30, 19)
            update('Fall', '04-07B CROSS-ADMIRALTY CANOE RTE', 31, 19)
            update('Fall', '04-08 NE ADMIRALTY', 32, 19)
            update('Fall', '04-09A SEYMOUR CANAL', 33, 19)
            update('Fall', '04-09B PACK CREEK ZOOLOGICAL AREA', 34, 19)
            update('Fall', '04-10A GREENS CREEK', 35, 19)
            update('Fall', '04-10B NW ADMIRALTY', 36, 19)
            update('Fall', '04-11A PORT FREDERICK', 37, 19)
            update('Fall', '04-11B FRESHWATER BAY', 38, 19)
            update('Fall', '04-12 TENAKEE INLET', 39, 19)
            update('Fall', '04-13 PERIL STRAIT', 40, 19)
            update('Fall', '04-14 SLOCUM ARM', 41, 19)
            update('Fall', '04-15A LISIANSKI', 42, 19)
            update('Fall', '04-15B WEST YAKOBI ISLAND', 43, 19)
            update('Fall', '04-15C STAG BAY', 44, 19)
            update('Fall', '04-15D PORTLOCK HARBOR', 45, 19)
            update('Fall', '04-16A POINT ADOLPHUS', 46, 19)
            update('Fall', '04-16B NORTH CHICHAGOF', 47, 19)
            update('Fall', '04-16C IDAHO INLET', 48, 19)
            update('Fall', '04-16D PLI WILDERNESS', 49, 19)
            update('Fall', '04-16E PORT ALTHORP', 50, 19)
            
            update('Winter', '01-01 SKAGWAY AREA', 3, 24)
            update('Winter', '01-02 HAINES AREA', 4, 24)
            update('Winter', '01-03 EAST CHILKATS', 5, 24)
            update('Winter', '01-04A BERNERS BAY', 6, 24)
            update('Winter', '01-04B N. JUNEAU COAST', 7, 24)
            update('Winter', '01-04C TAKU INLET', 8, 24)
            update('Winter', '01-04D SLOCUM INLET', 9, 24)
            update('Winter', '01-05A TAKU HARBOR', 10, 24)
            update('Winter', '01-05B PORT SNETTISHAM', 11, 24)
            update('Winter', '01-05C WINDHAM BAY', 12, 24)
            update('Winter', '01-05D TRACY ARM', 13, 24)
            update('Winter', '01-05E FORDS TERROR', 14, 24)
            update('Winter', '01-05F ENDICOTT ARM', 15, 24)
            update('Winter', '04-01A GUT BAY, BARANOF', 16, 24)
            update('Winter', '04-01B PORT ARMSTRONG', 17, 24)
            update('Winter', '04-01C NELSON BAY', 18, 24)
            update('Winter', '04-02A REDOUBT LAKE', 19, 24)
            update('Winter', '04-02B WHALE BAY', 20, 24)
            update('Winter', '04-02C NECKER ISLANDS', 21, 24)
            update('Winter', '04-02D SW BARANOF', 22, 24)
            update('Winter', '04-03 SITKA AREA', 23, 24)
            update('Winter', '04-04A RODMAN BAY', 24, 24)
            update('Winter', '04-04B KELP BAY', 25, 24)
            update('Winter', '04-04C BARANOF WARM SPRINGS', 26, 24)
            update('Winter', '04-05A SW ADMIRALTY', 27, 24)
            update('Winter', '04-06A PYBUS BAY', 28, 24)
            update('Winter', '04-06B ELIZA HARBOR', 29, 24)
            update('Winter', '04-07A GAMBIER BAY', 30, 24)
            update('Winter', '04-07B CROSS-ADMIRALTY CANOE RTE', 31, 24)
            update('Winter', '04-08 NE ADMIRALTY', 32, 24)
            update('Winter', '04-09A SEYMOUR CANAL', 33, 24)
            update('Winter', '04-09B PACK CREEK ZOOLOGICAL AREA', 34, 24)
            update('Winter', '04-10A GREENS CREEK', 35, 24)
            update('Winter', '04-10B NW ADMIRALTY', 36, 24)
            update('Winter', '04-11A PORT FREDERICK', 37, 24)
            update('Winter', '04-11B FRESHWATER BAY', 38, 24)
            update('Winter', '04-12 TENAKEE INLET', 39, 24)
            update('Winter', '04-13 PERIL STRAIT', 40, 24)
            update('Winter', '04-14 SLOCUM ARM', 41, 24)
            update('Winter', '04-15A LISIANSKI', 42, 24)
            update('Winter', '04-15B WEST YAKOBI ISLAND', 43, 24)
            update('Winter', '04-15C STAG BAY', 44, 24)
            update('Winter', '04-15D PORTLOCK HARBOR', 45, 24)
            update('Winter', '04-16A POINT ADOLPHUS', 46, 24)
            update('Winter', '04-16B NORTH CHICHAGOF', 47, 24)
            update('Winter', '04-16C IDAHO INLET', 48, 24)
            update('Winter', '04-16D PLI WILDERNESS', 49, 24)
            update('Winter', '04-16E PORT ALTHORP', 50, 24)   
            
            colHead = 'Actual Use (5-yr Avg, {}-{})'.format(start, end)
            
            ws.cell(row=2, column=4).value = colHead
            ws.cell(row=2, column=9).value = colHead
            ws.cell(row=2, column=14).value = colHead
            ws.cell(row=2, column=19).value = colHead
            ws.cell(row=2, column=24).value = colHead

            wb.save(path)

        return 

class NEPAReview_KMRD(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run NEPA Review-KMRD Allocation Report"
        self.description = ""
        self.canRunInBackground = False

    def getParameterInfo(self):
        """Define parameter definitions"""


        param0  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param0.value = startyear -5
        
        param1  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param1.value = endyear -1


        param2 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param2.filter.list = ["File System"]

        params = [param0, param1, param2]
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        """This tool is sued to populate the KMRD Allocation status template. It calculates the number of clients by use are by season."""
        
        startYear = parameters[0].value
        endYear = parameters[1].value
        savepath = parameters[2].value.value  
        start = str(int(startYear))
        end = str(int(endYear))


        template = r'T:\FS\BusOps\EnterpriseProgram\Project\R10\RegionalProjects\R10_OutfitterSurvey\Workspace\kmmiles\Reports\KMRDAllocationStatus_template.xlsx'
        wb= openpyxl.Workbook()
        wb= openpyxl.load_workbook(template)
        ws = wb['AllocStatus']
        ws.sheet_state = 'visible'
        savefile = "NEPAReview_KMRD" + "_" + start + "_" + end + ".xlsx"
        path = os.path.join(savepath, savefile)
        wb.save(path)

        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'
        
        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)


        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames

        """List of Full Name use areas that are in the Shoreline II area."""

        
        kmrd2 = ['K04 DUKE ISLAND', 'K05 SOUTH MISTY LAKES', 'K06 MISTY CORE LAKES', 'K07 WALKER CHICKAMIN', 'K08 BURROUGHS UNUK', 'K09 ALAVA PRINCESS MANZANITA', 'K10 RUDYERD WINSTANLEY', 
                     'K11 GRAVINA ISLAND', 'K12 BELL ISLAND', 'K13 EAST CLEVELAND', 'K14 WEST CLEVELAND', 'K15 WILSON BAKEWELL', 'K16 KETCHIKAN CORE SPNW', 'K17 GEORGE CARROLL THORNE', 
                     'K18 CENTRAL REVILLA SPNW', 'K19 NORTH REVILLA', 'K20 HYDER SPNW', 'K22 HYDER NA', 'K23 BETTON ISLAND', 'K24 KETCHIKAN CORE NA', 'K25  SOUTH REVILLA', 'K27 MARGARET BAY', 
                     'K28 NAHA BAY']


        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['USELOCATION'].str.upper()



        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
        
        start = int(startYear)
        end = int(endYear)

        """ Calculating the total sum number of clients by use area and by season. But because the same group of people can visit two stop locations in the same use are, if we only summed client number
        we would count that group twice. But if a group of people visited two use areas in one trip, we need to count them twice. So we sum the total nubmer of clients by use area and season, then make 
        sure that sum is not greater than the total number of clinets on the day by use area and season. If the client number sum is greater than the total clents on the day, we use total clients on 
        the day. """

        guideDayCol = getFieldNames(guideDay)
        guideDayWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        guideDayData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideDayWhere)]
        guideDayDF = pd.DataFrame(guideDayData, columns=guideDayCol)
       
        guideStopCol = getFieldNames(guideStop)
        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol) 
        
        guideActCol = getFieldNames(guideActivity)
        guideActData = [row for row in arcpy.da.SearchCursor(guideActivity, guideActCol)]
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)

        guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
        guideDF = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')                 

        guideDF['Year'] = guideDF['TRIPDATE'].dt.year
           
        guideDF['USELOCATION'].str.upper()
        guideLoc = guideDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID']) 
        
        guideLoc['keep'] = guideLoc['FULLNAME'].apply(lambda x: 'True' if x in kmrd2 else 'False')
        
        guideLoc1 = guideLoc[guideLoc["keep"].str.contains("True")]
        
        guideActGroup = guideLoc1.groupby(['DAY_GUID', 'FULLNAME'], as_index = False).agg({'Year': 'first','SEASON':'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
        guideActGroup.loc[guideActGroup['CLIENTNUMBER'] > guideActGroup['TOTALCLIENTSONDAY'], 'USE'] = guideActGroup['TOTALCLIENTSONDAY']
        guideActGroup['USE'].fillna(guideActGroup['CLIENTNUMBER'], axis = 0, inplace = True)
        
        guideRec = guideActGroup[['SEASON', 'FULLNAME', 'Year', 'USE']]
         
        
        huntDayCol = getFieldNames(huntDay)
        huntDayWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntDayWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns=huntDayCol)
       
        huntStopCol = getFieldNames(huntStop)
        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol)]  
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)
        
        huntActCol = getFieldNames(huntActivity)
        huntActData = [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol)] 
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
        
        huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
        huntDF = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')
        huntDF['Year'] = huntDF['TRIPDATE'].dt.year        
      
        huntDF['USELOCATION'].str.upper()
        huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
        
        huntLoc['keep'] = huntLoc['FULLNAME'].apply(lambda x: 'True' if x in kmrd2 else 'False')
        
        huntLoc1 = huntLoc[huntLoc["keep"].str.contains("True")]
        
        huntActGroup = huntLoc1.groupby(['DAY_GUID', 'FULLNAME'], as_index = False).agg({'Year': 'first', 'SEASON' : 'first', 'TOTALCLIENTSONDAY': 'max', 'CLIENTNUMBER': 'sum'})
        huntActGroup.loc[huntActGroup['CLIENTNUMBER'] > huntActGroup['TOTALCLIENTSONDAY'], 'USE'] = huntActGroup['TOTALCLIENTSONDAY']
        huntActGroup['USE'].fillna(huntActGroup['TOTALCLIENTSONDAY'], axis = 0, inplace = True)        
        
        hunting = huntActGroup[['SEASON', 'FULLNAME', 'Year', 'USE']]
        

        
        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)
        outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
        outfitDF['USELOCATION'].str.upper()
        outfitLoc = outfitDF.merge(locDF, how = 'inner', on = ['LOCATION_ID', 'USELOCATION'])
        outfitLoc.rename(columns = {'TOTALCLIENTSONDAY':'USE'}, inplace=True)
        
        outfitLoc['keep']= outfitLoc['FULLNAME'].apply(lambda x: 'True' if x in kmrd2 else 'False')

        outfitLoc1 = outfitLoc[outfitLoc["keep"].str.contains("True")]        
        
        outfitting = outfitLoc1[['SEASON', 'FULLNAME', 'Year', 'USE']] 
        
        allDF = guideRec.append([hunting, outfitting])
        

        allPivot = pd.pivot_table(allDF, index = ['SEASON', 'FULLNAME'], columns = ['Year'], values = ['USE'], aggfunc=np.sum, margins = False, dropna = True)
        
        allPivot.fillna(0, inplace=True)
        allPivot.reset_index(inplace=True)
        allPivot['ActualUseAvg'] = (allPivot.sum(axis=1))/5
        allPivot = allPivot.round()

        writer_args = {
            'path': path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        def update(season, fullname, startNumb, columnNumb):
            x = allPivot.loc[(allPivot.SEASON == season) & (allPivot.FULLNAME==fullname), 'ActualUseAvg']
            for row, text in enumerate(x, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)

        with pd.ExcelWriter(**writer_args) as xlsx:
            wb = openpyxl.load_workbook(path)
            ws = wb['AllocStatus']
            xlsx.book = wb
            
            update('Spring', 'K04 DUKE ISLAND', 3, 4)
            update('Spring', 'K05 SOUTH MISTY LAKES', 4, 4)
            update('Spring', 'K06 MISTY CORE LAKES', 5, 4)
            update('Spring', 'K07 WALKER CHICKAMIN', 6, 4)
            update('Spring', 'K08 BURROUGHS UNUK', 7, 4)
            update('Spring', 'K09 ALAVA PRINCESS MANZANITA', 8, 4)
            update('Spring', 'K10 RUDYERD WINSTANLEY', 9, 4)
            update('Spring', 'K11 GRAVINA ISLAND', 10, 4)
            update('Spring', 'K12 BELL ISLAND', 11, 4)
            update('Spring', 'K13 EAST CLEVELAND', 12, 4)
            update('Spring', 'K14 WEST CLEVELAND', 13, 4)
            update('Spring', 'K15 WILSON BAKEWELL', 14, 4)
            update('Spring', 'K16 KETCHIKAN CORE SPNW', 15, 4)
            update('Spring', 'K17 GEORGE CARROLL THORNE', 16, 4)
            update('Spring', 'K18 CENTRAL REVILLA SPNW', 17, 4)
            update('Spring', 'K19 NORTH REVILLA', 18, 4)
            update('Spring', 'K20 HYDER SPNW', 19, 4)
#            update('Spring', 'K21', 20, 4)
            update('Spring', 'K22 HYDER NA', 21, 4)
            update('Spring', 'K23 BETTON ISLAND', 22, 4)
            update('Spring', 'K24 KETCHIKAN CORE NA', 23, 4)
            update('Spring', 'K25  SOUTH REVILLA', 24, 4)
#            update('Spring', 'K26', 25, 4)
            update('Spring', 'K27 MARGARET BAY', 26, 4)
            update('Spring', 'K28 NAHA BAY', 27, 4)
            
            update('Summer', 'K04 DUKE ISLAND', 3, 9)
            update('Summer', 'K05 SOUTH MISTY LAKES', 4, 9)
            update('Summer', 'K06 MISTY CORE LAKES', 5, 9)
            update('Summer', 'K07 WALKER CHICKAMIN', 6, 9)
            update('Summer', 'K08 BURROUGHS UNUK', 7, 9)
            update('Summer', 'K09 ALAVA PRINCESS MANZANITA', 8, 9)
            update('Summer', 'K10 RUDYERD WINSTANLEY', 9, 9)
            update('Summer', 'K11 GRAVINA ISLAND', 10, 9)
            update('Summer', 'K12 BELL ISLAND', 11, 9)
            update('Summer', 'K13 EAST CLEVELAND', 12, 9)
            update('Summer', 'K14 WEST CLEVELAND', 13, 9)
            update('Summer', 'K15 WILSON BAKEWELL', 14, 9)
            update('Summer', 'K16 KETCHIKAN CORE SPNW', 15, 9)
            update('Summer', 'K17 GEORGE CARROLL THORNE', 16, 9)
            update('Summer', 'K18 CENTRAL REVILLA SPNW', 17, 9)
            update('Summer', 'K19 NORTH REVILLA', 18, 9)
            update('Summer', 'K20 HYDER SPNW', 19, 9)
#            update('Summer', 'K21', 20, 9)
            update('Summer', 'K22 HYDER NA', 21, 9)
            update('Summer', 'K23 BETTON ISLAND', 22, 9)
            update('Summer', 'K24 KETCHIKAN CORE NA', 23, 9)
            update('Summer', 'K25  SOUTH REVILLA', 24, 9)
#            update('Summer', 'K26', 25, 9)
            update('Summer', 'K27 MARGARET BAY', 26, 9)
            update('Summer', 'K28 NAHA BAY', 27, 9)   
            
            update('Fall', 'K04 DUKE ISLAND', 3, 14)
            update('Fall', 'K05 SOUTH MISTY LAKES', 4, 14)
            update('Fall', 'K06 MISTY CORE LAKES', 5, 14)
            update('Fall', 'K07 WALKER CHICKAMIN', 6, 14)
            update('Fall', 'K08 BURROUGHS UNUK', 7, 14)
            update('Fall', 'K09 ALAVA PRINCESS MANZANITA', 8, 14)
            update('Fall', 'K10 RUDYERD WINSTANLEY', 9, 14)
            update('Fall', 'K11 GRAVINA ISLAND', 10, 14)
            update('Fall', 'K12 BELL ISLAND', 11, 14)
            update('Fall', 'K13 EAST CLEVELAND', 12, 14)
            update('Fall', 'K14 WEST CLEVELAND', 13, 14)
            update('Fall', 'K15 WILSON BAKEWELL', 14, 14)
            update('Fall', 'K16 KETCHIKAN CORE SPNW', 15, 14)
            update('Fall', 'K17 GEORGE CARROLL THORNE', 16, 14)
            update('Fall', 'K18 CENTRAL REVILLA SPNW', 17, 14)
            update('Fall', 'K19 NORTH REVILLA', 18, 14)
            update('Fall', 'K20 HYDER SPNW', 19, 14)
#            update('Fall', 'K21', 20, 14)
            update('Fall', 'K22 HYDER NA', 21, 14)
            update('Fall', 'K23 BETTON ISLAND', 22, 14)
            update('Fall', 'K24 KETCHIKAN CORE NA', 23, 14)
            update('Fall', 'K25  SOUTH REVILLA', 24, 14)
#            update('Fall', 'K26', 25, 14)
            update('Fall', 'K27 MARGARET BAY', 26, 14)
            update('Fall', 'K28 NAHA BAY', 27, 14)            

            colHead = 'Actual Use (5-yr Avg, {}-{})'.format(start, end)
            ws.cell(row=2, column=4).value = colHead
            ws.cell(row=2, column=9).value = colHead
            ws.cell(row=2, column=14).value = colHead
            

            wb.save(path)

        return 


class ConfirmActualUse(object):
    
    """Confirm Actual Use Report Shows what businesses have submitted use by Forest"""
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Confirm Use Report Summary"
        self.description = ""
        self.canRunInBackground = False
        
    def getParameterInfo(self):
        """Define parameter definitions"""

        param0 = arcpy.Parameter(
            displayName="Forest",
            name="forest",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param0.filter.type = "ValueList"
        param0.filter.list = ['Tongass', 'Chugach']

        param1  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param1.value = startyear -5
        
        param2  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param2.value = endyear


        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]
        
        return params       
    
    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        """Modify the values and properties of parameters before internal
        validation is performed.  This method is called whenever a parameter
        has been changed."""
        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        """The source code of the tool."""
        
        forestname = parameters[0].value
        startYear = parameters[1].value
        endYear = parameters[2].value
        savepath = parameters[3].value.value
        
        start = str(int(startYear))
        end = str(int(endYear))


        self.savefile = "ConfirmUseReportSummary_" + start + "_" + end + ".xlsx"
        self.path = os.path.join(savepath, self.savefile)


        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames 
 
                
        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

            
        iceTable = '{}S_R10.R10_OFG_ICEFIELD_TRIPMONTH'.format(connection) 

        guideTable = '{}S_R10.R10_OFG_GUIDEDREC_TRIP'.format(connection)
 
        huntTable = '{}S_R10.R10_OFG_HUNTING_TRIP'.format(connection)
        
        heliTable ='{}S_R10.R10_OFG_HELISKI_TRIP'.format(connection)

        mendTable = '{}S_R10.R10_OFG_MENDENHALL_TRIPMONTH'.format(connection)

        outfitTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)



        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))


        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))   

        iceCol = getFieldNames(iceTable)
        iceWhere = "REPORTYEAR >= {} AND REPORTYEAR <= {}".format(startYear, endYear)
        iceData =[row for row in arcpy.da.SearchCursor(iceTable, iceCol, where_clause=iceWhere)]
        iceDF = pd.DataFrame(iceData, columns=iceCol)
        iceDF['Year'] = iceDF['REPORTYEAR'] 
        iceDF.loc[iceDF['NO_OPERATION'] == 'Yes', 'Number of Trips'] = 0.1
        iceDF['Number of Trips'].fillna(1, inplace = True)

        icePivot = pd.pivot_table(iceDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')

                             
        guideCol = getFieldNames(guideTable)
        guideWhere = "STARTDATE >= timestamp '{}' AND ENDDATE <= timestamp '{}' AND FORESTNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), forestname)
       
        guideData = [row for row in arcpy.da.SearchCursor(guideTable, guideCol, where_clause=guideWhere)]
        guideDF = pd.DataFrame(guideData, columns=guideCol)
        
        guideDF['Year'] = guideDF['STARTDATE'].dt.year
        guideDF.loc[guideDF['NO_OPERATION'] == 'Yes', 'Number of Trips'] = 0.1
        guideDF['Number of Trips'].fillna(1, inplace = True)
        guidePivot = pd.pivot_table(guideDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')

        
        huntCol = getFieldNames(huntTable)
        huntWhere = "STARTDATE >= timestamp '{}' AND ENDDATE <= timestamp '{}' AND FORESTNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), forestname)
        huntData = [row for row in arcpy.da.SearchCursor(huntTable, huntCol, where_clause=huntWhere)]
        huntDF = pd.DataFrame(huntData, columns=huntCol)
        huntDF['Year'] = huntDF['STARTDATE'].dt.year
        huntDF.loc[huntDF['NO_OPERATION'] == 'YES', 'Number of Trips'] = 0.1
        huntDF['Number of Trips'].fillna(1, inplace = True)
        huntPivot = pd.pivot_table(huntDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')        


        heliCol = getFieldNames(heliTable)
        heliWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND FORESTNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), forestname)
        heliData = [row for row in arcpy.da.SearchCursor(heliTable, heliCol, where_clause=heliWhere)]
        heliDF = pd.DataFrame(heliData, columns=heliCol) 
        heliDF['Year'] = heliDF['TRIPDATE'].dt.year
        heliDF.loc[heliDF['NO_OPERATION'] == 'Yes', 'Number of Trips'] = 0.1
        heliDF['Number of Trips'].fillna(1, inplace = True) 
        heliPivot = pd.pivot_table(heliDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')        

       
        mendCol = getFieldNames(mendTable)
        mendWhere = "REPORTYEAR >= {} AND REPORTYEAR <= {}".format(startYear, endYear)
        mendData = [row for row in arcpy.da.SearchCursor(mendTable, mendCol, where_clause=mendWhere)]
        mendDF = pd.DataFrame(mendData, columns=mendCol)
        mendDF['Year'] = mendDF['REPORTYEAR']
        mendDF.loc[mendDF['NO_OPERATION'] == 'Yes', 'Number of Trips'] = 0.1
        mendDF['Number of Trips'].fillna(1, inplace = True)  
        mendPivot = pd.pivot_table(mendDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')        

        
        outfitCol = getFieldNames(outfitTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND FORESTNAME = '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), forestname)
        outfitData = [row for row in arcpy.da.SearchCursor(outfitTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)
        outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
        outfitDF.loc[outfitDF['NO_OPERATION'] == 'Yes', 'Number of Trips'] = 0.1
        outfitDF['Number of Trips'].fillna(1, inplace = True)  
        outfitPivot = pd.pivot_table(outfitDF, index = ['BUSINESSNAME'], columns = ['Year'], values = ['Number of Trips'], aggfunc=np.sum, margins = False, dropna = False, margins_name = 'Total')

        
        writer_args = {
            'path': self.path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='left', wrapText = True, vertical='center')
        index_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        title_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def update(data, startNumb, columnNumb, ws):
            for row, text in enumerate(data, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)
        
    
        with pd.ExcelWriter(**writer_args) as xlsx:

            guidePivot.to_excel(xlsx, 'Guided Rec')
            huntPivot.to_excel(xlsx, 'Guided Hunting')
            outfitPivot.to_excel(xlsx, 'Outfitting')
            heliPivot.to_excel(xlsx, 'Heliski')            
            if forestname == 'Tongass':
                icePivot.to_excel(xlsx, 'Icefield')
                mendPivot.to_excel(xlsx, 'Mendenhall')
                iceWS = xlsx.sheets['Icefield']
                mendWS = xlsx.sheets['Mendenhall']


            guideWS = xlsx.sheets['Guided Rec']
            huntWS = xlsx.sheets['Guided Hunting']
            outfitWS = xlsx.sheets['Outfitting']
            heliWS = xlsx.sheets['Heliski']
            
      

            
            def excelRow(ws, rows, styles):
                for row in ws[rows]:
                    for cell in row:
                        cell.style = styles
            
            def excelStyle(ws, cells, styles):
                for cell in ws[cells]:
                    cell.style = styles
                    
            def noUse(ws):
                for r in range(3,ws.max_row+1):
                    for c in range(2,ws.max_column+1):
                        s = str(ws.cell(r,c).value)
                        if s == '0.1': 
                            ws.cell(r,c).value = s.replace('0.1',"No Ops Year")                     
                    
            def wsLayout(ws):
                col=get_column_letter(ws.max_column)
                colInt= (ws.max_column)
                rows=ws.max_row
                value_col = 2
                value_cells = 'B3:{}{}'.format(col, rows)
                index_column = 'A'
                ws.column_dimensions[index_column].width = 42.57
                grey = PatternFill(bgColor="D9D9D9")
                diff_style = DifferentialStyle(fill = grey)
                rule = Rule(type="expression", dxf=diff_style)
                rule.formula = ["=AND(LEN($B2)>0,MOD($B2,2)=0)"]
                title_row = '1'
                title_row1 = '2'
                title_row2 = '3'                
                while value_col <= colInt:
                     i = get_column_letter(value_col)
                     ws.column_dimensions[i].width = 12.29
                     value_col += 1                
                excelRow(ws, value_cells, value_style)
                excelStyle(ws, index_column, index_style)
                excelStyle(ws, title_row, title_style)
                excelStyle(ws, title_row1, title_style) 
                excelStyle(ws, title_row2, title_style)
               
            def title(ws, datasetName):
                ws.insert_rows(1, amount= 6)
                title = 'Actual Use Report Submissions - Number of Trip Records Submitted by Year'
                reportDate = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
                forestTitle = 'Forest: {} National Forest'.format(forestname)
                dataTitle = 'Dataset: {}'.format(datasetName)
                ws['A1'] = title
                ws['A1'].font = Font(name='Calibri', size=11, bold=True, color='000000')
                ws['A1'].alignment = Alignment(horizontal='left', wrapText = True, vertical='center')
                ws['A3'] = reportDate
                ws['A4'] = forestTitle
                ws['A5'] = dataTitle
                ws.delete_rows(7, 1)
                
            def title1(ws, datasetName):
                ws.insert_rows(1, amount= 6)
                title = 'Actual Use Report Submissions - Number of Trip-Month Records Submitted by Year'
                reportDate = 'Report Generated: {}'.format(datetime.today().strftime('%m/%d/%Y'))
                forestTitle = 'Forest: {} National Forest'.format(forestname)
                dataTitle = 'Dataset: {}'.format(datasetName)
                ws['A1'] = title
                ws['A1'].font = Font(name='Calibri', size=11, bold=True, color='000000')
                ws['A1'].alignment = Alignment(horizontal='left', wrapText = True, vertical='center')
                ws['A3'] = reportDate
                ws['A4'] = forestTitle
                ws['A5'] = dataTitle
                ws.delete_rows(7, 1)
                    
            wsLayout(guideWS)
            wsLayout(huntWS)
            wsLayout(outfitWS)
            wsLayout(heliWS)
            if forestname == 'Tongass':  
                wsLayout(iceWS)
                wsLayout(mendWS) 
                        
            noUse(guideWS)
            noUse(heliWS)
            noUse(outfitWS)
            if forestname == 'Tongass':
                noUse(iceWS)
                noUse(mendWS)

            title(guideWS, 'Guided Recreation')
            title(huntWS, 'Guided Hunting')
            title(outfitWS, 'Outfitting')
            title(heliWS, 'Heliski')
            if forestname == 'Tongass':
                title1(iceWS, 'Icefields')
                title1(mendWS, 'Mendenhall')
            
        return
    
class WildernessSummary(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run Wilderness Summary"
        self.description = ""
        self.canRunInBackground = False
        self.params = self.getParameterInfo()
        
    def getParameterInfo(self):
        """Define parameter definitions"""

        param0 = arcpy.Parameter(
            displayName="Wilderness",
            name="wilderneess",
            datatype="GPString",
            parameterType="Required",
            direction="Input", 
            multiValue=True)

        param0.filter.type = "ValueList"
        param0.filter.list = ['Chuck River Wilderness', 'Coronation Island Wilderness', 'Karta River Wilderness', 'Kootznoowoo Wilderness', 'Kuiu Wilderness', 'Maurille Islands Wilderness', 
                              'Misty Fiords National Monument Wilderness', 'Nellie Juan-College Fiord WSA', 'Petersburg Creek-Duncan Salt Chuck Wilderness', 'Pleasant/Lemusurier/Inian Islands Wilderness' ,
                              'Russell Fjord Wilderness', 'South Baranof Wilderness', 'South Etolin Wilderness', 'South Prince of Wales Wilderness', 'Stikine-LeConte Wilderness',
                              'Tebenkof Bay Wilderness', 'Tracy Arm-Fords Terror Wilderness', 'Warren Island Wilderness', 'West Chichagof-Yakobi Wilderness' ]

        param1  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param1.value = startyear -5
        
        param2  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param2.value = endyear -1


        param3 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param3.filter.list = ["File System"]

        params = [param0, param1, param2, param3]
        
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):

        return 

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        
        """ Function to get list of field names in a table"""

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames
            
           

        """The source code of the tool."""
        
        wilderness = parameters[0].values
        startYear = parameters[1].value
        endYear = parameters[2].value
        savepath = parameters[3].value.value
        
        allWild = tuple(wilderness)
    
        start = str(int(startYear))
        end = str(int(endYear))

        self.savefile = "WildernessSummary_" + start + "_" + end + ".xlsx"
        self.path = os.path.join(savepath, self.savefile)

        username = os.environ.get("USERNAME")

#        pd.set_option("display.width", 5000)
#        outcheck = os.path.join(savepath, "Visitation_OutputCheck.txt")
#        f1 = open(outcheck, 'w')         
        
        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

        guideTrip = '{}S_R10.R10_OFG_GUIDEDREC_TRIP'.format(connection)
        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntTrip = '{}S_R10.R10_OFG_HUNTING_TRIP'.format(connection)
        huntHunt = '{}S_R10.R10_OFG_HUNTING_HUNTER'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)
        

            
        """ ReFormatting start and end date for Guided Recreation and Hunting so that it matches the timestamp format"""

        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
            
        
        """Creating Guided Recreation DataFrames and filtering data by the report timeframe"""
        
        guideTripCol = getFieldNames(guideTrip)
        guideDayCol = getFieldNames(guideDay)
        guideStopCol = getFieldNames(guideStop)
        guideActCol = getFieldNames(guideActivity)        
 
        guideTripData = [row for row in arcpy.da.SearchCursor(guideTrip, guideTripCol)]   
        guideTripDF = pd.DataFrame(guideTripData, columns = guideTripCol)

        guideWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        guideData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideWhere)]
        guideDayDF = pd.DataFrame(guideData, columns = guideDayCol)

        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol)

        guideActData =[row for row in arcpy.da.SearchCursor(guideActivity, guideActCol)]  
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)
        
        
        if (guideDayDF.empty == False):
            if (guideStopDF.empty ==False):
                if (guideActDF.empty == False):
                    if (guideTripDF. empty == False): 
                        guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
                        guideDF2 = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')
                        guideDF = guideDF2.merge(guideTripDF, how='inner', on='TRIP_GUID')
        
        """Creating Hunting Data Frames and filtering data by the report timeframe"""        

        huntTripCol = getFieldNames(huntTrip)
        huntDayCol = getFieldNames(huntDay)
        huntStopCol = getFieldNames(huntStop)
        huntActCol = getFieldNames(huntActivity)
    
        huntTripData = [row for row in arcpy.da.SearchCursor(huntTrip, huntTripCol)]  
        huntTripDF = pd.DataFrame(huntTripData, columns = huntTripCol)
        

        huntWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns = huntDayCol)

        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol)]
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)
         
        huntActData =  [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol)]  
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
        
        if (huntDayDF.empty == False):
            if (huntStopDF.empty == False):
                if (huntActDF.empty == False):
                    huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
                    huntDF2 = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')
                    huntDF = huntDF2.merge(huntTripDF, how = 'inner', on = 'TRIP_GUID')
                                           
        
        """Creating Outfitting data frame in the date range"""

        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)        


        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['FULLNAME'].fillna(locDF['ZONE_NAME'], inplace=True)
        locDF['USE_AREA'] = locDF['FULLNAME']
        locDF['USE_AREA'].fillna('Use Area Unknown', inplace=True)
        locDF['USELOCATION'].str.upper()
        locDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
        
        
        """ Activity Dictionary to account for differences in the activity categories and what is listed in each of the datasets. """
        
        activities = {'Boating, Stand Up Paddle Boarding, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Pack Rafting, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 
           'Boating, Raft, Canoe, Kayak or Other Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Canoeing (Mendenhall form)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Rafting (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Canoeing':'Boating (Non-Motorized, Freshwater)', 'Boating (Non-Motorized, Freshwater)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking':'Boating (Non-Motorized, Freshwater)', 'Rafting':'Boating (Non-Motorized, Freshwater)', 'Camping':'Camping','Dog Sledding':'Dog Sled Tours', 'Dog Sled Tours':'Dog Sled Tours', 'Flightseeing Landing Tours':'Flightseeing Landing Tours',
           'Freshwater Fishing':'Freshwater Fishing', 'Glacier Trekking':'Helicopter Landing Tours', 'Helicopter Landing Tours':'Helicopter Landing Tours', 'Heli-skiing Tours':'Heli-skiing Tours', 
           'Heliski':'Heli-skiing Tours', 'Hunting, Brown Bear':'Hunting, Brown Bear', 'Hunting, Deer':'Hunting, Deer', 'Remote Setting Nature Tour, on Foot': 'Remote Setting Nature Tour', 
           'Hunting, Elk':'Hunting, Elk', 'Hunting, Moose':'Hunting, Moose', 'Hunting, Mountain Goat':'Hunting, Mountain Goat', 'Remote Setting Nature Tour':'Remote Setting Nature Tour',
           'Hunting, Wolf':'Hunting, Waterfowl/Small game/Wolf - Service Day', 'Outfitting':'Outfitting (Delivery and/or pick-up of vehicles, equipment, etc. to/from National Forest System lands; Total per day - no limit on equipment numbers or number of trips)', 
           'Over-Snow Vehicle Tours':'Over-Snow Vehicle Tours', 'Bikepacking':'Remote Setting Nature Tour', 'Horseback Riding':'Remote Setting Nature Tour', 'Nature Tour, on Foot':'Remote Setting Nature Tour',
           'Nature Tour, on Ski':'Remote Setting Nature Tour','Nature Tour, Bicycle':'Remote Setting Nature Tour', 'Biking (Mendenhall form)':'Remote Setting Nature Tour', 'Biking':'Remote Setting Nature Tour', 
           'Hiking (Mendenhall form)':'Remote Setting Nature Tour', 'Hiking':'Remote Setting Nature Tour', 'Nature Tour, ATV/OHV':'Road Based Nature Tours', 'Nature Tour, Vehicle':'Road Based Nature Tours', 
           'Visitor Center (Begich Boggs, MGVC, SEADC)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Visitor Center Transport (Mendenhall form)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 
           'Visitor_Center_Transport':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Hunting, Black Bear ':'Hunting, Black Bear', 'Hunting, Dall Sheep ':'Hunting, Dall Sheep', 'Hunting, Waterfowl/Small game ':'Hunting, Waterfowl/Small game/Wolf - Service Day',
           'Assigned Site' : 'Assigned Site', 'Minimum Fee': 'Minimum Fee', 'Hunting, Black Bear':'Hunting, Black Bear', 'Hunting, Waterfowl/Small game/Wolf - Service Day Rate':'Hunting, Waterfowl/Small game/Wolf - Service Day'}        
        
        """ This takes the start year and end year parameters and reformats it so that it matches the date format in Guided Recreation and Hunting."""
        
        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
          
            
        if 'guideDF' not in locals():
            guide = pd.DataFrame()
            arcpy.AddMessage('There are no guided recreation activities for this time period.')   
        else:
            if (guideDF['USELOCATION'].loc() == guideDF['USELOCATION'].loc()):
                guideDF['USELOCATION'] = guideDF.loc(guideDF['USELOCATION'].map(lambda x: x.title()))
            guideDF['Activity2'] = guideDF['ACTIVITY'].map(activities)
            guideDF['Year'] = guideDF['TRIPDATE'].dt.year

            
            guideDF['USELOCATION'].str.upper()
            guideLoc = guideDF.merge(locDF, how = 'left', on = ['LOCATION_ID'])

            guideLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'MAXCLIENTS':'Total Clients', 'CLIENTNUMBER':'Activity Client Number'}, inplace=True)
            guide = guideLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number', 'WILDERNESSNAME', 'SEASON']].copy() 
            
            
        if 'huntDF' not in locals():
            if 'nonHuntDF' not in locals():
                hunt = pd.DataFrame()
                arcpy.AddMessage('There are no hunting associated activities for this time period.')   
        else:
            huntDF['Activity2'] = huntDF['ACTIVITY'].map(activities)
            huntDF['Year'] = huntDF['TRIPDATE'].dt.year           
            huntDF['USELOCATION'].str.upper()

            huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID']) 
            huntLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'MAXCLIENTS':'Total Clients', 'CLIENTNUMBER':'Activity Client Number'}, inplace=True)
            hunt = huntLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number', 'WILDERNESSNAME', 'SEASON']].copy()
            
            
        if (outfitDF.empty == True):
            outfit = pd.DataFrame()
            arcpy.AddMessage('There are no outfitting activities for this time period.')  
        else:
            outfitDF['Activity2'] = outfitDF['ACTIVITY'].map(activities)
            outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
            outfitDF['USELOCATION'].str.upper()
            outfitLoc = outfitDF.merge(locDF, how = 'left', on = ['LOCATION_ID']) 
            outfitLoc.rename(columns = {'TOTALCLIENTSONDAY':'Total Clients'}, inplace=True) 
            outfitLoc['Activity Client Number'] = outfitLoc.loc[:, 'Total Clients']
            outfitLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
            outfit = outfitLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number', 'WILDERNESSNAME', 'SEASON']].copy()
            
            
            allDF = hunt.append([outfit, guide])

            finalDF = allDF.query('WILDERNESSNAME in @wilderness')
            
            finalDF.rename(columns={'FORESTNAME':'Forest', 'DISTRICTNAME': 'Ranger District', 'Activity2':'Activity', 'USELOCATION':'Use Location', 'WILDERNESSNAME': 'Wilderness', 'SEASON':'Season'}, inplace=True)
                                
            forestGroup = finalDF.groupby(['TRIP_GUID', 'Year'], as_index = False).agg({'Wilderness' : 'first', 'Total Clients': 'max'})   
            forestGroup.rename(columns={'Total Clients':'Number of Clients'}, inplace = True)
            
            forestPivot = pd.pivot_table(forestGroup, index = ['Wilderness'], columns = ['Year'], values = ['Number of Clients'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')            

            locGroup = finalDF.groupby(['TRIP_GUID', 'Use Location'], as_index = False).agg({'Year': 'first', 'Wilderness' : 'first', 'Total Clients': 'max', 'Activity Client Number': 'sum'})
            locGroup.loc[locGroup['Activity Client Number'] > locGroup['Total Clients'], 'Number of Clients'] = locGroup['Total Clients']
            locGroup['Number of Clients'].fillna(locGroup['Activity Client Number'], axis = 0, inplace = True)     
            
            locPivot = pd.pivot_table(locGroup, index = ['Wilderness', 'Use Location'], columns = ['Year'], values = ['Number of Clients'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')


            seasonGroup = finalDF.groupby(['TRIP_GUID', 'Season'], as_index = False).agg({'Year': 'first', 'Wilderness' : 'first', 'Total Clients': 'max', 'Activity Client Number': 'sum'})
            seasonGroup.loc[seasonGroup['Activity Client Number'] > seasonGroup['Total Clients'], 'Number of Clients'] = seasonGroup['Total Clients']
            seasonGroup['Number of Clients'].fillna(seasonGroup['Activity Client Number'], axis = 0, inplace = True)   
            
            seasonPivot = pd.pivot_table(seasonGroup, index = ['Wilderness', 'Season'], columns = ['Year'], values = ['Number of Clients'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')         
            
            
        """ Writing pivot tables to excel and additional formatting. The activity tab and hunt tab have to be added with conditional statements because not all businesses will have hunts, and some
        businesses that have hunts do not have other activities.
        """

        writer_args = {
            'path': self.path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='left', wrapText = True, vertical='top')
        index_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        title_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def update(data, startNumb, columnNumb, ws):
            for row, text in enumerate(data, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)
            
        def headers(ws):
            header1 = 'Wilderness Summary Report'
            ws.oddHeader.left.text = header1
            ws.oddHeader.left.size = 14
            ws.oddHeader.left.font = "Calibri, bold"
            ws.oddHeader.left.color = "000000"
            
            footer = 'Report Generated: {}\nSource: R10 Outfitter/Guide Database'.format(datetime.today().strftime('%m/%d/%Y'))
            ws.oddFooter.left.text = footer
            
        def excelRow(ws, rows, styles):
            for row in ws[rows]:
                for cell in row:
                    cell.style = styles
        
        def excelStyle(ws, cells, styles):
            for cell in ws[cells]:
                cell.style = styles

        def excelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 3
            index_column = 'A'
            index_column1 = 'B'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'C4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 28
            ws.column_dimensions[index_column1].width = 28
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, index_column1, index_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)

            
        def singleExcelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 2
            index_column = 'A'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'B4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws) 
            moveWild = "A3:A{}".format(rows)
            ws.move_range(moveWild, rows=-1)
            moveCells = "B4:{}{}".format(col, rows)
            ws.move_range(moveCells, rows=-1)
        
    
        with pd.ExcelWriter(**writer_args) as xlsx:
            forestPivot.to_excel(xlsx, 'Clients by Wilderness')
            ws = xlsx.sheets['Clients by Wilderness'] 
            singleExcelUpdate(ws)
            seasonPivot.to_excel(xlsx, 'Clients By Season')
            wsAct = xlsx.sheets['Clients By Season'] 
            excelUpdate(wsAct)            
            locPivot.to_excel(xlsx, 'Clients by Location')
            wsLoc = xlsx.sheets['Clients by Location'] 
            excelUpdate(wsLoc)
            wsLoc.oddFooter.center.text = "Page &P of &N"
            wsLoc.evenFooter.center.text = "Page &P of &N"

            
        return
        
class HeliskiSummary(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run Heliski Summary"
        self.description = ""
        self.canRunInBackground = False
        self.params = self.getParameterInfo()
        
    def getParameterInfo(self):
        """Define parameter definitions"""

        param0 = arcpy.Parameter(
            displayName="Forest",
            name="forest",
            datatype="GPString",
            parameterType="Required",
            direction="Input", 
            multiValue=True)

        param0.filter.type = "ValueList"
        param0.filter.list = ['Tongass', 'Chugach']
    
        param1 = arcpy.Parameter(
            displayName = "Ranger District",
            name = "rangerDist",  
            datatype = "GPString", 
            parameterType="Required", 
            direction="Input", 
            multiValue=True)
        
    
        param1.filter.type = "ValueList"        
        param1.filter.list = []

        param2  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param2.value = startyear -5
        
        param3  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param3.value = endyear -1


        param4 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param4.filter.list = ["File System"]

        params = [param0, param1, param2, param3, param4]
        
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        
        allForests = ['Tongass', 'Chugach']
        
        if all(elem in parameters[0].values for elem in allForests):
            parameters[1].filter.list = ["Admiralty National Monument", "Craig Ranger District", "Hoonah Ranger District", "Juneau Ranger District", "Ketchikan - Misty Ranger District", 
                      "Petersburg Ranger District", "Sitka Ranger District", "Thorne Bay Ranger District", "Wrangell Ranger District", "Yakutat Ranger District",
                      "Cordova Ranger District", "Glacier Ranger District", "Seward Ranger District"]
        elif'Tongass' in parameters[0].values:
            parameters[1].filter.list = ["Admiralty National Monument", "Craig Ranger District", "Hoonah Ranger District", "Juneau Ranger District", "Ketchikan - Misty Ranger District", 
                      "Petersburg Ranger District", "Sitka Ranger District", "Thorne Bay Ranger District", "Wrangell Ranger District", "Yakutat Ranger District"]
        elif 'Chugach' in parameters[0].values:
            parameters[1].filter.list = ["Cordova Ranger District", "Glacier Ranger District", "Seward Ranger District"]
            

        return parameters

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        
        """ Function to get list of field names in a table"""

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames
            
           

        """The source code of the tool."""
        
        forest = parameters[0].values
        rangerDist = parameters[1].values
        startYear = parameters[2].value
        endYear = parameters[3].value
        savepath = parameters[4].value.value
        
        allForests = tuple(forest)
    
        start = str(int(startYear))
        end = str(int(endYear))

        self.savefile = "HeliskiSummary_" + start + "_" + end + ".xlsx"
        self.path = os.path.join(savepath, self.savefile)

        username = os.environ.get("USERNAME")
       
        
        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

        heliTrip ='{}S_R10.R10_OFG_HELISKI_TRIP'.format(connection)
        heliActivity = '{}S_R10.R10_OFG_HELISKI_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)
        

            
        """ ReFormatting start and end date for Guided Recreation and Hunting so that it matches the timestamp format"""

        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
                    
        """ Creating Heliski dataframes in the date range and for the specific business name"""

        heliTripCol = getFieldNames(heliTrip)
        heliActCol = getFieldNames(heliActivity)

        heliWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        heliData = [row for row in arcpy.da.SearchCursor(heliTrip, heliTripCol, where_clause=heliWhere)]
        heliTripDF = pd.DataFrame(heliData, columns=heliTripCol)   

        heliActData = [row for row in arcpy.da.SearchCursor(heliActivity, heliActCol)]
        heliActDF = pd.DataFrame(heliActData, columns = heliActCol)
        
        if (heliTripDF.empty == False):
            if (heliActDF.empty == False):
                heliDF = heliActDF.merge(heliTripDF, how = 'inner', on='TRIP_GUID')
        

        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['FULLNAME'].fillna(locDF['ZONE_NAME'], inplace=True)
        locDF['USE_AREA'] = locDF['FULLNAME']
        locDF['USE_AREA'].fillna('Use Area Unknown', inplace=True)
        locDF['USELOCATION'].str.upper()
        locDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
        
        
        """ Activity Dictionary to account for differences in the activity categories and what is listed in each of the datasets. """
        
        activities = {'Boating, Stand Up Paddle Boarding, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Pack Rafting, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 
           'Boating, Raft, Canoe, Kayak or Other Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Canoeing (Mendenhall form)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Rafting (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Canoeing':'Boating (Non-Motorized, Freshwater)', 'Boating (Non-Motorized, Freshwater)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking':'Boating (Non-Motorized, Freshwater)', 'Rafting':'Boating (Non-Motorized, Freshwater)', 'Camping':'Camping','Dog Sledding':'Dog Sled Tours', 'Dog Sled Tours':'Dog Sled Tours', 'Flightseeing Landing Tours':'Flightseeing Landing Tours',
           'Freshwater Fishing':'Freshwater Fishing', 'Glacier Trekking':'Helicopter Landing Tours', 'Helicopter Landing Tours':'Helicopter Landing Tours', 'Heli-skiing Tours':'Heli-skiing Tours', 
           'Heliski':'Heli-skiing Tours', 'Hunting, Brown Bear':'Hunting, Brown Bear', 'Hunting, Deer':'Hunting, Deer', 'Remote Setting Nature Tour, on Foot': 'Remote Setting Nature Tour', 
           'Hunting, Elk':'Hunting, Elk', 'Hunting, Moose':'Hunting, Moose', 'Hunting, Mountain Goat':'Hunting, Mountain Goat', 'Remote Setting Nature Tour':'Remote Setting Nature Tour',
           'Hunting, Wolf':'Hunting, Waterfowl/Small game/Wolf - Service Day', 'Outfitting':'Outfitting (Delivery and/or pick-up of vehicles, equipment, etc. to/from National Forest System lands; Total per day - no limit on equipment numbers or number of trips)', 
           'Over-Snow Vehicle Tours':'Over-Snow Vehicle Tours', 'Bikepacking':'Remote Setting Nature Tour', 'Horseback Riding':'Remote Setting Nature Tour', 'Nature Tour, on Foot':'Remote Setting Nature Tour',
           'Nature Tour, on Ski':'Remote Setting Nature Tour','Nature Tour, Bicycle':'Remote Setting Nature Tour', 'Biking (Mendenhall form)':'Remote Setting Nature Tour', 'Biking':'Remote Setting Nature Tour', 
           'Hiking (Mendenhall form)':'Remote Setting Nature Tour', 'Hiking':'Remote Setting Nature Tour', 'Nature Tour, ATV/OHV':'Road Based Nature Tours', 'Nature Tour, Vehicle':'Road Based Nature Tours', 
           'Visitor Center (Begich Boggs, MGVC, SEADC)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Visitor Center Transport (Mendenhall form)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 
           'Visitor_Center_Transport':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Hunting, Black Bear ':'Hunting, Black Bear', 'Hunting, Dall Sheep ':'Hunting, Dall Sheep', 'Hunting, Waterfowl/Small game ':'Hunting, Waterfowl/Small game/Wolf - Service Day',
           'Assigned Site' : 'Assigned Site', 'Minimum Fee': 'Minimum Fee', 'Hunting, Black Bear':'Hunting, Black Bear', 'Hunting, Waterfowl/Small game/Wolf - Service Day Rate':'Hunting, Waterfowl/Small game/Wolf - Service Day'}        
        
        """ This takes the start year and end year parameters and reformats it so that it matches the date format in Guided Recreation and Hunting."""
        
        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
            
            
        if 'heliDF' not in locals():
            heli = pd.DataFrame()
            arcpy.AddMessage('There are no heliski activities for this time period.') 
        else:        
            heliDF['Activity2'] = heliDF['ACTIVITY'].map(activities)
            heliDF['Year'] = heliDF['TRIPDATE'].dt.year
            heliDF['USELOCATION'].str.upper()
            heliLoc = heliDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
            heliLoc.rename(columns = {'USELOCATION_x':'USELOCATION', 'TOTALCLIENTSONDAY':'Total Clients', 'CLIENTS_LOCATION':'Activity Client Number'}, inplace=True)
            heli = heliLoc[['Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number', 'FULLNAME']].copy()

            finalDF = heli.query('DISTRICTNAME in @rangerDist')
            
            finalDF.rename(columns={'FORESTNAME':'Forest', 'DISTRICTNAME': 'Ranger District', 'USELOCATION':'Use Location', 'FULLNAME':'Use Area'}, inplace=True)

            locGroup = finalDF.groupby(['TRIP_GUID', 'Use Area'], as_index = False).agg({'Ranger District':'first', 'Year': 'first', 'Forest':'first', 'Total Clients': 'max', 'Activity Client Number': 'sum'})
            locGroup.loc[locGroup['Activity Client Number'] > locGroup['Total Clients'], 'Clients By Heliski Region/Zone'] = locGroup['Total Clients']
            locGroup['Clients By Heliski Region/Zone'].fillna(locGroup['Clients By Heliski Region/Zone'], axis = 0, inplace = True)     
            
            locPivot = pd.pivot_table(locGroup, index = ['Ranger District', 'Use Area'], columns = ['Year'], values = ['Clients By Heliski Region/Zone'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')
        
            
            
        """ Writing pivot tables to excel and additional formatting. The activity tab and hunt tab have to be added with conditional statements because not all businesses will have hunts, and some
        businesses that have hunts do not have other activities.
        """

        writer_args = {
            'path': self.path,
            'mode': 'w',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='left', wrapText = True, vertical='center')
        index_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        title_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def update(data, startNumb, columnNumb, ws):
            for row, text in enumerate(data, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)
                
        def headers(ws):            
            footer = 'Prepared By: {}\nDate: {}'.format(username, datetime.today().strftime('%m/%d/%Y'))
            ws.oddFooter.left.text = footer
            
        def excelRow(ws, rows, styles):
            for row in ws[rows]:
                for cell in row:
                    cell.style = styles
        
        def excelStyle(ws, cells, styles):
            for cell in ws[cells]:
                cell.style = styles

        def excelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 3
            index_column = 'A'
            index_column1 = 'B'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'C4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            ws.column_dimensions[index_column1].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, index_column1, index_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)
            
        def singleExcelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 2
            index_column = 'A'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'B4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)            
        
    
        with pd.ExcelWriter(**writer_args) as xlsx:            
            locPivot.to_excel(xlsx, 'Clients by Use Area')
            wsLoc = xlsx.sheets['Clients by Use Area'] 
            excelUpdate(wsLoc)

            
        return  
        
class VisitationSummary(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "Run Visitation Summary"
        self.description = ""
        self.canRunInBackground = False
        self.params = self.getParameterInfo()
        
    def getParameterInfo(self):
        """Define parameter definitions"""

        param0 = arcpy.Parameter(
            displayName="For trips conducted on",
            name="forest",
            datatype="GPString",
            parameterType="Required",
            direction="Input", 
            multiValue=True)

        param0.filter.type = "ValueList"
        param0.filter.list = ['Tongass', 'Chugach']
        
        param1 = arcpy.Parameter(
            displayName = "Ranger District",
            name = "rangerDist",  
            datatype = "GPString", 
            parameterType="Required", 
            direction="Input", 
            multiValue=True)
        
    
        param1.filter.type = "ValueList"        
        param1.filter.list = []

        param2  = arcpy.Parameter(
            displayName="Start Year",
            name="startyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        startyear = int(datetime.today().year)
        param2.value = startyear -5
        
        param3  = arcpy.Parameter(
            displayName="End Year",
            name="endyear",
            datatype="GPDouble",
            parameterType="Required",
            direction="Input")

        endyear = int(datetime.today().year)
        param3.value = endyear -1


        param4 = arcpy.Parameter(
            displayName="Save report to",
            name="save_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")
        param4.filter.list = ["File System"]

        params = [param0, param1, param2, param3, param4]
        
        return params

    def isLicensed(self):
        """Set whether tool is licensed to execute."""
        return True

    def updateParameters(self, parameters):
        
        forestList = ['Tongass', 'Chugach']
        
        if all(elem in parameters[0].values for elem in forestList):
            parameters[1].filter.list = ["Admiralty National Monument", "Craig Ranger District", "Hoonah Ranger District", "Juneau Ranger District", "Ketchikan - Misty Ranger District", 
                      "Petersburg Ranger District", "Sitka Ranger District", "Thorne Bay Ranger District", "Wrangell Ranger District", "Yakutat Ranger District",
                      "Cordova Ranger District", "Glacier Ranger District", "Seward Ranger District"]
        elif'Tongass' in parameters[0].values:
            parameters[1].filter.list = ["Admiralty National Monument", "Craig Ranger District", "Hoonah Ranger District", "Juneau Ranger District", "Ketchikan - Misty Ranger District", 
                      "Petersburg Ranger District", "Sitka Ranger District", "Thorne Bay Ranger District", "Wrangell Ranger District", "Yakutat Ranger District"]
        elif 'Chugach' in parameters[0].values:
            parameters[1].filter.list = ["Cordova Ranger District", "Glacier Ranger District", "Seward Ranger District"]
            

        return parameters

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool
        parameter.  This method is called after internal validation."""
        return

    def execute(self, parameters, messages):
        
        """ Function to get list of field names in a table"""

        def getFieldNames(shp):
                fieldnames = [f.name for f in arcpy.ListFields(shp)]
                return fieldnames
            
           

        """The source code of the tool."""
        
        forest = parameters[0].values
        rangerDist = parameters[1].values
        startYear = parameters[2].value
        endYear = parameters[3].value
        savepath = parameters[4].value.value
        
        forestList = ['Tongass', 'Chugach']

        if all(elem in parameters[0].values for elem in forestList):        
            allForests = tuple(forest)
        else: allForests = ("'" + forest[0]+ "'")
        

    
        start = str(int(startYear))
        end = str(int(endYear))

        self.savefile = "VisitationSummary_" + start + "_" + end + ".xlsx"
        self.path = os.path.join(savepath, self.savefile)

        username = os.environ.get("USERNAME")

#        pd.set_option("display.width", 5000)
#        outcheck = os.path.join(savepath, "Visitation_OutputCheck.txt")
#        f1 = open(outcheck, 'w')         
        
        connection = r'T:/FS/Reference/GeoTool/r10/DatabaseConnection/r10_default_as_myself.sde/'

        iceTrip = '{}S_R10.R10_OFG_ICEFIELD_TRIPMONTH'.format(connection) 
        iceAct = '{}S_R10.R10_OFG_ICEFIELD_ACTIVITY'.format(connection)
        guideTrip = '{}S_R10.R10_OFG_GUIDEDREC_TRIP'.format(connection)
        guideDay = '{}S_R10.R10_OFG_GUIDEDREC_DAY'.format(connection)
        guideStop = '{}S_R10.R10_OFG_GUIDEDREC_STOP'.format(connection)
        guideActivity = '{}S_R10.R10_OFG_GUIDEDREC_ACTIVITY'.format(connection)
        huntTrip = '{}S_R10.R10_OFG_HUNTING_TRIP'.format(connection)
        huntHunt = '{}S_R10.R10_OFG_HUNTING_HUNTER'.format(connection)
        huntDay = '{}S_R10.R10_OFG_HUNTING_DAY'.format(connection)
        huntStop = '{}S_R10.R10_OFG_HUNTING_STOP'.format(connection)
        huntActivity = '{}S_R10.R10_OFG_HUNTING_ACTIVITY'.format(connection)
        heliTrip ='{}S_R10.R10_OFG_HELISKI_TRIP'.format(connection)
        heliActivity = '{}S_R10.R10_OFG_HELISKI_ACTIVITY'.format(connection)
        mendTrip = '{}S_R10.R10_OFG_MENDENHALL_TRIPMONTH'.format(connection)
        mendActivity = '{}S_R10.R10_OFG_MENDENHALL_ACTIVITY'.format(connection)
        outTable ='{}S_R10.R10_OFG_OUTFITTING_ACTIVITY'.format(connection)
        locTable = '{}S_R10.R10_OFG_LOCATION'.format(connection)
        

            
        """ ReFormatting start and end date for Guided Recreation and Hunting so that it matches the timestamp format"""

        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
            
        """Creating Icefield dataframes and filtering data by the report timeframe and business name"""

        iceTripCol = getFieldNames(iceTrip)
        iceActCol = getFieldNames(iceAct)
        
        iceWhere = "REPORTYEAR>= {} AND REPORTYEAR<= {}".format(startYear, endYear)
        iceTripData =[row for row in arcpy.da.SearchCursor(iceTrip, iceTripCol, where_clause=iceWhere)]

        iceActWhere = "FORESTNAME IN {}".format(allForests) 
        iceActData = [row for row in arcpy.da.SearchCursor(iceAct, iceActCol, where_clause = iceActWhere)]

        iceTripDF =  pd.DataFrame(iceTripData, columns = iceTripCol)
        iceActDF = pd.DataFrame(iceActData, columns = iceActCol)
        
        if (iceTripDF.empty == False):
            if (iceActDF.empty == False):
                iceDF = iceActDF.merge(iceTripDF, how = 'inner', on = 'TRIP_GUID')
        
        """Creating Guided Recreation DataFrames and filtering data by the report timeframe and business name"""
        
        guideTripCol = getFieldNames(guideTrip)
        guideDayCol = getFieldNames(guideDay)
        guideStopCol = getFieldNames(guideStop)
        guideActCol = getFieldNames(guideActivity)
        
        guideTripWhere = "FORESTNAME IN {}".format(allForests) 
        guideTripData = [row for row in arcpy.da.SearchCursor(guideTrip, guideTripCol, where_clause=guideTripWhere)]   
        guideTripDF = pd.DataFrame(guideTripData, columns = guideTripCol)

        guideWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        guideData = [row for row in arcpy.da.SearchCursor(guideDay, guideDayCol, where_clause=guideWhere)]
        guideDayDF = pd.DataFrame(guideData, columns = guideDayCol)

        guideStopWhere = "FORESTNAME IN {}".format(allForests) 
        guideStopData = [row for row in arcpy.da.SearchCursor(guideStop, guideStopCol)]
        guideStopDF = pd.DataFrame(guideStopData, columns = guideStopCol)

        guideActData =[row for row in arcpy.da.SearchCursor(guideActivity, guideActCol)]  
        guideActDF = pd.DataFrame(guideActData, columns = guideActCol)
        
        
        if (guideDayDF.empty == False):
            if (guideStopDF.empty ==False):
                if (guideActDF.empty == False):
                    if (guideTripDF. empty == False): 
                        guideDF1 = guideActDF.merge(guideStopDF, how = 'inner', on= 'STOP_GUID')
                        guideDF2 = guideDF1.merge(guideDayDF, how = 'inner', on = 'DAY_GUID')
                        guideDF = guideDF2.merge(guideTripDF, how='inner', on='TRIP_GUID')
        
        """Creating Hunting Data Frames to only get the non-hunting activities and filtering data by the report timeframe and business name"""        

        huntTripCol = getFieldNames(huntTrip)
        huntDayCol = getFieldNames(huntDay)
        huntStopCol = getFieldNames(huntStop)
        huntActCol = getFieldNames(huntActivity)
    
        huntTripWhere = "FORESTNAME IN {}".format(allForests) 
        huntTripData = [row for row in arcpy.da.SearchCursor(huntTrip, huntTripCol, where_clause=huntTripWhere)]  
        huntTripDF = pd.DataFrame(huntTripData, columns = huntTripCol)
        

        huntWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}'".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'))
        huntDayData = [row for row in arcpy.da.SearchCursor(huntDay, huntDayCol, where_clause=huntWhere)]
        huntDayDF = pd.DataFrame(huntDayData, columns = huntDayCol)

        huntStopData = [row for row in arcpy.da.SearchCursor(huntStop, huntStopCol)]
        huntStopDF = pd.DataFrame(huntStopData, columns = huntStopCol)
         
        huntActData =  [row for row in arcpy.da.SearchCursor(huntActivity, huntActCol)]  
        huntActDF = pd.DataFrame(huntActData, columns = huntActCol)
        
        if (huntDayDF.empty == False):
            if (huntStopDF.empty == False):
                if (huntActDF.empty == False):
                    huntDF1 = huntActDF.merge(huntStopDF, how = 'inner', on = 'STOP_GUID')
                    huntDF2 = huntDF1.merge(huntDayDF, how= 'inner', on = 'DAY_GUID')
                    huntDF = huntDF2.merge(huntTripDF, how = 'inner', on = 'TRIP_GUID')
                                           
        
        """ Creating Heliski dataframes in the date range and for the specific business name"""

        heliTripCol = getFieldNames(heliTrip)
        heliActCol = getFieldNames(heliActivity)

        heliWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND FORESTNAME IN {}".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), allForests)
        heliData = [row for row in arcpy.da.SearchCursor(heliTrip, heliTripCol, where_clause=heliWhere)]
        heliTripDF = pd.DataFrame(heliData, columns=heliTripCol)   
        
        heliActData = [row for row in arcpy.da.SearchCursor(heliActivity, heliActCol)]
        heliActDF = pd.DataFrame(heliActData, columns = heliActCol)
        
        if (heliTripDF.empty == False):
            if (heliActDF.empty == False):
                heliDF = heliActDF.merge(heliTripDF, how = 'inner', on='TRIP_GUID')
        
        
        """Creating Mendenhall datafrmaes in the date range and for the specific business name"""

        mendTripCol = getFieldNames(mendTrip)
        mendActCol = getFieldNames(mendActivity)
        
        
        mendWhere = "REPORTYEAR >= {} AND REPORTYEAR <= {}".format(startYear, endYear)
        mendData = [row for row in arcpy.da.SearchCursor(mendTrip, mendTripCol, where_clause=mendWhere)]
        mendTripDF = pd.DataFrame(mendData, columns=mendTripCol)

        mendActWhere = "FORESTNAME IN {}".format(allForests) 
        mendActData = [row for row in arcpy.da.SearchCursor(mendActivity, mendActCol, where_clause=mendActWhere)]
        mendActDF = pd.DataFrame(mendActData, columns=mendActCol) 
        
        if (mendTripDF.empty == False):
            if (mendActDF.empty == False):
                mendDF = mendActDF.merge(mendTripDF, how='inner', on = 'TRIP_GUID')
        
        """Creating Outfitting data frame in the date range and for the specific business name"""

        outfitCol = getFieldNames(outTable)
        outfitWhere = "TRIPDATE >= timestamp '{}' AND TRIPDATE <= timestamp '{}' AND FORESTNAME IN {}".format(stReplace.strftime('%Y-%m-%d %H:%M:%S'), endReplace.strftime('%Y-%m-%d %H:%M:%S'), allForests)
        outfitData = [row for row in arcpy.da.SearchCursor(outTable, outfitCol, where_clause=outfitWhere)]
        outfitDF = pd.DataFrame(outfitData, columns=outfitCol)        


        locCol = getFieldNames(locTable)
        locData =[row for row in arcpy.da.SearchCursor(locTable, locCol)]
        locDF = pd.DataFrame(locData, columns = locCol)
        locDF['FULLNAME'].fillna(locDF['ZONE_NAME'], inplace=True)
        locDF['USE_AREA'] = locDF['FULLNAME']
        locDF['USE_AREA'].fillna('Use Area Unknown', inplace=True)
        locDF['USELOCATION'].str.upper()
        locDF['DISTRICTNAME'].fillna('Ranger District Unknown', inplace = True)
        
        
        """ Activity Dictionary to account for differences in the activity categories and what is listed in each of the datasets. """
        
        activities = {'Boating, Stand Up Paddle Boarding, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Pack Rafting, Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 
           'Boating, Raft, Canoe, Kayak or Other Non-Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Boating, Motorized, Freshwater':'Boating (Non-Motorized, Freshwater)', 'Canoeing (Mendenhall form)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Rafting (Mendenhall form)':'Boating (Non-Motorized, Freshwater)', 'Canoeing':'Boating (Non-Motorized, Freshwater)', 'Boating (Non-Motorized, Freshwater)':'Boating (Non-Motorized, Freshwater)',
           'Kayaking':'Boating (Non-Motorized, Freshwater)', 'Rafting':'Boating (Non-Motorized, Freshwater)', 'Camping':'Camping','Dog Sledding':'Dog Sled Tours', 'Dog Sled Tours':'Dog Sled Tours', 'Flightseeing Landing Tours':'Flightseeing Landing Tours',
           'Freshwater Fishing':'Freshwater Fishing', 'Glacier Trekking':'Helicopter Landing Tours', 'Helicopter Landing Tours':'Helicopter Landing Tours', 'Heli-skiing Tours':'Heli-skiing Tours', 
           'Heliski':'Heli-skiing Tours', 'Hunting, Brown Bear':'Hunting, Brown Bear', 'Hunting, Deer':'Hunting, Deer', 'Remote Setting Nature Tour, on Foot': 'Remote Setting Nature Tour, on Foot', 
           'Hunting, Elk':'Hunting, Elk', 'Hunting, Moose':'Hunting, Moose', 'Hunting, Mountain Goat':'Hunting, Mountain Goat', 'Remote Setting Nature Tour':'Remote Setting Nature Tour',
           'Hunting, Wolf':'Hunting, Waterfowl/Small game/Wolf - Service Day', 'Outfitting':'Outfitting (Delivery and/or pick-up of vehicles, equipment, etc. to/from National Forest System lands; Total per day - no limit on equipment numbers or number of trips)', 
           'Over-Snow Vehicle Tours':'Over-Snow Vehicle Tours', 'Bikepacking':'Bikepacking', 'Horseback Riding':'Horseback Riding', 'Nature Tour, on Foot':'Nature Tour, on Foot',
           'Nature Tour, on Ski':'Nature Tour, on Ski','Nature Tour, Bicycle':'Nature Tour, Bicycle', 'Biking (Mendenhall form)':'Nature Tour, Bicycle', 'Biking':'Nature Tour, Bicycle', 
           'Hiking (Mendenhall form)':'Nature Tour, on Foot', 'Hiking':'Nature Tour, on Foot', 'Nature Tour, ATV/OHV':'Nature Tour, ATV/OHV', 'Nature Tour, Vehicle':'Nature Tour, Vehicle', 
           'Visitor Center (Begich Boggs, MGVC, SEADC)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Visitor Center Transport (Mendenhall form)':'Visitor Center (Begich Boggs, MGVC, SEADC)', 
           'Visitor_Center_Transport':'Visitor Center (Begich Boggs, MGVC, SEADC)', 'Hunting, Black Bear ':'Hunting, Black Bear', 'Hunting, Dall Sheep ':'Hunting, Dall Sheep', 'Hunting, Waterfowl/Small game ':'Hunting, Waterfowl/Small game/Wolf - Service Day',
           'Assigned Site' : 'Assigned Site', 'Minimum Fee': 'Minimum Fee', 'Hunting, Black Bear':'Hunting, Black Bear', 'Hunting, Waterfowl/Small game/Wolf - Service Day Rate':'Hunting, Waterfowl/Small game/Wolf - Service Day'}        
        
        """ This takes the start year and end year parameters and reformats it so that it matches the date format in Guided Recreation and Hunting."""
        
        startdate = '1980-01-01 00:00:00'
        st = datetime.strptime(startdate, '%Y-%m-%d %H:%M:%S')
        stReplace = st.replace(year = int(startYear))

        enddate = '1980-12-31 23:59:59'
        end = datetime.strptime(enddate, '%Y-%m-%d %H:%M:%S')
        endReplace = end.replace(year = int(endYear))
        
        if 'iceDF' not in locals():
            ice = pd.DataFrame()
            arcpy.AddMessage('This business had no icefield activities for this time period.')
        else:
            iceDF.loc[iceDF['LDNGOPS'] >0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['LDNGGRATUITY']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['LDNGPAIDCLIENTS']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['CLIENTSGLACTREK']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF.loc[iceDF['CLIENTSDOGSLED']>0, 'ACTIVITY'] = 'Dog Sled Tours'
            iceDF.loc[iceDF['CLIENTSHIKE']>0, 'ACTIVITY'] = 'Helicopter Landing Tours'
            iceDF['Activity2'] = iceDF['ACTIVITY'].map(activities)
            iceDF['USELOCATION'].str.upper()
            iceLoc = iceDF.merge(locDF, how='left', left_on = ['LOCATION_ID'] , right_on = ['LOCATION_ID'])
            iceLoc['Year'] = iceLoc['REPORTYEAR']
            iceLoc['LandTourTotal'] = iceLoc['CLIENTSGLACTREK'] + iceLoc['CLIENTSHIKE']
            iceLoc['LandTourTotal'].fillna(0, axis=0, inplace=True)
            iceLoc['DogSledTotal'] = iceLoc['CLIENTSDOGSLED']
            iceLoc['DogSledTotal'].fillna(0, axis=0, inplace=True)
            iceLoc['Activity Client Number'] = iceLoc['LandTourTotal']  + iceLoc['DogSledTotal']
            iceLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'TOTCLIENTSALLZONES':'Total Clients'}, inplace=True)

            
            ice = iceLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy()           
          
            
        if 'guideDF' not in locals():
            guide = pd.DataFrame()
            arcpy.AddMessage('There are no guided recreation activities for this time period.')   
        else:
            if (guideDF['USELOCATION'].loc() == guideDF['USELOCATION'].loc()):
                guideDF['USELOCATION'] = guideDF.loc(guideDF['USELOCATION'].map(lambda x: x.title()))
            guideDF['Activity2'] = guideDF['ACTIVITY'].map(activities)
            guideDF['Year'] = guideDF['TRIPDATE'].dt.year

            
            guideDF['USELOCATION'].str.upper()
            guideLoc = guideDF.merge(locDF, how = 'left', on = ['LOCATION_ID'])

            guideLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'MAXCLIENTS':'Total Clients', 'CLIENTNUMBER':'Activity Client Number'}, inplace=True)
            guide = guideLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy() 
            
            
        if 'huntDF' not in locals():
            if 'nonHuntDF' not in locals():
                hunt = pd.DataFrame()
                arcpy.AddMessage('There are no hunting associated activities for this time period.')   
        else:
            huntDF['Activity2'] = huntDF['ACTIVITY'].map(activities)
            huntDF['Year'] = huntDF['TRIPDATE'].dt.year           
            huntDF['USELOCATION'].str.upper()

            huntLoc = huntDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID']) 
            huntLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'MAXCLIENTS':'Total Clients', 'CLIENTNUMBER':'Activity Client Number'}, inplace=True)
            hunt = huntLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy()
            
            
        if 'heliDF' not in locals():
            heli = pd.DataFrame()
            arcpy.AddMessage('There are no heliski activities for this time period.') 
        else:        
            heliDF['Activity2'] = heliDF['ACTIVITY'].map(activities)
            heliDF['Year'] = heliDF['TRIPDATE'].dt.year
            heliDF['USELOCATION'].str.upper()
            heliLoc = heliDF.merge(locDF, how = 'left', left_on = ['LOCATION_ID'], right_on = ['LOCATION_ID'])
            heliLoc.rename(columns = {'USELOCATION_x':'USELOCATION', 'TOTALCLIENTSONDAY':'Total Clients', 'CLIENTS_LOCATION':'Activity Client Number'}, inplace=True)
            heli = heliLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy()

        if 'mendDF' not in locals():
            mend = pd.DataFrame()
            arcpy.AddMessage('There are no mendenhall activities for this time period.')      
        else:            
            mendDF.loc[mendDF['BIKING']>0, 'ACTIVITY'] = 'Nature Tour, Bicycle'
            mendDF.loc[mendDF['HIKING']>0, 'ACTIVITY'] = 'Nature Tour, on Foot'
            mendDF.loc[mendDF['CANOEING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['KAYAKING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['RAFTING']>0, 'ACTIVITY'] = 'Boating (Non-Motorized, Freshwater)'
            mendDF.loc[mendDF['VCTRANSPORT']>0, 'ACTIVITY'] = 'Visitor Center (Begich Boggs, MGVC, SEADC)'
            
            mendDF.rename(columns = {'USELOCATION':'USELOCATION'}, inplace=True)
            mendDF['Activity2'] = mendDF['ACTIVITY'].map(activities)
            mendDF['USELOCATION'].str.upper()
            mendLoc = mendDF.merge(locDF, how='left', left_on= ['LOCATION_ID'], right_on= ['LOCATION_ID'])
            mendLoc.rename(columns={'USELOCATION_x':'USELOCATION', 'CLIENTMONTH':'Total Clients', 'CLIENTSLOCATION':'Activity Client Number'}, inplace=True)
            mendLoc['Year'] = mendLoc['REPORTYEAR']  
            mend = mendLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy()
            
            
        if (outfitDF.empty == True):
            outfit = pd.DataFrame()
            arcpy.AddMessage('There are no outfitting activities for this time period.')  
        else:
            outfitDF['Activity2'] = outfitDF['ACTIVITY'].map(activities)
            outfitDF['Year'] = outfitDF['TRIPDATE'].dt.year
            outfitDF['USELOCATION'].str.upper()
            outfitLoc = outfitDF.merge(locDF, how = 'left', on = ['LOCATION_ID']) 
            outfitLoc.rename(columns = {'TOTALCLIENTSONDAY':'Total Clients'}, inplace=True) 
            outfitLoc['Activity Client Number'] = outfitLoc.loc[:, 'Total Clients']
            outfitLoc.rename(columns={'USELOCATION_x':'USELOCATION'}, inplace=True)
            outfit = outfitLoc[['Activity2', 'Year', 'FORESTNAME', 'DISTRICTNAME', 'Total Clients', 'USELOCATION', 'TRIP_GUID', 'Activity Client Number']].copy()
            
            
            allDF = ice.append([hunt, heli, mend, outfit, guide])
#            finalDF = allDF.loc[allDF['DISTRICTNAME'].isin([]])]
            
            
            allDF.rename(columns={'FORESTNAME':'Forest', 'Activity2':'Activity', 'USELOCATION':'Use Location'}, inplace=True)
            allDF['Forest'] = allDF['Forest'].astype(str) + " National Forest"
                                
            forestGroup = allDF.groupby(['TRIP_GUID', 'Year'], as_index = False).agg({'Forest' : 'first', 'Total Clients': 'max'})            
            
            forestPivot = pd.pivot_table(forestGroup, index = ['Forest'], columns = ['Year'], values = ['Total Clients'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')            
            
            finalDF = allDF.query('DISTRICTNAME in @rangerDist')
            
            finalDF.rename(columns={'DISTRICTNAME': 'Ranger District'}, inplace=True)
            
            rangerGroup = finalDF.groupby(['TRIP_GUID', 'Year'], as_index = False).agg({'Ranger District' : 'first', 'Total Clients': 'max'})  
            rangerPivot = pd.pivot_table(rangerGroup, index = ['Ranger District'], columns = ['Year'], values = ['Total Clients'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')

            locGroup = finalDF.groupby(['TRIP_GUID', 'Use Location'], as_index = False).agg({'Year': 'first', 'Ranger District' : 'first', 'Total Clients': 'max', 'Activity Client Number': 'sum'})
            locGroup.loc[locGroup['Activity Client Number'] > locGroup['Total Clients'], 'Clients By Location'] = locGroup['Total Clients']
            locGroup['Clients By Location'].fillna(locGroup['Activity Client Number'], axis = 0, inplace = True)     
            
            locPivot = pd.pivot_table(locGroup, index = ['Ranger District', 'Use Location'], columns = ['Year'], values = ['Clients By Location'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')


            actGroup = finalDF.groupby(['TRIP_GUID', 'Activity'], as_index = False).agg({'Year': 'first', 'Ranger District' : 'first', 'Total Clients': 'max', 'Activity Client Number': 'sum'})
            actGroup.loc[actGroup['Activity Client Number'] > actGroup['Total Clients'], 'Clients By Activity'] = actGroup['Total Clients']
            actGroup['Clients By Activity'].fillna(actGroup['Activity Client Number'], axis = 0, inplace = True)   
            
            actPivot = pd.pivot_table(actGroup, index = ['Ranger District', 'Activity'], columns = ['Year'], values = ['Clients By Activity'], aggfunc=np.sum, margins = False, dropna = True, margins_name = 'Total')         
            
            
        """ Writing pivot tables to excel and additional formatting. The activity tab and hunt tab have to be added with conditional statements because not all businesses will have hunts, and some
        businesses that have hunts do not have other activities.
        """

        writer_args = {
            'path': self.path,
            'mode': 'a',
            'engine': 'openpyxl'}
        
        thin = Side(border_style='thin', color='000000')        
        
        index_style = NamedStyle('IndexStyle')
        index_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        index_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        index_style.alignment = Alignment(horizontal='left', wrapText = True, vertical='top')
        index_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        title_style = NamedStyle('TitleStyle')
        title_style.font = Font(name='Calibri', size=11, bold=True, color='000000')
        title_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_style.alignment = Alignment(horizontal='center', wrapText = True, vertical='center')
        title_style.fill = PatternFill('solid', fgColor='D9D9D9')
        
        value_style = NamedStyle('ValueStyle')
        value_style.font = Font(name='Calibri', size=11, bold=False, color='000000')
        value_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        value_style.alignment = Alignment(horizontal='center', wrapText = False, vertical='center')
        
        fill_style = NamedStyle('FillStyle')
        fill_style.fill = PatternFill('solid', fgColor='00C0C0C0')
        
        def update(data, startNumb, columnNumb, ws):
            for row, text in enumerate(data, start=startNumb):
                ws.cell(column=columnNumb, row=row, value=text)
                
        def headers(ws):
            header1 = 'Visitation Summary Report'
            ws.oddHeader.left.text = header1
            ws.oddHeader.left.size = 14
            ws.oddHeader.left.font = "Calibri, bold"
            ws.oddHeader.left.color = "000000"
            
            footer = 'Report Generated: {}\nSource: R10 Outfitter/Guide Database'.format(datetime.today().strftime('%m/%d/%Y'))
            ws.oddFooter.left.text = footer
            
        def excelRow(ws, rows, styles):
            for row in ws[rows]:
                for cell in row:
                    cell.style = styles
        
        def excelStyle(ws, cells, styles):
            for cell in ws[cells]:
                cell.style = styles

        def excelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 3
            index_column = 'A'
            index_column1 = 'B'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'C4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            ws.column_dimensions[index_column1].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, index_column1, index_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)
            
        def singleExcelUpdate(ws):
            title_row = '1'
            title_row1 = '2'
            title_row2 = '3'
            value_col = 2
            index_column = 'A'
            col=get_column_letter(ws.max_column)
            colInt= (ws.max_column) 
            rows=ws.max_row                               
            value_cells = 'B4:{}{}'.format(col, rows)   
            ws.column_dimensions[index_column].width = 35
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')
            ws.sheet_view.view = 'pageLayout'
            while value_col <= colInt:
                i = get_column_letter(value_col)
                ws.column_dimensions[i].width = 8
                value_col += 1 
            excelRow(ws, value_cells, value_style)
            excelStyle(ws, index_column, index_style)
            excelStyle(ws, title_row, title_style)
            excelStyle(ws, title_row1, title_style)
            excelStyle(ws, title_row2, title_style)
            headers(ws)            
        
    
        with pd.ExcelWriter(**writer_args) as xlsx:
            forestPivot.to_excel(xlsx, 'Visitation by Forest')
            ws = xlsx.sheets['Visitation by Forest'] 
            singleExcelUpdate(ws)
            rangerPivot.to_excel(xlsx, 'Visitation by District')
            wsDist = xlsx.sheets['Visitation by District'] 
            singleExcelUpdate(wsDist)
            wsDist.delete_rows(3)
            
            locPivot.to_excel(xlsx, 'Visitation by Location')
            wsLoc = xlsx.sheets['Visitation by Location'] 
            excelUpdate(wsLoc)
            wsLoc.oddFooter.center.text = "Page &P of &N"
            wsLoc.evenFooter.center.text = "Page &P of &N"
            actPivot.to_excel(xlsx, 'Visitation by Activity')
            wsAct = xlsx.sheets['Visitation by Activity'] 
            excelUpdate(wsAct)
            wsAct.oddFooter.center.text = "Page &P of &N"
            wsAct.evenFooter.center.text = "Page &P of &N"
            
        return      
